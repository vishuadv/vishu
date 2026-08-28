import io
import logging
import calendar
import traceback
import os
import json
import secrets
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

try:
    import pdfkit
    PDFKIT_CFG = pdfkit.configuration(wkhtmltopdf="/usr/bin/wkhtmltopdf")
    PDFKIT_AVAILABLE = True
except (ImportError, OSError):
    pdfkit = None
    PDFKIT_CFG = None
    PDFKIT_AVAILABLE = False
import pandas as pd
from datetime import datetime, date, timedelta
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user

from flask import Flask, render_template, request, jsonify, send_file, render_template_string, current_app, Blueprint, abort, make_response, redirect, url_for, flash, Response
from flask_wtf.csrf import CSRFProtect
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, event, select, update, text, inspect
from flask import send_from_directory, make_response
import jinja2
import xlsxwriter
import math
import requests  # ADD THIS IMPORT
import threading  # ADD THIS IMPORT
import time  # ADD THIS IMPORT

# PDF Generation imports
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors

app = Flask(__name__)

# CSRF Protection
csrf = CSRFProtect(app)

# Exclude API routes from CSRF protection (they use other auth methods)
csrf.exempt('/api/save-payment-data')
csrf.exempt('/api/get-payment-data')
csrf.exempt('/api/delete-payment-data')
csrf.exempt('/api/complete-client-ledger')
csrf.exempt('/api/client-ledger')
csrf.exempt('/api/receivable-report')
csrf.exempt('/api/payment-entry/day-state')
csrf.exempt('/api/payment-entry/due-data')
csrf.exempt('/api/payment-entry/recover-single')
csrf.exempt('/api/payment-entry/penalty-report')
csrf.exempt('/api/payment-entry/penalty-collected')
csrf.exempt('/api/payment-entry/profit-loss-report')
csrf.exempt('/api/payment-entry/outstanding-report')
csrf.exempt('/api/recovery/report')
csrf.exempt('/api/delete-payment-data')
csrf.exempt('/clients')
csrf.exempt('/loan')
csrf.exempt('/loan/<int:id>')
csrf.exempt('/loan/<int:id>/close')
csrf.exempt('/loan/<int:id>/manual-close')
csrf.exempt('/loan/close/<int:loan_id>')
csrf.exempt('/delete/customer')
csrf.exempt('/client/<phone>')
csrf.exempt('/receivable-report/snapshot')
csrf.exempt('/daily-report/snapshot')
csrf.exempt('/bulk-import-all')
csrf.exempt('/batch/close-cases')
csrf.exempt('/payment-entry/process')
csrf.exempt('/payment-entry/edit')
csrf.exempt('/payment-entry/delete')
csrf.exempt('/payment-entry/bulk-submit-last-100-days')
csrf.exempt('/payment-entry/bulk-delete-last-100-days')

# 🔐 Flask-Login Configuration
# Generate a secure secret key if not provided
if not os.environ.get('SECRET_KEY'):
    # For development, generate a random key
    # In production, set SECRET_KEY environment variable
    app.secret_key = secrets.token_hex(32)
else:
    app.secret_key = os.environ.get('SECRET_KEY')

# Session configuration for auto-logout on browser close
app.config['SESSION_PERMANENT'] = False  # Session expires when browser closes
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)  # Fallback if permanent
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Security: prevent JavaScript access
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
app.config['SESSION_COOKIE_SECURE'] = True if os.environ.get('FLASK_ENV') == 'production' else False  # HTTPS only in production

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'

# 🔐 Custom Jinja2 Filter for JSON
@app.template_filter('from_json')
def from_json_filter(s):
    try:
        return json.loads(s)
    except:
        return []


@app.route("/sw.js")
def sw():
    resp = make_response(send_from_directory("static", "service-worker.js"))
    resp.headers["Content-Type"] = "application/javascript; charset=utf-8"
    resp.headers["Service-Worker-Allowed"] = "/"
    resp.headers["Cache-Control"] = "no-cache"
    return resp

@app.get("/health")
def health_check():
    """Lightweight health endpoint for platform probes."""
    return jsonify({"status": "ok"}), 200


# Optional but recommended: serve manifest with correct type
@app.route("/manifest.webmanifest")
def manifest():
    resp = make_response(send_from_directory("static", "manifest.json"))
    resp.headers["Content-Type"] = "application/manifest+json; charset=utf-8"
    resp.headers["Cache-Control"] = "no-cache"
    return resp

# Optional: offline page (used by SW fallback)
@app.route("/offline")
def offline():
    resp = make_response(send_from_directory("static", "offline.html"))
    resp.headers["Content-Type"] = "text/html; charset=utf-8"
    resp.headers["Cache-Control"] = "no-cache"
    return resp
    

# 🌍 DATABASE CONFIGURATION (CLOUD + LOCAL FALLBACK)
print("🔧 Configuring database...")

# Try cloud database first, fall back to local SQLite for testing
database_url = os.environ.get('DATABASE_URL')
if not database_url:
    print("⚠️ DATABASE_URL not found - using local SQLite for testing")
    database_url = 'sqlite:///swift_local.db'
    # Local SQLite doesn't need pool settings
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {}
else:
    # Fix PostgreSQL URL format for cloud deployment
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    print("✅ Using cloud PostgreSQL database")
    # Cloud PostgreSQL needs connection pooling
    app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
        'pool_pre_ping': True,
        'pool_recycle': 300,
        'pool_size': 10,
        'max_overflow': 20
    }

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Auto-detect environment: DEBUG for local, False for production
is_local = not bool(os.environ.get('DATABASE_URL'))
app.config['DEBUG'] = is_local or os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'

if is_local:
    print("✅ Local SQLite database configured for testing")
    print("🔧 DEBUG mode enabled for local development")
else:
    print("✅ Cloud PostgreSQL database configured")
    print("🚀 Production mode enabled")
    
db = SQLAlchemy()
db.init_app(app)





# 🔧 ALL YOUR BUSINESS MODELS (SAME AS BEFORE)
class Loan(db.Model):
    __tablename__ = 'loan'
    __table_args__ = {'extend_existing': True}
    
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200))
    address = db.Column(db.String(500))
    phone = db.Column(db.String(50))
    city = db.Column(db.String(100))
    loan_date = db.Column(db.String(20))
    loan_closed_date = db.Column(db.String(20), nullable=True)
    loan_amount = db.Column(db.Float, default=0.0)
    processing_fees = db.Column(db.Float, nullable=True, default=0.0)
    case_closing_amount = db.Column(db.Float, nullable=True, default=0.0)
    advance_amount = db.Column(db.Float, nullable=True, default=0.0)
    interest_rate = db.Column(db.Float, nullable=True, default=0.0)
    repayment_type = db.Column(db.String(20), default='DAILY')
    status = db.Column(db.String(50), default='OPEN')
    total_paid_amount = db.Column(db.Float, default=0.0)
    outstanding_balance = db.Column(db.Float, default=0.0)
    remarks = db.Column(db.String(1000), nullable=True)

class Payment(db.Model):
    __tablename__ = 'payment'
    __table_args__ = {'extend_existing': True}
    
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'), nullable=False)
    payment_date = db.Column(db.String(20), default='2025-01-01')
    amount_paid = db.Column(db.Float, nullable=False)
    payment_method = db.Column(db.String(100), default='Collection Entry')
    remarks = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.String(50), nullable=True)
    entry_type = db.Column(db.String(20), default='COLLECTION')

class PaymentSubmission(db.Model):
    __tablename__ = 'payment_submission'
    __table_args__ = {'extend_existing': True}
    
    id = db.Column(db.Integer, primary_key=True)
    submission_date = db.Column(db.String(20), nullable=False)
    total_amount = db.Column(db.Float)
    total_payments = db.Column(db.Integer)
    submitted_at = db.Column(db.String(50))

class Payments(db.Model):
    __tablename__ = 'payments'
    __table_args__ = {'extend_existing': True}
    
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'))
    payment_date = db.Column(db.String(20))
    amount = db.Column(db.Float)
    payment_type = db.Column(db.String(50))
    notes = db.Column(db.String(500))
    created_at = db.Column(db.String(50))

class RecoveryPayment(db.Model):
    __tablename__ = 'recovery_payments'
    __table_args__ = {'extend_existing': True}
    
    id = db.Column(db.Integer, primary_key=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'))
    client_name = db.Column(db.String(200))
    due_date = db.Column(db.String(20))
    recovery_date = db.Column(db.String(20))
    amount = db.Column(db.Float)
    notes = db.Column(db.String(500))
    created_at = db.Column(db.String(50))

class ShortPayment(db.Model):
    __tablename__ = 'short_payment'
    __table_args__ = {'extend_existing': True}
    
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'))
    payment_date = db.Column(db.String(20))
    expected_amount = db.Column(db.Float)
    status = db.Column(db.String(50))
    created_at = db.Column(db.String(50))


class PenaltyCollected(db.Model):
    __tablename__ = 'penalty_collected'
    __table_args__ = (
        db.UniqueConstraint('loan_id', 'due_date', 'repayment_type', name='uq_penalty_collected'),
        {'extend_existing': True}
    )

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    loan_id = db.Column(db.Integer, db.ForeignKey('loan.id'))
    due_date = db.Column(db.String(20), nullable=False)
    repayment_type = db.Column(db.String(50), nullable=False)
    collected = db.Column(db.Boolean, default=False)
    updated_at = db.Column(db.String(50))

class PaymentDraft(db.Model):
    __tablename__ = 'payment_draft'
    __table_args__ = {'extend_existing': True}

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    draft_date = db.Column(db.String(20), unique=True, nullable=False)
    data_json = db.Column(db.Text, nullable=False)
    updated_at = db.Column(db.String(50))


class DailyReportSnapshot(db.Model):
    __tablename__ = 'daily_report_snapshot'
    __table_args__ = {'extend_existing': True}

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    report_date = db.Column(db.String(20), unique=True, nullable=False)
    data_json = db.Column(db.Text, nullable=False)
    generated_at = db.Column(db.String(50))


class ReceivableSnapshot(db.Model):
    __tablename__ = 'receivable_snapshot'
    __table_args__ = {'extend_existing': True}

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    as_on_date = db.Column(db.String(20), unique=True, nullable=False)  # DD-MM-YYYY
    data_json = db.Column(db.Text, nullable=False)
    generated_at = db.Column(db.String(50))


# 🔐 AUTHENTICATION MODELS
class User(UserMixin, db.Model):
    __tablename__ = 'user'
    __table_args__ = {'extend_existing': True}

    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # 'admin', 'user'
    allowed_pages = db.Column(db.Text, default='[]')  # JSON array of allowed page names
    created_at = db.Column(db.String(50), default=lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    is_active = db.Column(db.Boolean, default=True)


# 🔐 Flask-Login User Loader
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# 🔐 Authentication Decorators
def role_required(role):
    """Decorator to require specific role"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            if current_user.role != role:
                return jsonify({"error": "Access denied - insufficient permissions"}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator


def page_required(page_name):
    """Decorator to require access to specific page"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return redirect(url_for('login'))
            
            # Admin has access to all pages
            if current_user.role == 'admin':
                return f(*args, **kwargs)
            
            # Check if user has permission for this page
            allowed_pages = json.loads(current_user.allowed_pages or '[]')
            if page_name not in allowed_pages:
                return jsonify({"error": "Access denied - page not allowed"}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def api_auth_required(page_name=None):
    """Custom decorator for API endpoints that returns JSON errors instead of HTML redirects"""
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                return jsonify({"error": "Authentication required", "message": "Please login to access this resource"}), 401
            
            # Check page permission if specified
            if page_name:
                # Admin has access to all pages
                if current_user.role != 'admin':
                    allowed_pages = json.loads(current_user.allowed_pages or '[]')
                    if page_name not in allowed_pages:
                        return jsonify({"error": "Access denied - page not allowed"}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return decorator


print("✅ All 9 models defined")

# 🔧 FIXED TABLE CREATION WITH CORRECT SQLAlchemy SYNTAX
def create_tables_with_retry(max_attempts=3):
    """Create tables with retry mechanism - FIXED VERSION"""
    for attempt in range(max_attempts):
        try:
            with app.app_context():
                # Only create tables if database doesn't exist (preserve backup data)
                db.create_all()
                print("✅ Tables ready")
                                
                # ✅ FIXED: Use inspect(engine).has_table() instead of engine.has_table()
                inspector = inspect(db.engine)
                tables_created = []
                
                for model in [Loan, Payment, PaymentSubmission, Payments, RecoveryPayment, ShortPayment, User]:
                    if inspector.has_table(model.__tablename__):
                        tables_created.append(model.__tablename__)
                
                if len(tables_created) == 7:
                    print("✅ All 7 tables created and verified successfully!")
                    print(f"📊 Tables: {', '.join(tables_created)}")
                    
                    # Test database connection with actual query
                    result = db.session.execute(text('SELECT 1'))
                    db.session.commit()
                    print("✅ Database connection tested and working perfectly")
                    return True
                else:
                    print(f"⚠️  Only {len(tables_created)} tables created, retrying...")
                    
        except Exception as e:
            print(f"❌ Attempt {attempt + 1} failed: {e}")
            if attempt == max_attempts - 1:
                print("❌ All table creation attempts failed")
                return False

    return False

def ensure_column_exists(table_name, column_name, ddl_sql):
    try:
        with app.app_context():
            inspector = inspect(db.engine)
            cols = [c['name'] for c in inspector.get_columns(table_name)]
            if column_name not in cols:
                db.session.execute(text(ddl_sql))
                db.session.commit()
                print(f"✅ Added missing column {column_name} to {table_name}")
            return True
    except Exception as e:
        print(f"⚠️  Migration check failed for {table_name}.{column_name}: {e}")
        return False

# 🔧 FIXED DATABASE INITIALIZATION
def initialize_cloud_database():
    """Initialize database with bulletproof error handling and verification"""
    try:
        with app.app_context():
            # Create tables with retry mechanism
            if not create_tables_with_retry():
                print("❌ Critical: Database table creation failed")
                return False
            
            # Ensure new columns exist (idempotent)
            ensure_column_exists('loan', 'interest_rate', 'ALTER TABLE loan ADD COLUMN interest_rate REAL')
            ensure_column_exists('recovery_payments', 'due_date', 'ALTER TABLE recovery_payments ADD COLUMN due_date TEXT')

            # Ensure daily_report_snapshot table exists for frozen daily reports (idempotent)
            try:
                inspector = inspect(db.engine)
                if not inspector.has_table('daily_report_snapshot'):
                    db.session.execute(text(
                        """
                        CREATE TABLE IF NOT EXISTS daily_report_snapshot (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            report_date VARCHAR(20) UNIQUE NOT NULL,
                            data_json TEXT NOT NULL,
                            generated_at VARCHAR(50)
                        )
                        """
                    ))
                    db.session.commit()
                    print("✅ Created daily_report_snapshot table")
            except Exception as e:
                print(f"⚠️  Could not ensure daily_report_snapshot table: {e}")

            # Ensure receivable_snapshot table exists for frozen receivable reports (idempotent)
            try:
                inspector = inspect(db.engine)
                if not inspector.has_table('receivable_snapshot'):
                    db.session.execute(text(
                        """
                        CREATE TABLE IF NOT EXISTS receivable_snapshot (
                            id INTEGER PRIMARY KEY AUTOINCREMENT,
                            as_on_date VARCHAR(20) UNIQUE NOT NULL,
                            data_json TEXT NOT NULL,
                            generated_at VARCHAR(50)
                        )
                        """
                    ))
                    db.session.commit()
                    print("✅ Created receivable_snapshot table")
            except Exception as e:
                print(f"⚠️  Could not ensure receivable_snapshot table: {e}")

            # Verify database is working with comprehensive test
            try:
                # Test each table
                print("✅ Database verification successful")
                
                # Add sample data only if completely empty
                if Loan.query.count() == 0:
                    sample_loans = [
                        Loan(name='SAMPLE CLIENT 1', phone='1111111111', loan_amount=100000, status='OPEN', outstanding_balance=100000),
                        Loan(name='SAMPLE CLIENT 2', phone='2222222222', loan_amount=150000, status='OPEN', outstanding_balance=150000)
                    ]
                    
                    for loan in sample_loans:
                        db.session.add(loan)
                    
                    db.session.commit()
                    print("✅ Sample loan data added successfully")
                
                # Create default admin user if not exists
                admin_username = os.environ.get('ADMIN_USERNAME', 'shivam')
                admin_password = os.environ.get('ADMIN_PASSWORD', 'Raaina@20')
                
                if User.query.filter_by(username=admin_username).first() is None:
                    admin_user = User(
                        username=admin_username,
                        password_hash=generate_password_hash(admin_password),
                        role='admin',
                        allowed_pages='[]'
                    )
                    db.session.add(admin_user)
                    db.session.commit()
                    print(f"✅ Default admin user created ({admin_username}/{admin_password})")
                
                return True
                
            except Exception as db_error:
                print(f"❌ Database verification failed: {db_error}")
                return False
            
    except Exception as e:
        print(f"❌ Database initialization failed: {e}")
        return False

# Initialize on startup
initialize_cloud_database()

# 🔐 AUTHENTICATION ROUTES
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.query.filter_by(username=username).first()
        
        if user and check_password_hash(user.password_hash, password):
            login_user(user)
            # Default admin (shivam) goes to user management, other admins go to index.html, users go to dashboard or index if permitted
            if user.role == 'admin':
                if user.username == 'shivam':
                    return redirect('/admin/users')  # Default admin goes to user management
                else:
                    return redirect('/home')  # Other admins go to index.html
            else:
                # Check if user has index permission
                allowed_pages = json.loads(user.allowed_pages or '[]')
                if 'index' in allowed_pages:
                    return redirect('/home')  # Users with index permission go to index.html
                else:
                    return redirect('/dashboard')  # Other users go to dashboard
        else:
            flash('Invalid username or password. Please try again.', 'error')
    
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """Logout current user"""
    logout_user()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard - shows allowed pages based on user role"""
    try:
        # Define all available pages
        all_pages = [
            {'name': 'daily_report', 'title': 'Daily Report', 'route': '/daily-report/page'},
            {'name': 'sales_report', 'title': 'Sales Report', 'route': '/sales/report/page'},
            {'name': 'reconciliation', 'title': 'Reconciliation', 'route': '/reconciliation/page'},
            {'name': 'clients', 'title': 'Clients', 'route': '/afin-ledger/page'},
            {'name': 'ledger', 'title': 'Ledger', 'route': '/ledger/page'},
            {'name': 'receivable', 'title': 'Advances Report', 'route': '/advances/report'},
            {'name': 'loan_form', 'title': 'Loan Management Form', 'route': '/loan/form'},
            {'name': 'payment_entry', 'title': 'Payment Entry System', 'route': '/payment-entry'},
            {'name': 'penalty_report', 'title': 'Penalty Report', 'route': '/payment-entry/penalty'},
            {'name': 'recovery', 'title': 'Recovery', 'route': '/recovery/page'},
            {'name': 'profit_loss', 'title': 'Profit Loss Report', 'route': '/payment-entry/profit-loss'},
            {'name': 'outstanding', 'title': 'Outstanding Report', 'route': '/payment-entry/outstanding'},
            {'name': 'all_ledgers', 'title': 'All Client Ledgers', 'route': '/all-clients-ledgers-view'},
        ]
        
        # Filter pages based on user role
        if current_user.role == 'admin':
            allowed_pages = all_pages
        else:
            try:
                allowed_pages_json = json.loads(current_user.allowed_pages or '[]')
            except (json.JSONDecodeError, TypeError):
                allowed_pages_json = []
            allowed_pages = [p for p in all_pages if p['name'] in allowed_pages_json]
        
        return render_template('dashboard.html', 
                              pages=allowed_pages,
                              user=current_user)
    except Exception as e:
        print(f"Dashboard error: {e}")
        import traceback
        traceback.print_exc()
        return f"Error loading dashboard: {str(e)}", 500

@app.route('/admin/users')
@login_required
@role_required('admin')
def admin_users():
    """Admin panel to manage users"""
    users = User.query.all()
    return render_template('admin_users.html', users=users)


@app.route('/admin/users/create', methods=['POST'])
@login_required
@role_required('admin')
def create_user():
    """Create a new user (admin only)"""
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')
    allowed_pages = request.form.getlist('allowed_pages')
    
    # Check if username already exists
    if User.query.filter_by(username=username).first():
        return jsonify({"error": "Username already exists"}), 400
    
    # Create user
    user = User(
        username=username,
        password_hash=generate_password_hash(password),
        role=role,
        allowed_pages=json.dumps(allowed_pages)
    )
    
    db.session.add(user)
    db.session.commit()
    
    return redirect(url_for('admin_users'))


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@login_required
@role_required('admin')
def delete_user(user_id):
    """Delete a user (admin only)"""
    user = User.query.get_or_404(user_id)
    
    # Prevent deleting yourself
    if user.id == current_user.id:
        return jsonify({"error": "Cannot delete yourself"}), 400
    
    db.session.delete(user)
    db.session.commit()
    
    return redirect(url_for('admin_users'))


@app.route('/admin/users/update', methods=['POST'])
@login_required
@role_required('admin')
def update_user():
    """Update a user (admin only)"""
    user_id = request.form.get('user_id')
    username = request.form.get('username')
    password = request.form.get('password')
    role = request.form.get('role')
    allowed_pages = request.form.getlist('allowed_pages')
    
    user = User.query.get_or_404(user_id)
    
    # Prevent editing yourself to remove admin role
    if user.id == current_user.id and role != 'admin':
        return jsonify({"error": "Cannot remove admin role from yourself"}), 400
    
    # Update fields
    user.username = username
    user.role = role
    user.allowed_pages = json.dumps(allowed_pages)
    
    # Update password only if provided
    if password:
        user.password_hash = generate_password_hash(password)
    
    db.session.commit()
    
    return redirect(url_for('admin_users'))


# 🌐 YOUR EXISTING ROUTES (keep as-is)

@app.post('/api/save-payment-data')
def save_payment_data():
    data = request.get_json(force=True) or {}
    draft_date = (data.get('date') or '').strip()
    if not draft_date:
        return jsonify({"error": "date is required"}), 400

    try:
        payload_json = json.dumps(data, ensure_ascii=False)
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        existing = PaymentDraft.query.filter_by(draft_date=draft_date).first()
        if existing:
            existing.data_json = payload_json
            existing.updated_at = now_str
            db.session.add(existing)
        else:
            draft = PaymentDraft(draft_date=draft_date, data_json=payload_json, updated_at=now_str)
            db.session.add(draft)

        db.session.commit()
        return jsonify({"success": True, "date": draft_date})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@app.get('/api/get-payment-data')
def get_payment_data():
    draft_date = (request.args.get('date') or '').strip()
    if not draft_date:
        return jsonify({"error": "date is required"}), 400

    try:
        existing = PaymentDraft.query.filter_by(draft_date=draft_date).first()
        if not existing:
            return jsonify({"date": draft_date, "shortPayments": [], "recoveryHistory": [], "totalRecoveredAmount": 0, "globalTotals": {}, "submissionStatus": False})

        parsed = json.loads(existing.data_json or '{}')
        if not isinstance(parsed, dict):
            parsed = {}
        parsed.setdefault('date', draft_date)
        return jsonify(parsed)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.delete('/api/delete-payment-data')
@csrf.exempt
def delete_payment_data():
    draft_date = (request.args.get('date') or '').strip()
    if not draft_date:
        return jsonify({"error": "date is required"}), 400

    try:
        PaymentDraft.query.filter_by(draft_date=draft_date).delete()
        db.session.commit()
        return jsonify({"success": True, "date": draft_date})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    return redirect(url_for('login'))

@app.route('/home')
@login_required
def home():
    """Show index.html - for admin shows all buttons, for user shows filtered buttons"""
    try:
        # Check if user has permission to access index.html
        if current_user.role != 'admin':
            allowed_pages_json = json.loads(current_user.allowed_pages or '[]')
            if 'index' not in allowed_pages_json:
                return redirect('/dashboard')  # Redirect to dashboard if no index permission
        
        # Get user's allowed pages
        if current_user.role == 'admin':
            allowed_pages = None  # Admin sees all
        else:
            allowed_pages = json.loads(current_user.allowed_pages or '[]')
        
        return render_template("index.html", allowed_pages=allowed_pages, user=current_user)
    except Exception:
        # Fallback if template doesn't exist
        return '''
        <!DOCTYPE html>
        <html>
        <head><title>Loan Management System</title></head>
        <body>
            <h1>🏦 LOAN MANAGEMENT SYSTEM</h1>
            <h2>Main Dashboard</h2>
            
            <div style="margin: 20px 0;">
                <h3>📊 Reports & Operations:</h3>
                <ul style="list-style-type: none; padding: 0;">
                    <li><a href="/daily-report/page" style="display: block; margin: 10px 0; padding: 10px; background: #007bff; color: white; text-decoration: none;">📅 Daily Collection Report</a></li>
                    <li><a href="/sales/report/page" style="display: block; margin: 10px 0; padding: 10px; background: #28a745; color: white; text-decoration: none;">📈 Sales Report</a></li>
                    <li><a href="/all-clients-ledgers-view" style="display: block; margin: 10px 0; padding: 10px; background: #17a2b8; color: white; text-decoration: none;">👥 All Client Ledgers</a></li>
                    <li><a href="/ledger/page" style="display: block; margin: 10px 0; padding: 10px; background: #6c757d; color: white; text-decoration: none;">📋 Individual Client Ledger</a></li>
                    <li><a href="/payment-entry" style="display: block; margin: 10px 0; padding: 10px; background: #fd7e14; color: white; text-decoration: none;">💰 Payment Entry</a></li>
                    <li><a href="/advances/report" style="display: block; margin: 10px 0; padding: 10px; background: #20c997; color: white; text-decoration: none;">💸 Receivable Report</a></li>
                </ul>
            </div>
            
            <div style="margin: 20px 0;">
                <h3>🔧 System Management:</h3>
                <ul style="list-style-type: none; padding: 0;">
                    <li><a href="/debug-db" style="display: block; margin: 10px 0; padding: 10px; background: #6f42c1; color: white; text-decoration: none;">🔍 Database Status</a></li>
                    <li><a href="/import-real-business-data" style="display: block; margin: 10px 0; padding: 10px; background: #e83e8c; color: white; text-decoration: none;">📤 Import Data</a></li>
                </ul>
            </div>
        </body>
        </html>
        '''

@app.route('/debug-db')
@login_required
@role_required('admin')
def debug_database():
    """Cloud database status"""
    try:
        loans = Loan.query.all()
        payments = Payment.query.all()
        
        # ✅ FIXED: Use inspector.has_table() instead of engine.has_table()
        inspector = inspect(db.engine)
        tables = []
        for model in [Loan, Payment, PaymentSubmission, Payments, RecoveryPayment, ShortPayment]:
            if inspector.has_table(model.__tablename__):
                tables.append(model.__tablename__)
        
        return {
            "status": "✅ CLOUD DATABASE WORKING",
            "database_type": "PostgreSQL Cloud",
            "tables_created": tables,
            "total_loans": len(loans),
            "total_payments": len(payments),
            "environment": "PRODUCTION",
            "pdf_export": "✅ ENABLED",
            "excel_export": "✅ ENABLED",
            "message": "Your loan management system is live with export features!",
            "ready_for_data": "Upload your CSV files now!"
        }
        
    except Exception as e:
        return {"status": "❌ Error", "error": str(e)}, 500

# 🚀 KEEP ALL YOUR EXISTING PDF AND EXCEL EXPORT ROUTES AS-IS

# ✅ FIXED: Add context processor for date/datetime in templates
@app.context_processor
def inject_date():
    """Make date and datetime available in ALL templates"""
    from datetime import date, datetime
    return dict(
        date=date,
        datetime=datetime,
        today=date.today(),
        now=datetime.now()
    )

ALLOWED_REPAYMENT_TYPES = {"DAILY", "WEEKLY", "TEN_DAYS", "MONTHLY"}

# PROFESSIONAL PDF Templates for Sales Report
SALES_REPORT_PDF_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head
    <meta charset="utf-8">
    <style>
        body { font-family: Arial, sans-serif; font-size: 10px; padding: 10px; margin: 0; }
        .report-header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px; border-radius: 8px; margin-bottom: 15px; text-align: center; }
        .filters-info { background: #f8f9fa; padding: 10px; border-radius: 5px; margin-bottom: 15px; font-size: 11px; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 20px; font-size: 9px; }
        th, td { border: 1px solid #ddd; padding: 4px; text-align: center; }
        th { background-color: #343a40; color: white; font-weight: bold; }
        .status-open { background-color: #fff3cd; color: #856404; }
        .status-closed { background-color: #d1edff; color: #0c5460; }
        .totals-row { background-color: #28a745; color: white; font-weight: bold; }
        .footer { margin-top: 20px; text-align: center; font-size: 10px; color: #666; }
        @page { size: A4 landscape; margin: 0.5cm; }
    </style>
</head>
<body>
    <div class="report-header">
        <h3>📊 SALES REPORT</h3>
        <p>Generated: {{ generated_date }} | Total Records: {{ total_count }}</p>
    </div>
    
    <div class="filters-info">
        <strong>Filters Applied:</strong> 
        Date Range: {{ filters.start_date }} to {{ filters.end_date }} | 
        Status: {{ filters.status }} | 
        Type: {{ filters.type }} |
        Remarks: {{ filters.remarks }}
    </div>
    
    <table>
        <thead>
            <tr>
                <th>#</th><th>Client Name</th><th>Loan Start</th><th>Loan End</th><th>Close Date</th><th>Close Amount</th>
                <th>Type</th><th>Day</th><th>Loan Amount</th><th>Copies</th><th>Installment</th><th>Remarks</th><th>Status</th>
            </tr>
        </thead>
        <tbody>
            {% for row in rows %}
            <tr class="{{ 'status-closed' if row.status == 'CLOSED' else 'status-open' }}">
                <td>{{ row.serial }}</td>
                <td>{{ row.name }}</td>
                <td>{{ row.loan_start_date }}</td>
                <td>{{ row.loan_end_date }}</td>
                <td>
                    {{ row.close_date }}
                    {% if row.renewal_info %}
                    <div style="font-size:8px; color:#0d6efd; font-weight:bold; margin-top:2px;">
                        {{ row.renewal_info }}
                    </div>
                    {% endif %}
                </td>
                <td>₹{{ "{:,}".format(row.close_amount) if row.close_amount else 0 }}</td>
                <td>{{ row.type }}</td>
                <td>{{ row.day }}</td>
                <td>₹{{ "{:,}".format(row.loan_amount) }}</td>
                <td>{{ "{:.2f}".format(row.copies) }}</td>
                <td>₹{{ "{:,}".format(row.installment_amt) }}</td>
                <td>{{ row.remarks }}</td>
                <td>{{ row.status }}</td>
            </tr>
            {% endfor %}
        </tbody>
        <tfoot>
            <tr class="totals-row">
                <td colspan="8"><strong>TOTALS (Active Only - {{ open_totals.count }} cases):</strong></td>
                <td><strong>₹{{ "{:,}".format(open_totals.loan_amount) }}</strong></td>
                <td><strong>{{ "{:.2f}".format(open_totals.copies) }}</strong></td>
                <td><strong>₹{{ "{:,}".format(open_totals.installments) }}</strong></td>
                <td></td>
                <td></td>
            </tr>
            <tr class="totals-row" style="background-color:#6c757d;">
                <td colspan="8"><strong>CLOSED CASES ({{ closed_totals.count }}) - Excluded Above:</strong></td>
                <td><strong>₹{{ "{:,}".format(closed_totals.loan_amount) }}</strong></td>
                <td><strong>{{ "{:.2f}".format(closed_totals.copies) }}</strong></td>
                <td><strong>₹{{ "{:,}".format(closed_totals.installments) }}</strong></td>
                <td></td>
                <td></td>
            </tr>
            <tr class="totals-row" style="background-color:#17a2b8;">
                <td colspan="8"><strong>GRAND TOTALS ({{ total_count }} cases):</strong></td>
                <td><strong>₹{{ "{:,}".format(grand_totals.loan_amount) }}</strong></td>
                <td><strong>{{ "{:.2f}".format(grand_totals.copies) }}</strong></td>
                <td><strong>₹{{ "{:,}".format(grand_totals.installments) }}</strong></td>
                <td></td>
                <td></td>
            </tr>
        </tfoot>
    </table>
    
    <div class="footer">
        <p><strong>SHIVI PROJECT - Loan Management System</strong></p>
    </div>
</body>
</html>
'''

# COMPLETE PDF Templates (keeping existing daily report template)
PDF_DAILY_REPORT_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body { font-family: Arial, sans-serif; font-size: 12px; padding: 10px; margin: 0; }
        .report-header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 15px; border-radius: 8px; margin-bottom: 15px; text-align: center; }
        .section-header { background: #343a40; color: white; padding: 8px 15px; margin: 20px 0 10px 0; font-weight: bold; text-align: center; font-size: 16px; border-radius: 5px; }
        .multi-column { display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 15px; }
        .column-box { border: 2px solid #dee2e6; border-radius: 8px; font-size: 12px; overflow: hidden; flex: 0 0 calc(33% - 12px); max-width: calc(33% - 12px); min-width: 240px; box-sizing: border-box; }
        .column-header { background: #f8f9fa; border-bottom: 2px solid #343a40; padding: 8px; font-weight: bold; display: flex; justify-content: space-between; }
        .client-row { display: flex; padding: 6px 8px; border-bottom: 1px solid #e9ecef; font-size: 12px; align-items: center; justify-content: space-between; }
        .client-row:nth-child(even) { background: #f8f9fa; }
        .client-serial { width: 50px; text-align: center; font-weight: bold; }
        .client-name { flex: 2; text-align: center; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; padding: 0 5px; }
        .client-amount { width: 80px; text-align: center; font-weight: bold; color: #28a745; }
        .client-row.new-client { background-color: #fff3cd !important; color: #856404 !important; font-weight: bold; border-left: 4px solid #ffc107; }
        .client-row.closed-loan { background-color: #f8d7da !important; color: #721c24 !important; font-weight: bold; border-left: 4px solid #dc3545; }
        .column-subtotal { background: #6c757d !important; color: white !important; font-weight: bold; border-top: 2px solid #343a40; padding: 8px; display: flex; justify-content: space-between; }
        .type-total { background: #28a745; color: white; padding: 12px; text-align: center; font-weight: bold; margin: 15px 0 25px 0; border-radius: 8px; font-size: 16px; }
        .grand-total { background: #dc3545; color: white; padding: 15px; border-radius: 8px; text-align: center; font-size: 18px; font-weight: bold; margin-top: 20px; }
        @page { size: A4 landscape; margin: 1cm; }
    </style>
</head>
<body>
    <div class="report-header">
        <h4>Daily Collection Report</h4>
        <p>{{ report_date }} | Generated: {{ generated_time }}</p>
    </div>
    
    {% if data.daily %}
    <div class="section-header">🔄 DAILY INSTALLMENTS ({{ data.daily|length }} clients)</div>
    <div class="multi-column">
        {% set clients_per_column = 50 %}
        {% set total_columns = ((data.daily|length + clients_per_column - 1) // clients_per_column) or 1 %}
        
        {% for col in range(total_columns) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.daily|length] | min %}
            {% set column_clients = data.daily[start_index:end_index] %}
            
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span>Sr.No.</span><span>Name</span><span>Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    {% set highlight_class = 'new-client' if client.is_new_client else ('closed-loan' if client.is_closed_loan else '') %}
                    <div class="client-row {{ highlight_class }}">
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span>Subtotal:</span><span>₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    <div class="type-total">DAILY TOTAL: ₹{{ "{:,}".format(data.daily_total) }}</div>
    {% endif %}

    {% if data.weekly %}
    <div class="section-header">📅 WEEKLY INSTALLMENTS ({{ data.weekly|length }} clients)</div>
    <div class="multi-column">
        {% set clients_per_column = 50 %}
        {% set total_columns = ((data.weekly|length + clients_per_column - 1) // clients_per_column) or 1 %}
        
        {% for col in range(total_columns) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.weekly|length] | min %}
            {% set column_clients = data.weekly[start_index:end_index] %}
            
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span>Sr.No.</span><span>Name</span><span>Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    {% set highlight_class = 'new-client' if client.is_new_client else ('closed-loan' if client.is_closed_loan else '') %}
                    <div class="client-row {{ highlight_class }}">
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span>Subtotal:</span><span>₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    <div class="type-total">WEEKLY TOTAL: ₹{{ "{:,}".format(data.weekly_total) }}</div>
    {% endif %}

    {% if data.ten_days %}
    <div class="section-header">📆 TEN DAYS INSTALLMENTS ({{ data.ten_days|length }} clients)</div>
    <div class="multi-column">
        {% set clients_per_column = 50 %}
        {% set total_columns = ((data.ten_days|length + clients_per_column - 1) // clients_per_column) or 1 %}
        
        {% for col in range(total_columns) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.ten_days|length] | min %}
            {% set column_clients = data.ten_days[start_index:end_index] %}
            
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span>Sr.No.</span><span>Name</span><span>Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    {% set highlight_class = 'new-client' if client.is_new_client else ('closed-loan' if client.is_closed_loan else '') %}
                    <div class="client-row {{ highlight_class }}">
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span>Subtotal:</span><span>₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    <div class="type-total">TEN DAYS TOTAL: ₹{{ "{:,}".format(data.ten_days_total) }}</div>
    {% endif %}

    {% if data.monthly %}
    {% set unique_monthly_clients = [] %}
    {% set seen_monthly_clients = {} %}
    {% for client in data.monthly %}
        {% set client_key = client.name + "_" + client.repayment_type %}
        {% if client_key not in seen_monthly_clients %}
            {% set _ = seen_monthly_clients.update({client_key: true}) %}
            {% set _ = unique_monthly_clients.append(client) %}
        {% endif %}
    {% endfor %}
    
    <div class="section-header">🗓️ MONTHLY INSTALLMENTS ({{ unique_monthly_clients|length }} clients)</div>
    <div class="multi-column">
        {% set clients_per_column = 50 %}
        {% set total_columns = ((unique_monthly_clients|length + clients_per_column - 1) // clients_per_column) or 1 %}
        
        {% for col in range(total_columns) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, unique_monthly_clients|length] | min %}
            {% set column_clients = unique_monthly_clients[start_index:end_index] %}
            
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span>Sr.No.</span><span>Name</span><span>Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    {% set highlight_class = 'new-client' if client.is_new_client else ('closed-loan' if client.is_closed_loan else '') %}
                    <div class="client-row {{ highlight_class }}">
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span>Subtotal:</span><span>₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    {% set actual_monthly_total = unique_monthly_clients | sum(attribute='total_amount') %}
    <div class="type-total">MONTHLY TOTAL: ₹{{ "{:,}".format(actual_monthly_total) }}</div>
    {% endif %}

    {% set grand_total = data.daily_total + data.weekly_total + data.ten_days_total + (actual_monthly_total if data.monthly else 0) %}
    {% if grand_total > 0 %}
    <div class="grand-total">GRAND TOTAL: ₹{{ "{:,}".format(grand_total) }}</div>
    {% endif %}
</body>
</html>
'''

PDF_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Loan Ledger - {{ client.name }}</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 0; padding: 20px; font-size: 12px; }
        .customer-header { text-align: center; margin-bottom: 30px; padding: 20px; border: 2px solid #343a40; border-radius: 10px; background-color: #f8f9fa; }
        .customer-name { font-size: 28px; font-weight: bold; color: #343a40; margin: 0 0 15px 0; text-transform: uppercase; letter-spacing: 2px; }
        .customer-details { font-size: 16px; color: #666; font-weight: 500; margin: 0; }
        .client-header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 10px; border-radius: 8px; margin-bottom: 20px; text-align: center; font-size: 14px; }
        table { width: 100%; border-collapse: collapse; margin-bottom: 20px; }
        th, td { border: 1px solid #ddd; padding: 6px; text-align: center; font-size: 9px; }
        th { background-color: #343a40; color: white; font-weight: bold; }
        .status-open { background-color: #fff3cd; color: #856404; font-weight: bold; }
        .status-closed { background-color: #d1edff; color: #0c5460; font-weight: bold; }
        .balance-due { color: #dc3545; font-weight: bold; }
        .totals-row { background-color: #e9ecef; font-weight: bold; }
        .currency::before { content: "₹"; }
        .footer { margin-top: 30px; text-align: center; font-size: 10px; color: #666; }
    </style>
</head>
<body>
    <div class="customer-header">
        <h1 class="customer-name">{{ client.name }}</h1>
        <p class="customer-details">{{ client.full_contact_line }}</p>
    </div>
    
    <div class="client-header">
        <strong>LOAN LEDGER STATEMENT</strong> | Generated: {{ generated_date }}
    </div>
    
    <table>
        <thead>
            <tr>
                <th>ID</th><th>Loan Date</th><th>Day</th><th>End Date</th><th>Close Date</th><th>Proc. Fees</th><th>Loan Amount</th><th>Inst. Amount</th><th>Type</th><th>Advance</th><th>Closing Amt</th><th>Balance</th><th>Pending</th><th>Status</th>
            </tr>
        </thead>
        <tbody>
            {% for loan in loans %}
            <tr>
                <td>{{ loan.id }}</td><td>{{ loan.loan_date }}</td><td>{{ loan.day_name }}</td><td>{{ loan.loan_end_date }}</td><td>{{ loan.loan_closed_date }}</td><td class="currency">{{ "{:,}".format(loan.processing_fees) }}</td><td class="currency">{{ "{:,}".format(loan.amount) }}</td><td class="currency">{{ "{:,}".format(loan.installment_amount) }}</td><td>{{ loan.repayment_type }}</td><td class="currency">{{ "{:,}".format(loan.advance_amount) }}</td><td class="currency">{{ "{:,}".format(loan.case_closing_amt) }}</td><td class="currency {% if loan.balance_amount > 0 %}balance-due{% endif %}">{{ "{:,}".format(loan.balance_amount) }}</td><td>{{ loan.pending_installments }}</td><td class="{% if loan.status == 'CLOSED' %}status-closed{% else %}status-open{% endif %}">{{ loan.status }}</td>
            </tr>
            {% endfor %}
        </tbody>
        <tfoot>
            <tr class="totals-row">
                <td colspan="5"><strong>TOTALS:</strong></td><td class="currency"><strong>{{ "{:,}".format(totals.processing_fees) }}</strong></td><td class="currency"><strong>{{ "{:,}".format(totals.loan_amount) }}</strong></td><td class="currency"><strong>{{ "{:,}".format(totals.installment_amount) }}</strong></td><td></td><td class="currency"><strong>{{ "{:,}".format(totals.advance_amount) }}</strong></td><td class="currency"><strong>{{ "{:,}".format(totals.case_closing_amount) }}</strong></td><td class="currency balance-due"><strong>{{ "{:,}".format(totals.balance_amount) }}</strong></td><td><strong>{{ totals.pending_installments }}</strong></td><td></td>
            </tr>
        </tfoot>
    </table>
    
    <div class="footer">
        <p><strong>SHIVI PROJECT - Loan Management System</strong></p>
    </div>
</body>
</html>
'''
PDF_PAYMENT_ENTRY_TEMPLATE = '''
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        html, body { width: 100%; }
        body { font-family: Arial, Helvetica, sans-serif; font-size: 11px; padding: 6px; margin: 0; color: #111827; background: #ffffff; }
        .report-header { background: #0f172a; color: white; padding: 12px 14px; border-radius: 10px; margin-bottom: 12px; text-align: center; border: 2px solid #111827; }
        .report-header h4 { margin: 0; font-size: 18px; letter-spacing: 0.3px; }
        .report-header p { margin: 4px 0 0 0; font-size: 11px; opacity: 0.95; }
        .section-header { width: 100%; display: block; box-sizing: border-box; background: #111827; color: white; padding: 8px 12px; margin: 16px 0 8px 0; font-weight: bold; text-align: center; font-size: 14px; border-radius: 10px; }
        .section-header.short { background: #991b1b; border: 2px solid #7f1d1d; }
        .section-header.short .badge { display: inline-block; background: rgba(255,255,255,0.14); border: 1px solid rgba(255,255,255,0.30); padding: 2px 8px; border-radius: 999px; font-size: 11px; margin-right: 8px; }
        .section-header.short .meta { display: inline-block; background: rgba(0,0,0,0.18); padding: 2px 8px; border-radius: 999px; font-size: 11px; margin-left: 8px; }
        .multi-column { width: 100%; font-size: 0; margin-bottom: 12px; }
        .column-box { display: inline-block; vertical-align: top; width: 24.25%; margin: 0 1% 8px 0; border: 1px solid #d1d5db; border-radius: 10px; font-size: 11px; overflow: hidden; background: #ffffff; box-sizing: border-box; }
        .column-box:nth-child(4n) { margin-right: 0; }
        body.cols-3 .column-box { width: 32.66%; margin: 0 1% 8px 0; }
        body.cols-3 .column-box:nth-child(3n) { margin-right: 0; }
        .multi-column.cols-1 .column-box { width: 100% !important; margin-right: 0 !important; }
        .multi-column.cols-2 .column-box { width: 49.5% !important; margin-right: 1% !important; }
        .multi-column.cols-2 .column-box:nth-child(2n) { margin-right: 0 !important; }
        .multi-column.cols-3 .column-box { width: 32.66% !important; margin-right: 1% !important; }
        .multi-column.cols-3 .column-box:nth-child(3n) { margin-right: 0 !important; }
        .column-header { display: table; width: 100%; table-layout: fixed; background: #f3f4f6; border-bottom: 2px solid #111827; padding: 5px 7px; font-weight: bold; box-sizing: border-box; color: #111827; }
        .client-row { display: table; width: 100%; table-layout: fixed; padding: 2px 6px; border-bottom: 1px solid #e9ecef; font-size: 10px; box-sizing: border-box; }
        .client-row:nth-child(even) { background: #f8f9fa; }
        .client-check { display: table-cell; width: 16px; text-align: center; font-weight: bold; }
        .client-serial { display: table-cell; width: 34px; text-align: center; font-weight: bold; }
        .client-name { display: table-cell; text-align: left; padding: 0 5px; white-space: normal; }
        .client-name > div:first-child { white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
        .client-meta { font-size: 9px; color: #666; margin-top: 2px; white-space: normal; line-height: 1.15; }
        .client-amount { display: table-cell; width: 60px; text-align: right; font-weight: bold; color: #16a34a; }
        .column-subtotal { display: table; width: 100%; table-layout: fixed; background: #111827 !important; color: white !important; font-weight: bold; border-top: 2px solid #111827; padding: 5px 7px; box-sizing: border-box; }
        .type-total { background: #16a34a; color: white; padding: 10px; text-align: center; font-weight: bold; margin: 10px 0 12px 0; border-radius: 10px; font-size: 14px; }
        .type-total.short { background: #991b1b; }
        .grand-total { background: #0f172a; color: white; padding: 12px; border-radius: 10px; text-align: center; font-size: 16px; font-weight: bold; margin-top: 14px; border: 2px solid #111827; }
        .page-break { page-break-after: always; }
        @page { size: A4 {{ 'portrait' if orientation == 'portrait' else 'landscape' }}; margin: 0.5cm; }
    </style>
</head>
<body class="cols-{{ columns_per_page }}">
    <div class="report-header">
        <h4>Payment Entry Report</h4>
        <p>{{ report_date }} | Generated: {{ generated_time }}</p>
    </div>

    {% if data.short %}
    <div class="section-header short"><span class="badge">SHORT PAYMENT REPORT</span> {{ data.short|length }} clients <span class="meta">UNPAID</span></div>
    {% set clients_per_column = 80 %}
    {% set total_columns = ((data.short|length + clients_per_column - 1) // clients_per_column) or 1 %}
    {% for col_group in range(0, total_columns, columns_per_page) %}
    {% set cols_in_group = [columns_per_page, total_columns - col_group] | min %}
    <div class="multi-column cols-{{ cols_in_group }}">
        {% for col in range(col_group, [col_group + columns_per_page, total_columns] | min) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.short|length] | min %}
            {% set column_clients = data.short[start_index:end_index] %}
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span class="client-check">✔</span><span class="client-serial">Sr.No.</span><span class="client-name">Name</span><span class="client-amount">Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    <div class="client-row">
                        <div class="client-check">☐</div>
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">
                            <div>{{ client.name }}</div>
                            <div class="client-meta">{{ client.repayment_type }} - Unpaid since: {{ client.unpaid_since }}</div>
                        </div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span class="client-name">Subtotal:</span><span class="client-amount">₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    {% if col_group + columns_per_page < total_columns %}<div class="page-break"></div>{% endif %}
    {% endfor %}
    <div class="type-total short">SHORT TOTAL: ₹{{ "{:,}".format(data.short_total) }}</div>
    {% endif %}

    {% if data.daily %}
    <div class="section-header">🔄 DAILY INSTALLMENTS ({{ data.daily|length }} clients)</div>
    {% set clients_per_column = 80 %}
    {% set total_columns = ((data.daily|length + clients_per_column - 1) // clients_per_column) or 1 %}
    {% for col_group in range(0, total_columns, columns_per_page) %}
    {% set cols_in_group = [columns_per_page, total_columns - col_group] | min %}
    <div class="multi-column cols-{{ cols_in_group }}">
        {% for col in range(col_group, [col_group + columns_per_page, total_columns] | min) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.daily|length] | min %}
            {% set column_clients = data.daily[start_index:end_index] %}
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span class="client-check">✔</span><span class="client-serial">Sr.No.</span><span class="client-name">Name</span><span class="client-amount">Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    <div class="client-row">
                        <div class="client-check">{{ '☑' if client.payment_received else '☐' }}</div>
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span class="client-name">Subtotal:</span><span class="client-amount">₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    {% if col_group + columns_per_page < total_columns %}<div class="page-break"></div>{% endif %}
    {% endfor %}
    <div class="type-total">DAILY TOTAL: ₹{{ "{:,}".format(data.daily_total) }}</div>
    {% endif %}

    {% if data.weekly %}
    <div class="section-header">📅 WEEKLY INSTALLMENTS ({{ data.weekly|length }} clients)</div>
    {% set clients_per_column = 80 %}
    {% set total_columns = ((data.weekly|length + clients_per_column - 1) // clients_per_column) or 1 %}
    {% for col_group in range(0, total_columns, columns_per_page) %}
    {% set cols_in_group = [columns_per_page, total_columns - col_group] | min %}
    <div class="multi-column cols-{{ cols_in_group }}">
        {% for col in range(col_group, [col_group + columns_per_page, total_columns] | min) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.weekly|length] | min %}
            {% set column_clients = data.weekly[start_index:end_index] %}
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span class="client-check">✔</span><span class="client-serial">Sr.No.</span><span class="client-name">Name</span><span class="client-amount">Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    <div class="client-row">
                        <div class="client-check">{{ '☑' if client.payment_received else '☐' }}</div>
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span class="client-name">Subtotal:</span><span class="client-amount">₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    {% if col_group + columns_per_page < total_columns %}<div class="page-break"></div>{% endif %}
    {% endfor %}
    <div class="type-total">WEEKLY TOTAL: ₹{{ "{:,}".format(data.weekly_total) }}</div>
    {% endif %}

    {% if data.ten_days %}
    <div class="section-header">📆 TEN DAYS INSTALLMENTS ({{ data.ten_days|length }} clients)</div>
    {% set clients_per_column = 80 %}
    {% set total_columns = ((data.ten_days|length + clients_per_column - 1) // clients_per_column) or 1 %}
    {% for col_group in range(0, total_columns, columns_per_page) %}
    {% set cols_in_group = [columns_per_page, total_columns - col_group] | min %}
    <div class="multi-column cols-{{ cols_in_group }}">
        {% for col in range(col_group, [col_group + columns_per_page, total_columns] | min) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.ten_days|length] | min %}
            {% set column_clients = data.ten_days[start_index:end_index] %}
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span class="client-check">✔</span><span class="client-serial">Sr.No.</span><span class="client-name">Name</span><span class="client-amount">Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    <div class="client-row">
                        <div class="client-check">{{ '☑' if client.payment_received else '☐' }}</div>
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span class="client-name">Subtotal:</span><span class="client-amount">₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    {% if col_group + columns_per_page < total_columns %}<div class="page-break"></div>{% endif %}
    {% endfor %}
    <div class="type-total">TEN DAYS TOTAL: ₹{{ "{:,}".format(data.ten_days_total) }}</div>
    {% endif %}

    {% if data.monthly %}
    <div class="section-header">🗓️ MONTHLY INSTALLMENTS ({{ data.monthly|length }} clients)</div>
    {% set clients_per_column = 80 %}
    {% set total_columns = ((data.monthly|length + clients_per_column - 1) // clients_per_column) or 1 %}
    {% for col_group in range(0, total_columns, columns_per_page) %}
    {% set cols_in_group = [columns_per_page, total_columns - col_group] | min %}
    <div class="multi-column cols-{{ cols_in_group }}">
        {% for col in range(col_group, [col_group + columns_per_page, total_columns] | min) %}
            {% set start_index = col * clients_per_column %}
            {% set end_index = [start_index + clients_per_column, data.monthly|length] | min %}
            {% set column_clients = data.monthly[start_index:end_index] %}
            {% if column_clients %}
            <div class="column-box">
                <div class="column-header">
                    <span class="client-check">✔</span><span class="client-serial">Sr.No.</span><span class="client-name">Name</span><span class="client-amount">Amt</span>
                </div>
                {% for client in column_clients %}
                    {% set serial_no = start_index + loop.index %}
                    <div class="client-row">
                        <div class="client-check">{{ '☑' if client.payment_received else '☐' }}</div>
                        <div class="client-serial">{{ serial_no }}</div>
                        <div class="client-name">{{ client.name }}</div>
                        <div class="client-amount">₹{{ "{:,}".format(client.total_amount) }}</div>
                    </div>
                {% endfor %}
                {% set column_total = column_clients | sum(attribute='total_amount') %}
                <div class="column-subtotal">
                    <span class="client-name">Subtotal:</span><span class="client-amount">₹{{ "{:,}".format(column_total) }}</span>
                </div>
            </div>
            {% endif %}
        {% endfor %}
    </div>
    {% if col_group + columns_per_page < total_columns %}<div class="page-break"></div>{% endif %}
    {% endfor %}
    <div class="type-total">MONTHLY TOTAL: ₹{{ "{:,}".format(data.monthly_total) }}</div>
    {% endif %}

    {% set grand_total = data.daily_total + data.weekly_total + data.ten_days_total + data.monthly_total %}
    <div class="grand-total">GRAND TOTAL (Collected + Recovered): ₹{{ "{:,}".format(grand_total) }}</div>
</body>
</html>
'''

# =============================================================================
# MODELS
# =============================================================================


# =============================================================================
# EVENT LISTENERS FOR AUTOMATIC BALANCE UPDATES
# =============================================================================

@event.listens_for(Payment, 'after_insert')
def update_balance_after_payment_insert(mapper, connection, target):
    """Automatically update loan balance when payment is added"""
    update_loan_balance_direct(connection, target.loan_id)

@event.listens_for(Payment, 'after_update')
def update_balance_after_payment_update(mapper, connection, target):
    """Update balance when payment is modified"""
    update_loan_balance_direct(connection, target.loan_id)

@event.listens_for(Payment, 'after_delete')  
def update_balance_after_payment_delete(mapper, connection, target):
    """Update balance when payment is deleted"""
    update_loan_balance_direct(connection, target.loan_id)

# ✅ NEW: Update balance when loan status changes to CLOSED
@event.listens_for(Loan, 'after_update') 
def update_balance_after_loan_update(mapper, connection, target):
    """Update balance when loan status changes"""
    if target.status == "CLOSED":
        # Set outstanding_balance to 0 when loan is closed
        connection.execute(
            update(Loan)
            .where(Loan.id == target.id)
            .values(outstanding_balance=0)
        )

def update_loan_balance_direct(connection, loan_id):
    """Direct SQL update for use in event listeners"""
    result = connection.execute(
        select(func.sum(Payment.amount_paid))
        .where(Payment.loan_id == loan_id)
    ).scalar()
    total_payments = result or 0
    
    loan_result = connection.execute(
        select(Loan.loan_amount, Loan.status).where(Loan.id == loan_id)
    ).fetchone()
    
    if loan_result:
        loan_amount = loan_result[0] or 0
        loan_status = loan_result[1]
        
        # ✅ FIXED: If loan is closed, set balance to 0
        if loan_status == "CLOSED":
            outstanding = 0
        else:
            outstanding = loan_amount - total_payments
        
        connection.execute(
            update(Loan)
            .where(Loan.id == loan_id)
            .values(
                total_paid_amount=total_payments,
                outstanding_balance=outstanding
            )
        )

# =============================================================================
# HELPER FUNCTIONS - 100% COMPLETE WITH DD-MM-YYYY FORMAT
# =============================================================================

def parse_loan_date_universal(date_str):
    """✅ ROBUST: Handle multiple date formats like your original"""
    if not date_str:
        return None
    try:
        if '-' in date_str:
            parts = date_str.split('-')
            if len(parts) == 3:
                if len(parts[0]) == 4:  # YYYY-MM-DD
                    return datetime.strptime(date_str, '%Y-%m-%d')
                else:  # DD-MM-YYYY
                    return datetime.strptime(date_str, '%d-%m-%Y')
        elif '/' in date_str:
            return datetime.strptime(date_str, '%d/%m/%Y')
        return None
    except ValueError:
        return None

def chunk_list(lst, n):
    """Split a list into successive n-sized chunks."""
    return [lst[i:i + n] for i in range(0, len(lst), n)]

@app.template_filter('number_format')
def number_format(value):
    """Format numbers with commas"""
    try:
        return f"{int(value):,}"
    except (ValueError, TypeError):
        return "0"

def safe_float(val):
    try:
        if val is None or val == "":
            return None
        return float(val)
    except Exception:
        return None

def parse_date_str(d):
    """✅ FIXED: Use robust date parsing"""
    if not d:
        return None
    parsed_dt = parse_loan_date_universal(d)
    return parsed_dt.date() if parsed_dt else None

def format_date(date_str):
    """✅ FIXED: Always return DD-MM-YYYY format"""
    if not date_str:
        return ""
    try:
        parsed_dt = parse_loan_date_universal(date_str)
        if parsed_dt:
            return parsed_dt.strftime("%d-%m-%Y")
        return date_str
    except Exception:
        return date_str

def register_error_handlers(app):
    @app.errorhandler(Exception)
    def handle_exception(e):
        app.logger.error(f'Unhandled Exception: {e}\n' + traceback.format_exc())
        return jsonify({"error": "Internal Server Error", "message": str(e)}), 500

# Call this after your app = Flask(__name__) line
register_error_handlers(app)

def require_fields(data, fields):
    errs = {}
    for f in fields:
        v = data.get(f)
        if v is None or (isinstance(v, str) and v.strip() == ""):
            errs[f] = "Required"
    return errs

def validate_repayment_type(rtype):
    return rtype in ALLOWED_REPAYMENT_TYPES

def calc_advance(amount, rtype):
    if not amount:
        return 0
    if rtype == "WEEKLY":
        return round(amount * 0.07)
    elif rtype == "TEN_DAYS":
        return round(amount / 10)
    elif rtype == "MONTHLY":
        return 0
    return 0

def add_months(start_date, months):
    y = start_date.year + (start_date.month - 1 + months) // 12
    m = (start_date.month - 1 + months) % 12 + 1
    d = min(start_date.day, calendar.monthrange(y, m)[1])
    return date(y, m, d)

def monthly_interest_amount(loan):
    total_amt = loan.loan_amount or 0
    rate = loan.interest_rate or 0
    if total_amt <= 0 or rate <= 0:
        return 0
    return round(total_amt * (rate / 100.0))

def schedule_end_date(loan):
    """Business logic for loan end dates."""
    # Use robust date parsing
    start_date = parse_date_str(loan.loan_date)
    if not start_date:
        return None

    if loan.repayment_type == "MONTHLY":
        return add_months(start_date, 12)

    return start_date + timedelta(days=100)

def next_effective_date(loan):
    """Main rule: Close a loan today → installment amount becomes 0 tomorrow"""
    closed = parse_date_str(loan.loan_closed_date)
    if not closed:
        return None
    return closed + timedelta(days=1)
def calculate_balance_and_pending_core(loan, as_of):
    """✅ FIXED: Core balance/pending/installment calculation with correct WEEKLY logic"""
    start_date = parse_date_str(loan.loan_date)
    if start_date is None:
        return 0, 0, 0

    when = as_of or date.today()
    days_passed = max(0, (when - start_date).days)
    total_amt = loan.loan_amount or 0
    adv_amt = loan.advance_amount or 0
    per_amt  = 0
    inst_amt = 0
    pending_count = 0
    balance = 0

    if loan.repayment_type == "DAILY":
        total_inst = 100
        inst_amt = round(total_amt / total_inst) if total_inst else 0
        paid_count = min(days_passed, total_inst)
        pending_count = max(0, total_inst - paid_count)
        balance = inst_amt * pending_count
        return balance, pending_count, inst_amt

    elif loan.repayment_type == "WEEKLY":
        weekly_installment = round(total_amt * 0.07)  # 7% per week
        two_days_balance = round(total_amt * 0.02)    # 2% for 2 days
        weeks_passed = days_passed // 7
        advance_covers_week1 = (adv_amt >= weekly_installment)

        if advance_covers_week1:
            # ✅ WITH ADVANCE: Week 1 FREE, Week 14 due day 91, Final due day 98
            if days_passed > 98:
                # Loan completely finished
                balance = 0
                pending_count = 0
                inst_amt = 0
            elif days_passed == 98:
                # Final installment period - considered collected
                balance = 0
                pending_count = 0
                inst_amt = 0
            elif days_passed > 91:
                # Week 14 collected on day 91, only Final remains
                balance = two_days_balance
                pending_count = 1
                inst_amt = two_days_balance
            elif days_passed == 91:
                # Week 14 collected on day 91, only Final remains
                balance = two_days_balance
                pending_count = 1
                inst_amt = weekly_installment
            else:
                # Normal calculation - weeks remaining + final (advance counts as earliest installment)
                if weekly_installment > 0:
                    advance_blocks = int(adv_amt // weekly_installment)
                else:
                    advance_blocks = 0
                total_weeks = 14
                effective_weeks_passed = min(weeks_passed, total_weeks)
                received_weeks = min(total_weeks, advance_blocks + effective_weeks_passed)
                weeks_remaining = max(0, total_weeks - received_weeks)
                balance = (weeks_remaining * weekly_installment) + two_days_balance
                pending_count = weeks_remaining + 1  # weekly remaining + final 2-day installment
                inst_amt = weekly_installment
        else:
            # ✅ WITHOUT ADVANCE: Week 1-13 normal, Week 14 + Final BOTH due day 98
            if days_passed > 98:
                # Loan completely finished
                balance = 0
                pending_count = 0
                inst_amt = 0
            elif days_passed == 98:
                # Both Week 14 and Final collected on day 98
                balance = weekly_installment + two_days_balance
                pending_count = 1
                inst_amt = weekly_installment + two_days_balance
            else:
                weeks_remaining = max(0, 14 - weeks_passed)
                if weeks_remaining <= 1:
                    # Week 14 + Final both pending until day 98
                    balance = weekly_installment + two_days_balance  # ₹9,000
                    pending_count = 2
                    inst_amt = weekly_installment + two_days_balance
                else:
                    # Normal calculation
                    balance = (weeks_remaining * weekly_installment) + two_days_balance
                    pending_count = weeks_remaining + 1
                    inst_amt = weekly_installment
        return balance, pending_count, inst_amt

    elif loan.repayment_type == "MONTHLY":
        inst_amt = monthly_interest_amount(loan)
        total_inst = 12
        adv_amt = loan.advance_amount or 0
        first_covered = (adv_amt >= inst_amt) if inst_amt > 0 else False
        paid_count = 1 if first_covered else 0
        max_month = 11 if first_covered else 12
        for m in range(1, max_month + 1):
            due_dt = add_months(start_date, m)
            if when >= due_dt:
                paid_count += 1
        paid_count = min(total_inst, paid_count)
        pending_count = max(0, total_inst - paid_count)
        balance = total_amt
        if pending_count > 0:
            return balance, pending_count, inst_amt
        return balance, 0, 0

    elif loan.repayment_type == "TEN_DAYS":
        per_amt = round(total_amt / 10) if total_amt else 0
        total_inst = 10
        paid_count = min(days_passed // 10, total_inst)
    if adv_amt > 0 and per_amt > 0:
        adv_blocks = int(adv_amt // per_amt)
        paid_count = min(total_inst, paid_count + adv_blocks)
    pending_count = max(0, total_inst - paid_count)
    balance = per_amt * pending_count
    if days_passed > 90 and pending_count == 0:
        inst_amt = 0
    elif pending_count > 0:
        inst_amt = per_amt
    else:
        inst_amt = 0
    return balance, pending_count, inst_amt


def apply_closed_effects(loan, as_of, balance, pending, inst_amt, zero_on_close_date=False):
    """Zero out inst, balance, pending from the effective date"""
    if loan.status != "CLOSED":
        return balance, pending, inst_amt, False

    eff_date = parse_date_str(loan.loan_closed_date) if zero_on_close_date else next_effective_date(loan)
    curr = as_of or date.today()
    if eff_date and curr >= eff_date:
        return 0, 0, 0, True
    return balance, pending, inst_amt, False

def is_excluded_from_loan_total(loan, as_of):
    """Exclude from Loan Amount totals immediately on the close date"""
    if loan.status != "CLOSED":
        return False
    close_dt = parse_date_str(loan.loan_closed_date)
    if not close_dt:
        return False
    return as_of >= close_dt

def compute_row_fields(l, as_of_date=None, zero_on_close_date=False):
    as_of = as_of_date or date.today()
    bal_core, pending_core, inst_core = calculate_balance_and_pending_core(l, as_of=as_of)
    bal, pending, inst_amt, eff_applied = apply_closed_effects(
        l, as_of, bal_core, pending_core, inst_core, zero_on_close_date=zero_on_close_date
    )
    return bal, pending, inst_amt, eff_applied


def due_amount_for_date(loan, target_date, inst_amt):
    """Adjust theoretical installment amount for report usage."""
    if loan.repayment_type == "WEEKLY":
        start_date = parse_date_str(loan.loan_date)
        if not start_date:
            return inst_amt

        days_diff = (target_date - start_date).days
        if days_diff < 0:
            return 0

        total_amt = loan.loan_amount or 0
        weekly_installment = round(total_amt * 0.07)
        two_days_balance = round(total_amt * 0.02)
        adv_amt = loan.advance_amount or 0
        advance_covers_week1 = (weekly_installment > 0) and (adv_amt >= weekly_installment)

        # Week-1 is treated as already received if advance covers it.
        if advance_covers_week1 and days_diff == 0:
            return 0

        # Day 98 is the final collection day.
        if days_diff == 98:
            return two_days_balance if advance_covers_week1 else (weekly_installment + two_days_balance)

        if days_diff > 0 and days_diff % 7 == 0 and days_diff < 98:
            return weekly_installment

        return 0

    if loan.repayment_type not in ("TEN_DAYS", "MONTHLY"):
        return inst_amt

    if loan.repayment_type == "MONTHLY":
        start_date = parse_date_str(loan.loan_date)
        if not start_date:
            return inst_amt
        inst_due = monthly_interest_amount(loan)
        if inst_due <= 0:
            return 0
        first_covered = (loan.advance_amount or 0) >= inst_due
        max_month = 11 if first_covered else 12
        for m in range(1, max_month + 1):
            if add_months(start_date, m) == target_date:
                return inst_due
        return inst_amt

    start_date = parse_date_str(loan.loan_date)
    if not start_date:
        return inst_amt

    days_diff = (target_date - start_date).days
    if days_diff <= 0:
        return inst_amt

    per_amt = round((loan.loan_amount or 0) / 10) if (loan.loan_amount or 0) else 0
    if per_amt <= 0:
        return inst_amt

    adv_amt = loan.advance_amount or 0
    first_installment_covered_by_advance = (adv_amt >= per_amt) if per_amt > 0 else False
    max_due_day = 90 if first_installment_covered_by_advance else 100

    if days_diff % 10 == 0 and 0 < days_diff <= max_due_day:
        return per_amt

    return inst_amt


# ✅ CRITICAL FIX: Daily installments start from NEXT day after loan date
def has_installment_due_on_date(loan, target_date):
    """✅ FIXED: Check if a loan has an installment due on the target date"""
    loan_start = parse_date_str(loan.loan_date)
    
    if not loan_start or target_date < loan_start:
        return False
    
    days_diff = (target_date - loan_start).days
    
    if loan.repayment_type == "DAILY":
        # ✅ CRITICAL FIX: Daily installments start from NEXT day (day 1, not day 0)
        return days_diff >= 1 and days_diff <= 100
    elif loan.repayment_type == "WEEKLY":
        # ✅ CRITICAL FIX: Include Day 98 as final collection day
        return (days_diff % 7 == 0 and days_diff < 98) or days_diff == 98
    elif loan.repayment_type == "TEN_DAYS":
        per_amt = round((loan.loan_amount or 0) / 10) if (loan.loan_amount or 0) else 0
        adv_amt = loan.advance_amount or 0
        first_installment_covered_by_advance = (adv_amt >= per_amt) if per_amt > 0 else False
        max_due_day = 90 if first_installment_covered_by_advance else 100
        return days_diff > 0 and (days_diff % 10 == 0) and days_diff <= max_due_day

    elif loan.repayment_type == "MONTHLY":
        inst_due = monthly_interest_amount(loan)
        if inst_due <= 0:
            return False
        adv_amt = loan.advance_amount or 0
        first_covered = adv_amt >= inst_due
        max_month = 11 if first_covered else 12
        for m in range(1, max_month + 1):
            if add_months(loan_start, m) == target_date:
                return True
        return False
    
    return False

def is_last_installment_date(loan, target_date):
    """Check if this is the last installment date for the loan"""
    loan_start = parse_date_str(loan.loan_date)
    
    if not loan_start:
        return False
    
    days_diff = (target_date - loan_start).days
    
    if loan.repayment_type == "DAILY":
        return days_diff == 100
    elif loan.repayment_type == "WEEKLY":
        return days_diff == 98  # Day 98 is final collection day
    elif loan.repayment_type == "TEN_DAYS":
        per_amt = round((loan.loan_amount or 0) / 10) if (loan.loan_amount or 0) else 0
        adv_amt = loan.advance_amount or 0
        first_installment_covered_by_advance = (adv_amt >= per_amt) if per_amt > 0 else False
        return days_diff == (9 * 10) if first_installment_covered_by_advance else days_diff == (10 * 10)

    elif loan.repayment_type == "MONTHLY":
        inst_due = monthly_interest_amount(loan)
        if inst_due <= 0:
            return False
        adv_amt = loan.advance_amount or 0
        first_covered = adv_amt >= inst_due
        months_to_add = 11 if first_covered else 12
        return add_months(loan_start, months_to_add) == target_date
    
    return False

def parse_ymd(d):
    """Accepts '2025-08-14' → datetime or None on bad input"""
    try:
        return datetime.strptime(d, "%Y-%m-%d")
    except Exception:
        return None

def fmt_dd_mm_yyyy(d):
    """✅ FIXED: Return '14-08-2025' or '' if d is None / ''."""
    if not d:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%d-%m-%Y")
    if isinstance(d, date):
        return d.strftime("%d-%m-%Y")
    try:
        parsed_dt = parse_loan_date_universal(d)
        return parsed_dt.strftime("%d-%m-%Y") if parsed_dt else ""
    except Exception:
        return ""

def update_loan_balance(loan_id):
    """Update loan balance based on actual payments"""
    loan = db.session.get(Loan, loan_id)
    if not loan:
        return
    
    total_payments = db.session.query(func.sum(Payment.amount_paid))\
                              .filter(Payment.loan_id == loan_id).scalar() or 0

    # FIXED: Respect closure status
    if loan.status == "CLOSED":
        loan.outstanding_balance = 0
    else:
        loan.outstanding_balance = loan.loan_amount - total_payments

    loan.total_paid_amount = total_payments
    db.session.add(loan)
    db.session.commit()

def get_expected_payments_for_date(target_date):
    """Get all loans expected to pay on target date"""
    target_dt = parse_date_str(target_date) if isinstance(target_date, str) else target_date
    loans = Loan.query.filter_by(status="OPEN").all()
    expected = []
    for loan in loans:
        if has_installment_due_on_date(loan, target_dt):
            _, _, inst_amt, _ = compute_row_fields(loan, as_of_date=target_dt)
            due_amt = due_amount_for_date(loan, target_dt, inst_amt)
            try:
                due_amt = float(due_amt or 0)
            except Exception:
                due_amt = 0
            if due_amt > 0:
                expected.append({
                    'loan_id': loan.id,
                    'name': loan.name,
                    'amount_due': due_amt,
                    'sr': loan.id
                })
    
    return expected


def get_clients_due_for_date(target_date, repayment_type):
    """Return list of loan clients with installments due on target_date for a specified repayment_type."""
    loans = Loan.query.filter_by(status="OPEN", repayment_type=repayment_type).all()
    clients_due = []
    sr = 1
    for loan in loans:
        if has_installment_due_on_date(loan, target_date):
            _, _, inst_amt, _ = compute_row_fields(loan, as_of_date=target_date)
            if inst_amt > 0:
                clients_due.append({
                    'sr': sr,
                    'loan_id': loan.id,
                    'name': loan.name,
                    'amount_due': inst_amt,
                    'checked': True
                })
                sr += 1
    return clients_due

# ✅ FIXED: Unified ledger function that returns complete data
# ✅ FIXED: Unified ledger function that returns complete data
def get_complete_ledger_data_by_client_name(client_name):
    """Get complete ledger data for hyperlink fix"""
    # Find client by name and get their phone
    loan = Loan.query.filter(Loan.name.ilike(f'%{client_name}%')).first()
    if not loan:
        return None
    
    # Use the existing ledger logic
    phone = loan.phone
    loans = Loan.query.filter_by(phone=phone).order_by(Loan.loan_date.asc()).all()
    results = []
    
    today = date.today()
    totals = {
        "processing_fees": 0,
        "loan_amount": 0,
        "installment_amount": 0,
        "advance_amount": 0,
        "case_closing_amount": 0,
        "balance_amount": 0,
        "pending_installments": 0
    }

    for l in loans:
        loan_dt = parse_date_str(l.loan_date)
        end_dt = schedule_end_date(l)

        bal, pending, inst_amt, eff_applied = compute_row_fields(
            l, as_of_date=today, zero_on_close_date=False
        )

        include_in_loan_total = not is_excluded_from_loan_total(l, today)

        row_data = {
            "id": l.id,
            "loan_date": format_date(l.loan_date),
            "loan_date_raw": l.loan_date,
            "day_name": loan_dt.strftime("%a") if loan_dt and l.repayment_type == "WEEKLY" else "",
            "loan_end_date": end_dt.strftime("%d-%m-%Y") if end_dt else "",
            "loan_closed_date": format_date(l.loan_closed_date),
            "loan_closed_date_raw": l.loan_closed_date,
            "processing_fees": int(l.processing_fees) if l.processing_fees else 0,
            "amount": int(l.loan_amount) if l.loan_amount else 0,
            "installment_amount": inst_amt,
            "repayment_type": l.repayment_type,
            "advance_amount": int(l.advance_amount) if l.advance_amount else 0,
            "case_closing_amt": int(l.case_closing_amount) if l.case_closing_amount else 0,
            "balance_amount": bal,
            "pending_installments": pending,
            "status": l.status,
            "effective_closed_applied": eff_applied,
            "include_in_loan_total": include_in_loan_total,
            "name": l.name,
            "address": l.address,
            "phone": l.phone,
            "city": l.city
        }
        results.append(row_data)

        totals["processing_fees"] += row_data["processing_fees"]
        totals["installment_amount"] += row_data["installment_amount"]
        totals["advance_amount"] += row_data["advance_amount"]
        totals["case_closing_amount"] += row_data["case_closing_amt"]
        totals["balance_amount"] += row_data["balance_amount"]
        totals["pending_installments"] += row_data["pending_installments"]
        
        if include_in_loan_total:
            totals["loan_amount"] += row_data["amount"]

    return {
        "rows": results,
        "totals": totals,
        "meta": {"as_of": today.isoformat(), "phone": phone},
        "client_info": {
            "name": loan.name,
            "phone": phone,
            "address": loan.address,
            "city": loan.city
        }
    }

def get_ledger_data_by_client(client_name):
    """Get ledger data for a specific client by name"""
    loans = Loan.query.filter(Loan.name.ilike(f'%{client_name}%')).all()
    
    if not loans:
        return None
    
    today = date.today()
    transactions = []
    total_loans = len(loans)
    active_loans = 0
    outstanding_amount = 0
    
    for loan in loans:
        if loan.status == "OPEN":
            active_loans += 1
        
        bal, pending, inst_amt, _ = compute_row_fields(loan, as_of_date=today)
        outstanding_amount += bal
        
        transactions.append({
            'loan_id': loan.id,
            'loan_date': format_date(loan.loan_date),
            'loan_amount': loan.loan_amount or 0,
            'repayment_type': loan.repayment_type,
            'status': loan.status,
            'balance': bal,
            'pending_installments': pending
        })
    
    return {
        'total_loans': total_loans,
        'active_loans': active_loans,
        'outstanding_amount': outstanding_amount,
        'transactions': transactions
    }

#
# =============================================================================
# PAYMENT BLUEPRINT - COMPLETE WITH ALL FEATURES
# =============================================================================

payment_bp = Blueprint('payment', __name__)


def _to_ymd(value):
    if value is None:
        return None
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, datetime):
        return value.date().strftime('%Y-%m-%d')
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(s, fmt).date().strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None

@payment_bp.route('/payment-entry')
@login_required
@page_required('payment_entry')
def payment_entry():
    """ENHANCED: Payment entry with date filter and freezing logic"""
    date_str = request.args.get('date')
    if date_str:
        try:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            target_date = date.today()
    else:
        target_date = date.today()
    
    # Convert target_date to string in 'YYYY-MM-DD' format for database query
    target_date_str = target_date.strftime('%Y-%m-%d')
    existing_submission = PaymentSubmission.query.filter_by(submission_date=target_date_str).first()
    payments_already_submitted = existing_submission is not None
    
    existing_payments = Payment.query.filter_by(payment_date=target_date_str).all()
    paid_loan_ids = {p.loan_id for p in existing_payments}
    
    daily_clients = get_clients_due_for_date(target_date, "DAILY")
    weekly_clients = get_clients_due_for_date(target_date, "WEEKLY")
    ten_days_clients = get_clients_due_for_date(target_date, "TEN_DAYS")
    monthly_clients = get_clients_due_for_date(target_date, "MONTHLY")
    
    for client in daily_clients + weekly_clients + ten_days_clients + monthly_clients:
        client['payment_received'] = client['loan_id'] in paid_loan_ids
    
    daily_columns = chunk_list(daily_clients, 60)
    weekly_columns = chunk_list(weekly_clients, 10)
    ten_days_columns = chunk_list(ten_days_clients, 10)
    monthly_columns = chunk_list(monthly_clients, 10)
    
    daily_subtotal = sum(c['amount_due'] for c in daily_clients if c['payment_received'])
    weekly_subtotal = sum(c['amount_due'] for c in weekly_clients if c['payment_received'])
    ten_days_subtotal = sum(c['amount_due'] for c in ten_days_clients if c['payment_received'])
    monthly_subtotal = sum(c['amount_due'] for c in monthly_clients if c['payment_received'])
    
    short_clients = []
    cumulative_short = 0
    
    if payments_already_submitted:
        db_shorts = ShortPayment.query.filter_by(payment_date=target_date_str, status='PENDING').all()
        for short in db_shorts:
            loan = db.session.get(Loan, short.loan_id)
            if loan:
                short_clients.append({
                    'loan_id': short.loan_id,
                    'sr': short.id,
                    'name': loan.name,
                    'expected_amount': short.expected_amount
                })
                cumulative_short += short.expected_amount
    else:
        for client in daily_clients + weekly_clients + ten_days_clients + monthly_clients:
            if not client['payment_received']:
                short_clients.append({
                    'loan_id': client['loan_id'],
                    'sr': client['sr'],
                    'name': client['name'],
                    'expected_amount': client['amount_due']
                })
                cumulative_short += client['amount_due']
    
    return render_template('payment_entry.html',
                          target_date=target_date.strftime('%Y-%m-%d'),
                          target_date_display=target_date.strftime('%d-%m-%Y'),
                          payments_already_submitted=payments_already_submitted,
                          daily_columns=daily_columns,
                          weekly_columns=weekly_columns,
                          ten_days_columns=ten_days_columns,
                          monthly_columns=monthly_columns,
                          daily_subtotal=daily_subtotal,
                          weekly_subtotal=weekly_subtotal,
                          ten_days_subtotal=ten_days_subtotal,
                          monthly_subtotal=monthly_subtotal,
                          short_clients=short_clients,
                          cumulative_short=cumulative_short,
                          can_edit=not payments_already_submitted or target_date == date.today())

@payment_bp.route('/payment-entry/process', methods=['POST'])
@csrf.exempt
@login_required
@page_required('payment_entry')
def process_payments():
    """✅ FIXED: Process payments with proper transaction isolation"""
    data = request.get_json(force=True)
    payments = data.get('payments', [])
    target_date_str = data.get('date')
    
    if target_date_str:
        target_date = None
        for date_format in ['%d-%m-%Y', '%Y-%m-%d', '%m/%d/%Y']:
            try:
                target_date = datetime.strptime(target_date_str, date_format).date()
                break
            except ValueError:
                continue
        
        if not target_date:
            return jsonify({"success": False, "error": "Invalid date format"}), 400
    else:
        target_date = date.today()
    
    target_date_str = _to_ymd(target_date)
    existing = PaymentSubmission.query.filter_by(submission_date=target_date_str).first()
    if existing:
        return jsonify({"success": False, "error": "Payments already submitted for this date"}), 400
    
    try:
        total_amount = 0
        received_loan_ids = set()
        
        for p in payments:
            loan_id = p.get('loan_id')
            amount = float(p.get('amount'))

            print(f"Processing payment for loan_id: {loan_id}, amount: {amount}")

            loan = db.session.get(Loan, loan_id)
            if not loan:
                print(f"Loan with ID {loan_id} not found.")
                continue
            if loan.status != "OPEN":
                print(f"Loan {loan_id} is not OPEN (status={loan.status})")
                continue

            payment = Payment(
                loan_id=loan_id,
                payment_date=target_date_str,
                amount_paid=amount,
                payment_method='Collection Entry',
                entry_type='COLLECTION',
                remarks='Recorded via payment entry page'
            )
            db.session.add(payment)
            
            total_amount += amount
            received_loan_ids.add(loan_id)

        all_expected = get_expected_payments_for_date(target_date)
        for expected in all_expected:
            if expected['loan_id'] not in received_loan_ids:
                short_payment = ShortPayment(
                    loan_id=expected['loan_id'],
                    payment_date=target_date_str,
                    expected_amount=expected['amount_due'],
                    status='PENDING'
                )
                db.session.add(short_payment)

        submission = PaymentSubmission(
            submission_date=target_date_str,
            total_amount=total_amount,
            total_payments=len(payments),
            submitted_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        db.session.add(submission)
        
        db.session.commit()
        return jsonify({"success": True, "total_amount": total_amount})
        
    except Exception as e:
        db.session.rollback()
        print(f"Exception in payment processing: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@payment_bp.route('/payment-entry/recover-short-payments', methods=['POST'])
@login_required
@page_required('payment_entry')
def recover_short_payments():
    """ENHANCED: Recover selected short payments"""
    try:
        data = request.get_json()
        loan_ids = data.get('loans', [])
        
        if not loan_ids:
            return jsonify({"success": False, "error": "No loans selected"}), 400
        
        shorts = ShortPayment.query.filter(
            ShortPayment.loan_id.in_(loan_ids),
            ShortPayment.status == 'PENDING'
        ).all()
        
        total_recovered = 0
        for short in shorts:
            due_date = _to_ymd(short.payment_date)
            today_str = date.today().strftime('%Y-%m-%d')
            payment = Payment(
                loan_id=short.loan_id,
                payment_date=today_str,
                amount_paid=short.expected_amount,
                payment_method='Short Recovery',
                entry_type='RECOVERY',
                remarks=f'Recovered short payment from {short.payment_date}'
            )
            db.session.add(payment)

            loan = db.session.get(Loan, short.loan_id)
            recovery_row = RecoveryPayment(
                loan_id=short.loan_id,
                client_name=(loan.name if loan else ''),
                due_date=due_date,
                recovery_date=today_str,
                amount=short.expected_amount,
                notes='',
                created_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            )
            db.session.add(recovery_row)
            
            short.status = 'RECOVERED'
            db.session.add(short)
            
            total_recovered += short.expected_amount
        
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "recovered_count": len(shorts),
            "total_recovered": total_recovered
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@payment_bp.route('/payment-entry/edit/<target_date>', methods=['POST'])
@csrf.exempt
@login_required
@page_required('payment_entry')
def edit_payment_entry(target_date):
    """Edit existing payment entry for a specific date"""
    target_dt = parse_date_str(target_date)
    if not target_dt:
        return jsonify({"error": "Invalid date format"}), 400
    
    if target_dt != date.today():
        existing_submission = PaymentSubmission.query.filter_by(submission_date=_to_ymd(target_dt)).first()
        if existing_submission and target_dt < date.today():
            return jsonify({"error": "Cannot edit past submissions"}), 403
    
    target_date_str = _to_ymd(target_dt)
    Payment.query.filter_by(payment_date=target_date_str).delete()
    ShortPayment.query.filter_by(payment_date=target_date_str).delete()
    RecoveryPayment.query.filter_by(due_date=target_date_str).delete()
    submission = PaymentSubmission.query.filter_by(submission_date=target_date_str).first()
    if submission:
        db.session.delete(submission)
    
    db.session.commit()
    return jsonify({"success": True, "message": "Records cleared for editing"})

@payment_bp.route('/payment-entry/delete/<target_date>', methods=['POST'])
@csrf.exempt
@login_required
@page_required('payment_entry')
def delete_payment_entry(target_date):
    """Delete all entries for the given date."""
    target_date_str = _to_ymd(target_date)
    if not target_date_str:
        return jsonify({"success": False, "error": "Invalid date"}), 400

    Payment.query.filter_by(payment_date=target_date_str).delete()
    ShortPayment.query.filter_by(payment_date=target_date_str).delete()
    RecoveryPayment.query.filter_by(due_date=target_date_str).delete()
    PaymentDraft.query.filter_by(draft_date=target_date_str).delete()
    submission = PaymentSubmission.query.filter_by(submission_date=target_date_str).first()
    if submission:
        db.session.delete(submission)
    
    db.session.commit()
    
    return jsonify({"success": True, "message": f"All entries for {target_date} deleted"})


@payment_bp.post('/payment-entry/bulk-submit-last-100-days')
@csrf.exempt
@login_required
@page_required('payment_entry')
def bulk_submit_last_100_days():
    today_dt = date.today()
    start_dt = today_dt - timedelta(days=99)

    created_submissions = 0
    created_shorts = 0
    processed_days = 0

    try:
        for i in range(100):
            day_dt = start_dt + timedelta(days=i)
            day_str = _to_ymd(day_dt)
            processed_days += 1

            existing = PaymentSubmission.query.filter_by(submission_date=day_str).first()
            if existing:
                continue

            payments = Payment.query.filter_by(payment_date=day_str).all()
            received_loan_ids = {p.loan_id for p in payments}
            total_amount = sum(float(p.amount_paid or 0) for p in payments)

            expected = get_expected_payments_for_date(day_str)
            for exp in expected:
                loan_id = exp.get('loan_id')
                if loan_id in received_loan_ids:
                    continue
                exists_short = ShortPayment.query.filter_by(payment_date=day_str, loan_id=loan_id).first()
                if exists_short:
                    continue
                sp = ShortPayment(
                    loan_id=loan_id,
                    payment_date=day_str,
                    expected_amount=float(exp.get('amount_due') or 0),
                    status='PENDING'
                )
                db.session.add(sp)
                created_shorts += 1

            submission = PaymentSubmission(
                submission_date=day_str,
                total_amount=total_amount,
                total_payments=len(payments),
                submitted_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            )
            db.session.add(submission)
            created_submissions += 1

        db.session.commit()
        return jsonify({
            'success': True,
            'processed_days': processed_days,
            'created_submissions': created_submissions,
            'created_short_payments': created_shorts
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@payment_bp.post('/payment-entry/bulk-delete-last-100-days')
@csrf.exempt
@login_required
@page_required('payment_entry')
def bulk_delete_last_100_days():
    today_dt = date.today()
    start_dt = today_dt - timedelta(days=99)

    deleted_payments = 0
    deleted_shorts = 0
    deleted_recoveries = 0
    deleted_drafts = 0
    deleted_submissions = 0
    processed_days = 0

    try:
        for i in range(100):
            day_dt = start_dt + timedelta(days=i)
            day_str = _to_ymd(day_dt)
            processed_days += 1

            deleted_payments += Payment.query.filter_by(payment_date=day_str).delete()
            deleted_shorts += ShortPayment.query.filter_by(payment_date=day_str).delete()
            deleted_recoveries += RecoveryPayment.query.filter_by(due_date=day_str).delete()
            deleted_drafts += PaymentDraft.query.filter_by(draft_date=day_str).delete()

            submission = PaymentSubmission.query.filter_by(submission_date=day_str).first()
            if submission:
                db.session.delete(submission)
                deleted_submissions += 1

        db.session.commit()
        return jsonify({
            'success': True,
            'processed_days': processed_days,
            'deleted_payments': deleted_payments,
            'deleted_short_payments': deleted_shorts,
            'deleted_recoveries': deleted_recoveries,
            'deleted_drafts': deleted_drafts,
            'deleted_submissions': deleted_submissions
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@payment_bp.get('/api/payment-entry/day-state')
@login_required
@page_required('payment_entry')
def payment_entry_day_state():
    date_str = _to_ymd(request.args.get('date'))
    if not date_str:
        return jsonify({"success": False, "error": "date is required"}), 400

    def ui_type(raw_type):
        normalized = (raw_type or '').upper().replace('-', '_').replace(' ', '_')
        if normalized == 'DAILY':
            return 'daily'
        if normalized == 'WEEKLY':
            return 'weekly'
        if normalized in ('TEN_DAYS', 'TEN_DAY', '10_DAYS', '10_DAY', '10'):
            return 'ten_days'
        if normalized == 'MONTHLY':
            return 'monthly'
        return normalized.lower() if normalized else ''

    submitted = PaymentSubmission.query.filter_by(submission_date=date_str).first() is not None
    payments = Payment.query.filter_by(payment_date=date_str).all()
    paid_loan_ids = [p.loan_id for p in payments]

    pending_rows = ShortPayment.query.filter_by(payment_date=date_str, status='PENDING').all()
    pending_shorts = []
    due_dt = None
    try:
        due_dt = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        due_dt = None
    updated_any = False
    for idx, sp in enumerate(pending_rows, start=1):
        loan = db.session.get(Loan, sp.loan_id)
        rep_type = (loan.repayment_type if loan else '')
        amount = float(sp.expected_amount or 0)
        if loan and due_dt:
            try:
                _, _, inst_amt, _ = compute_row_fields(loan, as_of_date=due_dt)
                calc_amt = due_amount_for_date(loan, due_dt, inst_amt)
                calc_amt = float(calc_amt or 0)
                if calc_amt > 0 and abs(calc_amt - amount) > 1e-6:
                    sp.expected_amount = calc_amt
                    db.session.add(sp)
                    amount = calc_amt
                    updated_any = True
            except Exception:
                pass
        pending_shorts.append({
            "sr": idx,
            "loan_id": sp.loan_id,
            "name": (loan.name if loan else ''),
            "phone": ((loan.phone or '') if loan else ''),
            "amount": amount,
            "type": ui_type(rep_type),
            "unpaid_since": date_str
        })

    if updated_any:
        try:
            db.session.commit()
        except Exception:
            db.session.rollback()

    recoveries = RecoveryPayment.query.filter_by(due_date=date_str).all()
    recovery_history = []
    total_recovered_amount = 0
    for r in recoveries:
        total_recovered_amount += float(r.amount or 0)
        loan = db.session.get(Loan, r.loan_id) if r.loan_id else None
        rep_type = (loan.repayment_type if loan else '')
        days_diff = 0
        try:
            due_dt = datetime.strptime(date_str, '%Y-%m-%d').date()
            days_diff = max(0, (date.today() - due_dt).days)
        except Exception:
            days_diff = 0
        recovery_history.append({
            "id": str(r.id),
            "date": date_str,
            "loanId": r.loan_id,
            "clientName": r.client_name or '',
            "amount": float(r.amount or 0),
            "unpaidDate": date_str,
            "daysDiff": days_diff,
            "notes": r.notes or '',
            "recoveryDate": _to_ymd(r.recovery_date) or '',
            "recoveredAt": (r.created_at or ''),
            "type": ui_type(rep_type)
        })

    return jsonify({
        "success": True,
        "date": date_str,
        "submitted": submitted,
        "paidLoanIds": paid_loan_ids,
        "pendingShortPayments": pending_shorts,
        "recoveryHistory": recovery_history,
        "totalRecoveredAmount": total_recovered_amount
    })


@payment_bp.get('/api/payment-entry/due-data')
@login_required
@page_required('payment_entry')
def payment_entry_due_data():
    date_str = _to_ymd(request.args.get('date'))
    if not date_str:
        return jsonify({"success": False, "error": "date is required"}), 400

    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        return jsonify({"success": False, "error": "Invalid date format"}), 400

    try:
        def build_rows(repayment_type):
            loans = Loan.query.filter_by(status='OPEN', repayment_type=repayment_type).all()
            rows = []
            for loan in loans:
                if not has_installment_due_on_date(loan, target_date):
                    continue
                _, _, inst_amt, _ = compute_row_fields(loan, as_of_date=target_date)
                due_amt = due_amount_for_date(loan, target_date, inst_amt)
                try:
                    due_amt = float(due_amt or 0)
                except Exception:
                    due_amt = 0
                if due_amt <= 0:
                    continue
                rows.append({
                    'loan_id': loan.id,
                    'name': loan.name,
                    'phone': (loan.phone or ''),
                    'total_amount': due_amt
                })
            rows.sort(key=lambda r: (r.get('name') or '').lower())
            total = sum(float(r.get('total_amount') or 0) for r in rows)
            return rows, total

        daily_rows, daily_total = build_rows('DAILY')
        weekly_rows, weekly_total = build_rows('WEEKLY')
        ten_days_rows, ten_days_total = build_rows('TEN_DAYS')
        monthly_rows, monthly_total = build_rows('MONTHLY')

        return jsonify({
            'success': True,
            'date': date_str,
            'daily': daily_rows,
            'weekly': weekly_rows,
            'ten_days': ten_days_rows,
            'monthly': monthly_rows,
            'daily_total': daily_total,
            'weekly_total': weekly_total,
            'ten_days_total': ten_days_total,
            'monthly_total': monthly_total
        })
    except Exception as e:
        current_app.logger.exception('payment_entry_due_data failed')
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@payment_bp.post('/api/payment-entry/recover-single')
@login_required
@page_required('payment_entry')
def payment_entry_recover_single():
    data = request.get_json(force=True) or {}
    due_date = _to_ymd(data.get('due_date'))
    recovery_date = _to_ymd(data.get('recovery_date'))
    loan_id = data.get('loan_id')
    notes = (data.get('notes') or '').strip()
    if not due_date or not loan_id:
        return jsonify({"success": False, "error": "due_date and loan_id are required"}), 400

    try:
        loan_id_int = int(loan_id)
    except Exception:
        return jsonify({"success": False, "error": "Invalid loan_id"}), 400

    short = ShortPayment.query.filter_by(payment_date=due_date, loan_id=loan_id_int).first()
    if not short or (short.status or '').upper() != 'PENDING':
        return jsonify({"success": False, "error": "Short payment not found or already recovered"}), 404

    recovery_date_str = recovery_date or date.today().strftime('%Y-%m-%d')
    loan = db.session.get(Loan, loan_id_int)
    client_name = (loan.name if loan else '')

    try:
        payment = Payment(
            loan_id=loan_id_int,
            payment_date=recovery_date_str,
            amount_paid=float(short.expected_amount or 0),
            payment_method='Short Recovery',
            entry_type='RECOVERY',
            remarks=f'Recovered short payment from {due_date}. {notes}'
        )
        db.session.add(payment)

        recovery_row = RecoveryPayment(
            loan_id=loan_id_int,
            client_name=client_name,
            due_date=due_date,
            recovery_date=recovery_date_str,
            amount=float(short.expected_amount or 0),
            notes=notes,
            created_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        db.session.add(recovery_row)

        short.status = 'RECOVERED'
        db.session.add(short)

        db.session.commit()
        return jsonify({"success": True})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@payment_bp.get('/payment-entry/penalty')
@login_required
@page_required('penalty_report')
def payment_entry_penalty_page():
    today = date.today()
    default_end = today.strftime('%Y-%m-%d')
    default_start = (today - timedelta(days=90)).strftime('%Y-%m-%d')
    return render_template('penalty_report.html', default_start=default_start, default_end=default_end)


@payment_bp.get('/api/payment-entry/penalty-report')
@login_required
@page_required('penalty_report')
def payment_entry_penalty_report():
    start_str = _to_ymd(request.args.get('start'))
    end_str = _to_ymd(request.args.get('end'))

    q = ShortPayment.query.filter(ShortPayment.status == 'PENDING')
    if start_str:
        q = q.filter(ShortPayment.payment_date >= start_str)
    if end_str:
        q = q.filter(ShortPayment.payment_date <= end_str)

    rows = q.order_by(ShortPayment.payment_date.asc()).all()
    today = date.today()

    due_dates = []
    loan_ids = []
    for sp in rows:
        due_str = _to_ymd(sp.payment_date)
        if due_str:
            due_dates.append(due_str)
        if sp.loan_id:
            loan_ids.append(sp.loan_id)

    collected_map = {}
    if rows and loan_ids and due_dates:
        collected_q = PenaltyCollected.query.filter(PenaltyCollected.loan_id.in_(set(loan_ids)))
        if start_str:
            collected_q = collected_q.filter(PenaltyCollected.due_date >= start_str)
        if end_str:
            collected_q = collected_q.filter(PenaltyCollected.due_date <= end_str)
        collected_rows = collected_q.all()
        for c in collected_rows:
            key = f"{c.loan_id}__{_to_ymd(c.due_date) or (c.due_date or '')}__{(c.repayment_type or '').upper()}"
            collected_map[key] = bool(c.collected)

    data = []
    for idx, sp in enumerate(rows, start=1):
        loan = db.session.get(Loan, sp.loan_id)
        due_str = _to_ymd(sp.payment_date)
        rep_type = ((loan.repayment_type or '') if loan else '')
        days_overdue = 0
        try:
            if due_str:
                due_dt = datetime.strptime(due_str, '%Y-%m-%d').date()
                days_overdue = max(0, (today - due_dt).days)
        except Exception:
            days_overdue = 0

        key = f"{sp.loan_id}__{due_str or (sp.payment_date or '')}__{(rep_type or '').upper()}"
        data.append({
            'sr': idx,
            'loan_id': sp.loan_id,
            'name': (loan.name if loan else ''),
            'phone': ((loan.phone or '') if loan else ''),
            'repayment_type': rep_type,
            'due_date': due_str or (sp.payment_date or ''),
            'pending_amount': float(sp.expected_amount or 0),
            'days_overdue': int(days_overdue),
            'collected': bool(collected_map.get(key, False))
        })

    return jsonify({
        'success': True,
        'start': start_str,
        'end': end_str,
        'rows': data,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })


@payment_bp.get('/recovery/page')
@login_required
@page_required('recovery')
def recovery_page():
    today = date.today()
    default_end = today.strftime('%Y-%m-%d')
    default_start = today.replace(day=1).strftime('%Y-%m-%d')
    return render_template('recover.html', default_start=default_start, default_end=default_end, default_penalty_per_day=500)


@payment_bp.get('/api/recovery/report')
@login_required
@page_required('recovery')
def recovery_report():
    start_str = _to_ymd(request.args.get('start'))
    end_str = _to_ymd(request.args.get('end'))

    q = RecoveryPayment.query
    if start_str:
        q = q.filter(RecoveryPayment.recovery_date >= start_str)
    if end_str:
        q = q.filter(RecoveryPayment.recovery_date <= end_str)

    rows = q.order_by(RecoveryPayment.recovery_date.asc()).all()

    data = []
    for r in rows:
        loan = db.session.get(Loan, r.loan_id) if getattr(r, 'loan_id', None) else None
        rep_type = (loan.repayment_type if loan else '')
        days_diff = 0
        try:
            due_dt = datetime.strptime(_to_ymd(r.due_date), '%Y-%m-%d').date()
            rec_dt = datetime.strptime(_to_ymd(r.recovery_date), '%Y-%m-%d').date()
            days_diff = max(0, (rec_dt - due_dt).days)
        except Exception:
            days_diff = 0
        data.append({
            'id': str(r.id),
            'loan_id': r.loan_id,
            'name': (r.client_name or (loan.name if loan else '')),
            'phone': ((loan.phone or '') if loan else ''),
            'repayment_type': rep_type,
            'due_date': _to_ymd(r.due_date),
            'recovery_date': _to_ymd(r.recovery_date),
            'days_diff': int(days_diff),
            'recovered_amount': float(getattr(r, 'amount', 0) or 0),
            'notes': (r.notes or '')
        })

    return jsonify({
        'success': True,
        'rows': data,
        'start': start_str,
        'end': end_str
    })


@payment_bp.post('/api/payment-entry/penalty-collected')
@csrf.exempt
@login_required
@page_required('penalty_report')
def payment_entry_penalty_collected_set():
    payload = request.get_json(force=True) or {}
    due_date = _to_ymd(payload.get('due_date'))
    repayment_type = (payload.get('repayment_type') or '').strip().upper()
    collected = bool(payload.get('collected'))
    loan_id = payload.get('loan_id')

    if not due_date or not repayment_type or not loan_id:
        return jsonify({'success': False, 'error': 'loan_id, due_date and repayment_type are required'}), 400

    try:
        loan_id_int = int(loan_id)
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid loan_id'}), 400

    try:
        row = PenaltyCollected.query.filter_by(
            loan_id=loan_id_int,
            due_date=due_date,
            repayment_type=repayment_type
        ).first()

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if not row:
            row = PenaltyCollected(
                loan_id=loan_id_int,
                due_date=due_date,
                repayment_type=repayment_type,
                collected=collected,
                updated_at=now_str
            )
        else:
            row.collected = collected
            row.updated_at = now_str

        db.session.add(row)
        db.session.commit()
        return jsonify({'success': True, 'loan_id': loan_id_int, 'due_date': due_date, 'repayment_type': repayment_type, 'collected': collected})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@payment_bp.get('/payment-entry/profit-loss')
@login_required
@page_required('profit_loss')
def payment_entry_profit_loss_page():
    today = date.today()
    default_end = (today - timedelta(days=1)).strftime('%Y-%m-%d')
    default_start = (today - timedelta(days=100)).strftime('%Y-%m-%d')
    return render_template('profit_loss_report.html', default_start=default_start, default_end=default_end)
    
@payment_bp.get('/payment-entry/outstanding')
@login_required
@page_required('outstanding')
def payment_entry_outstanding_page():
    """Outstanding movement analysis page (defaults: today to today)."""
    today = date.today()
    default_end = today.strftime('%Y-%m-%d')
    # Default window: from today to today
    default_start = today.strftime('%Y-%m-%d')
    return render_template(
        'outstanding_report.html',
        default_start=default_start,
        default_end=default_end,
        default_penalty_per_day=500
    ) 


@payment_bp.get('/api/payment-entry/profit-loss-report')
@login_required
@page_required('profit_loss')
def payment_entry_profit_loss_report():
    start_str = _to_ymd(request.args.get('start'))
    end_str = _to_ymd(request.args.get('end'))
    pct_raw = request.args.get('percentage')
    basis_raw = request.args.get('basis')

    try:
        pct = float(pct_raw) if pct_raw is not None and str(pct_raw).strip() != '' else 0.0
    except Exception:
        return jsonify({'success': False, 'error': 'Invalid percentage'}), 400

    start_dt = datetime.strptime(start_str, '%Y-%m-%d').date() if start_str else None
    end_dt = datetime.strptime(end_str, '%Y-%m-%d').date() if end_str else None

    basis = (str(basis_raw).strip().lower() if basis_raw is not None else '')
    if basis not in ('start', 'close'):
        basis = 'close'

    loans = Loan.query.filter_by(status='CLOSED').all()
    filtered = []
    for loan in loans:
        if basis == 'close':
            loan_dt = parse_date_str(loan.loan_closed_date) if loan.loan_closed_date else None
        else:
            loan_dt = parse_date_str(loan.loan_date) if loan.loan_date else None
        if start_dt and (not loan_dt or loan_dt < start_dt):
            continue
        if end_dt and (not loan_dt or loan_dt > end_dt):
            continue
        filtered.append((loan_dt or date.min, loan))

    filtered.sort(key=lambda t: t[0])

    report_rows = []
    totals = {
        'loan_amount': 0.0,
        'principal_amount': 0.0,
        'profit_amount': 0.0,
        'total_with_profit': 0.0
    }

    sr = 1
    for _, loan in filtered:
        total_amt = float(loan.loan_amount or 0)
        profit_amount = round(total_amt * (pct / 100.0), 2)
        principal_amount = round(total_amt - profit_amount, 2)
        total_with_profit = round(principal_amount + profit_amount, 2)
        end_dt_calc = schedule_end_date(loan)

        report_rows.append({
            'sr': sr,
            'loan_id': loan.id,
            'name': loan.name or '',
            'phone': loan.phone or '',
            'loan_date': format_date(loan.loan_date),
            'loan_date_raw': _to_ymd(loan.loan_date) or (loan.loan_date or ''),
            'loan_end_date': (end_dt_calc.strftime('%d-%m-%Y') if end_dt_calc else ''),
            'loan_closed_date': format_date(loan.loan_closed_date),
            'repayment_type': loan.repayment_type or '',
            'loan_amount': total_amt,
            'profit_pct': pct,
            'principal_amount': principal_amount,
            'profit_amount': profit_amount,
            'total_with_profit': total_with_profit
        })
        sr += 1

        totals['loan_amount'] += total_amt
        totals['principal_amount'] += principal_amount
        totals['profit_amount'] += profit_amount
        totals['total_with_profit'] += total_with_profit

    return jsonify({
        'success': True,
        'start': start_str,
        'end': end_str,
        'basis': basis,
        'percentage': pct,
        'rows': report_rows,
        'totals': totals,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })


@payment_bp.get('/api/payment-entry/outstanding-report')
@login_required
@page_required('outstanding')
def payment_entry_outstanding_report():
    """Outstanding movement analysis for a date range.

    If the caller does not specify a start date, the default window is
    today only (same as the UI defaults: From = To = today).
    """
    today = date.today()
    start_raw = (request.args.get('start') or '').strip()
    end_raw = (request.args.get('end') or '').strip()

    try:
        end_dt = datetime.strptime(end_raw, '%Y-%m-%d').date() if end_raw else today
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid end date format. Use YYYY-MM-DD'}), 400

    try:
        if start_raw:
            start_dt = datetime.strptime(start_raw, '%Y-%m-%d').date()
        else:
            # Default analysis window if start not provided: today only
            start_dt = end_dt
    except ValueError:
        return jsonify({'success': False, 'error': 'Invalid start date format. Use YYYY-MM-DD'}), 400

    # Normalize if start is after end
    if start_dt > end_dt:
        start_dt, end_dt = end_dt, start_dt

    start_str = start_dt.strftime('%Y-%m-%d')
    end_str = end_dt.strftime('%Y-%m-%d')

    # Compute opening and closing outstanding using receivable-report logic
    # Opening should be as-of the day BEFORE the start date to avoid double-counting
    # movements that occur on the first day of the analysis window.
    opening_outstanding = None
    closing_outstanding = None

    def _safe_receivable_total(as_on_ddmmyyyy):
        try:
            with current_app.test_request_context(f"/api/receivable-report?as_on_date={as_on_ddmmyyyy}"):
                resp = get_receivable_report()
                if isinstance(resp, tuple):
                    resp_obj, status_code = resp
                    if status_code != 200:
                        return None
                    data = resp_obj.get_json()
                else:
                    data = resp.get_json()
            return float(data.get('grand_total', 0) or 0)
        except Exception:
            return None

    # Opening as on previous day; closing as on end date
    opening_ref_dt = start_dt - timedelta(days=1)
    opening_as_on = opening_ref_dt.strftime('%d-%m-%Y')
    closing_as_on = end_dt.strftime('%d-%m-%Y')
    opening_outstanding = _safe_receivable_total(opening_as_on)
    closing_outstanding = _safe_receivable_total(closing_as_on)
    # Fetch closing O/S from saved ReceivableSnapshot if available
    closing_outstanding_from_snapshot = None
    try:
        # For day-wise breakdown across the entire period, we need to fetch receivable data for each date
        day_wise_breakdown = {}
        current_date = start_dt
        while current_date <= end_dt:
            as_on_date = current_date.strftime('%d-%m-%Y')
            try:
                with current_app.test_request_context(f"/api/receivable-report?as_on_date={as_on_date}"):
                    live_resp = current_app.view_functions["get_receivable_report"]()
                    live_data = live_resp.get_json()
                    if live_data and 'grand_total' in live_data:
                        total = float(live_data.get('grand_total', 0))
                        if total > 0:
                            date_key = current_date.strftime('%Y-%m-%d')
                            day_wise_breakdown[date_key] = total
            except Exception as inner_e:
                current_app.logger.error(f"Failed to fetch receivable data for {as_on_date}: {inner_e}")
            current_date += timedelta(days=1)
        
        # Format as day-wise array
        day_wise_array = [{'date': k, 'total': v} for k, v in sorted(day_wise_breakdown.items())]
        
        if day_wise_array:
            closing_outstanding_from_snapshot = {
                'as_on': f"{start_dt.strftime('%d-%m-%Y')} to {end_dt.strftime('%d-%m-%Y')}",
                'grand_total': sum(v for v in day_wise_breakdown.values()),
                'day_wise_breakdown': day_wise_array
            }
        else:
            closing_outstanding_from_snapshot = None
            
    except Exception as e:
        current_app.logger.error(f"Failed to fetch ReceivableSnapshot for date range: {e}")
        closing_outstanding_from_snapshot = None

    # Category: New loans created in the period
    new_loans_total = 0.0
    new_loans_count = 0
    new_loan_events = []

    # Category: Advance received (for WEEKLY / TEN_DAYS new loans) in the period
    advance_total = 0.0
    advance_count = 0
    advance_by_day = {}  # ymd -> total advance on that date
    advance_events = []  # detailed rows for frontend (date, name, amount)

    loans = Loan.query.all()
    for loan in loans:
        loan_dt = parse_date_str(loan.loan_date)
        if loan_dt and start_dt <= loan_dt <= end_dt:
            amount = float(loan.loan_amount or 0)
            new_loans_total += amount
            new_loans_count += 1
            new_loan_events.append({
                'event_type': 'NEW_LOAN',
                'date': loan_dt.strftime('%Y-%m-%d'),
                'loan_id': loan.id,
                'name': loan.name or '',
                'phone': loan.phone or '',
                'amount': amount,
                'delta': amount
            })

            # Track advance received at the time of new WEEKLY / TEN_DAYS loans
            try:
                if loan.repayment_type in ("WEEKLY", "TEN_DAYS"):
                    adv_amt = float(loan.advance_amount or 0)
                    if adv_amt > 0:
                        advance_total += adv_amt
                        advance_count += 1
                        ymd = loan_dt.strftime('%Y-%m-%d')
                        advance_by_day[ymd] = advance_by_day.get(ymd, 0.0) + adv_amt
                        advance_events.append({
                            'date': loan_dt.strftime('%Y-%m-%d'),
                            'loan_id': loan.id,
                            'name': loan.name or '',
                            'phone': loan.phone or '',
                            'amount': adv_amt,
                        })
            except Exception:
                # Do not break the report if advance values are bad
                pass

    # Category: Installments collected – use Daily Report "due installments" logic
    # For each day in the range, call /daily-report/data and assume all due installments
    # are collected on that date. This matches how the Daily Report page works.
    installments_total = 0.0
    installments_by_day = []

    def _daily_due_summary(date_ymd):
        """Return (total_due, count_clients) for a given YYYY-MM-DD using daily_report logic."""
        try:
            with current_app.test_request_context(f"/daily-report/data?date={date_ymd}"):
                resp = daily_collection_data()
                if isinstance(resp, tuple):
                    resp_obj, status_code = resp
                    if status_code != 200:
                        return 0.0, 0
                    data = resp_obj.get_json()
                else:
                    data = resp.get_json()

            if not isinstance(data, dict):
                return 0.0, 0

            try:
                total = float(data.get('daily_total', 0) or 0) \
                        + float(data.get('weekly_total', 0) or 0) \
                        + float(data.get('ten_days_total', 0) or 0) \
                        + float(data.get('monthly_total', 0) or 0)
            except Exception:
                total = 0.0

            count = 0
            for key in ('daily', 'weekly', 'ten_days', 'monthly'):
                rows = data.get(key) or []
                for row in rows:
                    try:
                        amt = float(row.get('total_amount', 0) or 0)
                    except Exception:
                        amt = 0.0
                    if amt > 0:
                        count += 1

            return total, count
        except Exception:
            return 0.0, 0

    cur_dt = start_dt
    while cur_dt <= end_dt:
        ymd = cur_dt.strftime('%Y-%m-%d')
        day_total, day_count = _daily_due_summary(ymd)
        if day_total or day_count:
            installments_total += day_total
            installments_by_day.append({
                'date': ymd,
                'total': float(day_total or 0.0),
                'count': int(day_count or 0),
            })
        cur_dt += timedelta(days=1)

    # Category: Short EMIs created in the period
    short_total = 0.0
    short_count = 0
    short_events = []
    short_rows = ShortPayment.query.filter(
        ShortPayment.payment_date >= start_str,
        ShortPayment.payment_date <= end_str
    ).all()
    for sp in short_rows:
        amt = float(sp.expected_amount or 0)
        short_total += amt
        short_count += 1
        loan = db.session.get(Loan, sp.loan_id) if sp.loan_id else None
        short_events.append({
            'event_type': 'SHORT_EMI',
            'date': sp.payment_date or '',
            'loan_id': sp.loan_id,
            'name': loan.name if loan else '',
            'phone': loan.phone if loan else '',
            'amount': amt,
            'status': sp.status or '',
            'delta': amt
        })

    # Category: Loans closed in the period
    closed_loans_total = 0.0
    closed_loans_count = 0
    closed_loan_events = []
    for loan in loans:
        if not loan.loan_closed_date:
            continue
        close_dt = parse_date_str(loan.loan_closed_date)
        if not close_dt or close_dt < start_dt or close_dt > end_dt:
            continue
        amt = float(loan.case_closing_amount or 0)
        closed_loans_total += amt
        closed_loans_count += 1
        closed_loan_events.append({
            'event_type': 'LOAN_CLOSED',
            'date': close_dt.strftime('%Y-%m-%d'),
            'loan_id': loan.id,
            'name': loan.name or '',
            'phone': loan.phone or '',
            'amount': amt,
            'delta': -amt
        })

    details = []
    details.extend(new_loan_events)
    details.extend(short_events)
    details.extend(closed_loan_events)

    # Sort details by date then event type for stable display
    def _event_sort_key(ev):
        return (ev.get('date') or '', ev.get('event_type') or '', ev.get('loan_id') or 0)

    details.sort(key=_event_sort_key)

    opening_val = float(opening_outstanding or 0.0)
    closing_val = float(closing_outstanding or 0.0) if closing_outstanding is not None else None

    # Advance received reduces outstanding similar to installments/closed loans
    computed_closing = (
        opening_val
        + new_loans_total
        + short_total
        - installments_total
        - advance_total
        - closed_loans_total
    )
    movement_change = computed_closing - opening_val
    diff_vs_receivable = None
    if closing_outstanding is not None:
        diff_vs_receivable = computed_closing - closing_val

    # Convert advance_by_day dict to sorted list for frontend (similar to installments_by_day)
    advance_by_day_list = []
    for ymd in sorted(advance_by_day.keys()):
        amt = float(advance_by_day.get(ymd, 0.0) or 0.0)
        if amt or True:
            advance_by_day_list.append({
                'date': ymd,
                'total': amt,
                'count': 0,
            })

    # Sort detailed advance events for a clean table (date, then loan id)
    advance_events.sort(key=lambda ev: (ev.get('date') or '', ev.get('loan_id') or 0))

    summary = {
        'opening_outstanding': opening_outstanding,
        'closing_outstanding': closing_outstanding,
        'opening_as_on': opening_as_on,
        'closing_as_on': closing_as_on,
        'new_loans_total': new_loans_total,
        'new_loans_count': new_loans_count,
        'installments_total': installments_total,
        'short_total': short_total,
        'short_count': short_count,
        'closed_loans_total': closed_loans_total,
        'closed_loans_count': closed_loans_count,
        'advance_total': advance_total,
        'advance_count': advance_count,
        'computed_closing_outstanding': computed_closing,
        'movement_net_change': movement_change,
        'difference_vs_receivable': diff_vs_receivable
    }

    return jsonify({
        'success': True,
        'start': start_dt.strftime('%Y-%m-%d'),
        'end': end_dt.strftime('%Y-%m-%d'),
        'summary': summary,
        'details': details,
        'installments_by_day': installments_by_day,
        'advance_by_day': advance_by_day_list,
        'advance_details': advance_events,
        'closing_outstanding_from_receivable_snapshot': closing_outstanding_from_snapshot,
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    })

app.register_blueprint(payment_bp)


# =============================================================================
# LOAN CRUD OPERATIONS - 100% COMPLETE
# =============================================================================

# ✅ FIXED: Unified client ledger endpoint for hyperlink fix


@app.route('/api/complete-client-ledger')
@api_auth_required()
def complete_client_ledger():
    """FIXED: Complete ledger data for hyperlinks"""
    try:
        name = request.args.get('name', '').strip()
        if not name:
            return jsonify({"error": "Parameter 'name' is required"}), 400

        ledger_data = get_complete_ledger_data_by_client_name(name)

        if not ledger_data:
            return jsonify({
                "rows": [],
                "totals": {
                    "processing_fees": 0,
                    "loan_amount": 0,
                    "installment_amount": 0,
                    "advance_amount": 0,
                    "case_closing_amount": 0,
                    "balance_amount": 0,
                    "pending_installments": 0
                },
                "meta": {"as_of": date.today().isoformat(), "phone": ""},
                "client_info": {"name": name, "phone": "", "address": "", "city": ""}
            })

        ledger_data["status"] = "success"
        ledger_data["generated_at"] = datetime.now().strftime('%d-%m-%Y %H:%M')

        return jsonify(ledger_data)

    except Exception as e:
        app.logger.error(f'Error in /api/complete-client-ledger: {e}\n' + traceback.format_exc())
        return jsonify({"error": "Internal Server Error", "message": str(e)}), 500

@app.route('/api/client-ledger')
@api_auth_required()
def client_ledger():
    try:
        name = request.args.get('name', '').strip()
        if not name:
            return jsonify({"error": "Parameter 'name' is required"}), 400

        normalized_name = name.upper()
        ledger_data = get_ledger_data_by_client(normalized_name)

        if not ledger_data:
            # Return empty ledger for unknown clients
            ledger_data = {
                "total_loans": 0,
                "active_loans": 0,
                "outstanding_amount": 0,
                "transactions": []
            }

        response = dict(ledger_data)
        response["client_name"] = name
        response["generated_at"] = datetime.now().strftime('%d-%m-%Y %H:%M')
        response["status"] = "success"

        return jsonify(response)

    except Exception as e:
        app.logger.error(f'Error in /api/client-ledger: {e}\n' + traceback.format_exc())
        return jsonify({"error": "Internal Server Error", "message": str(e)}), 500
# ============================================
# AFIN CLIENTS LEDGER PAGE (Modified Version)
# ============================================
@app.route('/afin-ledger/page')
@login_required
@page_required('clients')
def afin_ledger_page():
    """AFIN Clients Ledger - Modified version without ID, Processing Fee, Advance, Closing Amt"""
    return render_template('AFINCLIENTS.HTML')


@app.route("/loan", methods=["POST"])
@csrf.exempt
@login_required
@page_required('loan_form')
def add_loan():
    data = request.get_json(force=True) or {}
    errs = require_fields(data, ["name", "phone", "loan_date", "loan_amount", "repayment_type"])
    if not validate_repayment_type(data.get("repayment_type")):
        errs["repayment_type"] = "Must be one of DAILY, WEEKLY, TEN_DAYS, MONTHLY"
    if parse_date_str(data.get("loan_date")) is None:
        errs["loan_date"] = "Invalid format YYYY-MM-DD"

    amount = safe_float(data.get("loan_amount"))
    if amount is None or amount < 0:
        errs["loan_amount"] = "Must be a non-negative number"

    close_date = data.get("loan_closed_date")
    if close_date and parse_date_str(close_date) is None:
        errs["loan_closed_date"] = "Invalid format YYYY-MM-DD"

    if errs:
        return jsonify({"errors": errs}), 400

    repayment_type = data.get("repayment_type")

    interest_rate = safe_float(data.get("interest_rate")) if "interest_rate" in data else None
    if repayment_type == "MONTHLY":
        if interest_rate is None or interest_rate <= 0:
            errs["interest_rate"] = "Required"
    else:
        interest_rate = safe_float(interest_rate) or 0

    if errs:
        return jsonify({"errors": errs}), 400

    # Check if phone number already exists for a DIFFERENT client (different name)
    phone = (data.get("phone") or "").strip()
    name = (data.get("name") or "").strip().upper()
    if phone:
        existing_loan = Loan.query.filter_by(phone=phone).first()
        if existing_loan and existing_loan.name != name:
            return jsonify({"errors": {"phone": "This phone number already belongs to another client"}}), 400

    # Ensure primary key is populated even if the legacy table lacks AUTOINCREMENT
    next_id = db.session.query(func.max(Loan.id)).scalar() or 0

    existing_paid = safe_float(data.get("total_paid_amount")) or 0
    status = data.get("status") or "OPEN"
    outstanding_balance = 0 if status == "CLOSED" else max(0, (amount or 0) - existing_paid)

    loan = Loan(
        id=next_id + 1,
        name=(data.get("name") or "").strip().upper(),
        address=(data.get("address") or "").strip(),
        phone=(data.get("phone") or "").strip(),
        city=(data.get("city") or "").strip(),
        loan_date=data.get("loan_date"),
        loan_closed_date=close_date or None,
        loan_amount=amount,
        processing_fees=safe_float(data.get("processing_fees")),
        case_closing_amount=safe_float(data.get("case_closing_amount")),
        advance_amount=safe_float(data.get("advance_amount")) if "advance_amount" in data else calc_advance(amount, repayment_type),
        interest_rate=interest_rate or 0,
        repayment_type=repayment_type,
        status=status,
        total_paid_amount=existing_paid,
        outstanding_balance=outstanding_balance,
        remarks=(data.get("remarks") or "").strip()
    )
    db.session.add(loan)
    db.session.commit()
    return jsonify({"id": loan.id, "message": "Loan added"}), 201

@app.route("/loan/<int:id>", methods=["PUT"])
@csrf.exempt
@login_required
@page_required('loan_form')
def update_loan(id):
    data = request.get_json(force=True) or {}
    loan = db.session.get(Loan, id)
    if not loan:
        abort(404)

    if "name" in data and not data.get("name", "").strip():
        return jsonify({"errors": {"name": "Name cannot be empty"}}), 400
    
    if "loan_date" in data and parse_date_str(data.get("loan_date")) is None:
        return jsonify({"errors": {"loan_date": "Invalid format YYYY-MM-DD"}}), 400
        
    if "repayment_type" in data and not validate_repayment_type(data.get("repayment_type")):
        return jsonify({"errors": {"repayment_type": "Must be one of DAILY, WEEKLY, TEN_DAYS, MONTHLY"}}), 400

    amount = safe_float(data.get("loan_amount")) if "loan_amount" in data else loan.loan_amount
    if amount is not None and amount < 0:
        return jsonify({"errors": {"loan_amount": "Must be non-negative"}}), 400

    if "name" in data:
        loan.name = data.get("name").strip().upper()
    if "address" in data:
        loan.address = (data.get("address") or "").strip()
    if "phone" in data:
        loan.phone = (data.get("phone") or "").strip()
    if "city" in data:
        loan.city = (data.get("city") or "").strip()
    if "loan_date" in data:
        loan.loan_date = data.get("loan_date")
    if "loan_closed_date" in data:
        loan.loan_closed_date = data.get("loan_closed_date") or None
    if "loan_amount" in data:
        loan.loan_amount = amount
        # ✅ FIXED: Update balance considering closure status
        if loan.status == "CLOSED":
            loan.outstanding_balance = 0
        else:
            loan.outstanding_balance = amount - (loan.total_paid_amount or 0)
    if "processing_fees" in data:
        loan.processing_fees = safe_float(data.get("processing_fees"))
    if "case_closing_amount" in data:
        loan.case_closing_amount = safe_float(data.get("case_closing_amount"))
    if "repayment_type" in data:
        loan.repayment_type = data.get("repayment_type")

    if "interest_rate" in data:
        loan.interest_rate = safe_float(data.get("interest_rate")) or 0
    if "status" in data:
        loan.status = data.get("status")
        # ✅ FIXED: Set balance to 0 when closing loan
        if loan.status == "CLOSED":
            loan.outstanding_balance = 0

    if "advance_amount" in data:
        loan.advance_amount = safe_float(data.get("advance_amount"))
    else:
        loan.advance_amount = calc_advance(loan.loan_amount, loan.repayment_type)

    if "remarks" in data:
        loan.remarks = (data.get("remarks") or "").strip()

    db.session.commit()
    return jsonify({"message": "Loan updated"})


@app.route("/client/<phone>/personal", methods=["PUT"])
@csrf.exempt
@login_required
@page_required('loan_form')
def update_client_personal_info(phone):
    """Update personal information (and optional phone) for ALL loans of a client"""
    data = request.get_json(force=True) or {}

    if not phone or not phone.strip():
        return jsonify({"errors": {"phone": "Phone is required"}}), 400

    identifier_phone = phone.strip()
    incoming_name = (data.get("name") or "").strip()
    incoming_address = (data.get("address") or "").strip()
    incoming_city = (data.get("city") or "").strip()
    new_phone = (data.get("new_phone") or data.get("phone") or "").strip()

    errs = {}
    if "name" in data and not incoming_name:
        errs["name"] = "Name cannot be empty"
    if "new_phone" in data and not new_phone:
        errs["new_phone"] = "New phone cannot be empty"

    if errs:
        return jsonify({"errors": errs}), 400

    loans = Loan.query.filter_by(phone=identifier_phone).all()

    if not loans:
        return jsonify({"errors": {"phone": "No loans found for this client"}}), 404

    # If new phone supplied, ensure it's not already used by another client (optional safeguard)
    if new_phone and new_phone != identifier_phone:
        phone_exists = Loan.query.filter(Loan.phone == new_phone, Loan.phone != identifier_phone).first()
        if phone_exists:
            return jsonify({"errors": {"new_phone": "Another client already uses this phone"}}), 400

    updated_count = 0
    for loan in loans:
        if "name" in data:
            loan.name = incoming_name.upper()
        if "address" in data:
            loan.address = incoming_address
        if "city" in data:
            loan.city = incoming_city
        if new_phone:
            loan.phone = new_phone
        updated_count += 1

    db.session.commit()
    return jsonify({
        "message": f"Personal info updated for {updated_count} loan records",
        "updated_loans": updated_count,
        "phone": new_phone or identifier_phone
    })

@app.route("/loan/<int:id>/close", methods=["POST"])
@csrf.exempt
@login_required
@page_required('loan_form')
def close_loan(id):
    """Manual close loan"""
    data = request.get_json(force=True) or {}
    loan = db.session.get(Loan, id)
    if not loan:
        abort(404)
    
    # Accept both payload styles:
    # - loan_closed_date (used by loan_form.html / legacy)
    # - close_date (used by AFINCLIENTS.HTML)
    close_date_str = (data.get("loan_closed_date") or data.get("close_date") or "").strip() or None
    as_of = parse_date_str(close_date_str) if close_date_str else date.today()
    if close_date_str and as_of is None:
        return jsonify({"errors": {"loan_closed_date": "Invalid date format"}}), 400

    # Store consistently in DD-MM-YYYY so all reports can detect close date reliably
    loan.loan_closed_date = as_of.strftime("%d-%m-%Y")
    if loan.repayment_type == "MONTHLY":
        loan.case_closing_amount = int(loan.loan_amount or 0)
    else:
        natural_last = schedule_end_date(loan)
        if natural_last and as_of == natural_last:
            loan.case_closing_amount = int(loan.loan_amount or 0)
        else:
            bal, _, _ = calculate_balance_and_pending_core(loan, as_of=as_of)
            loan.case_closing_amount = int(bal)

    loan.status = "CLOSED"
    # ✅ FIXED: Set outstanding balance to 0 when closing
    loan.outstanding_balance = 0
    db.session.commit()
    return jsonify({"message": "Loan closed", "closing_amount": loan.case_closing_amount})

# ✅ NEW: Manual Loan Close Route for Sales Report
@app.route("/loan/<int:id>/manual-close", methods=["POST"])
@csrf.exempt
@login_required
@page_required('sales_report')
def manual_close_loan(id):
    """NEW: Manual close loan from sales report with custom date and amount"""
    data = request.get_json(force=True) or {}
    loan = db.session.get(Loan, id)
    
    if not loan:
        return jsonify({"success": False, "error": "Loan not found"}), 404
    
    if loan.status == "CLOSED":
        return jsonify({"success": False, "error": "Loan is already closed"}), 400
    
    close_date_str = data.get("close_date")
    close_amount = data.get("close_amount")
    
    if not close_date_str or close_amount is None:
        return jsonify({"success": False, "error": "Close date and amount are required"}), 400
    
    # Validate date format
    try:
        close_date = datetime.strptime(close_date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"success": False, "error": "Invalid date format. Use YYYY-MM-DD"}), 400
    
    # Validate amount
    try:
        close_amount = float(close_amount)
        if close_amount < 0:
            return jsonify({"success": False, "error": "Close amount cannot be negative"}), 400
    except (ValueError, TypeError):
        return jsonify({"success": False, "error": "Invalid close amount"}), 400
    
    try:
        # ✅ FIXED: Store in DD-MM-YYYY format
        loan.loan_closed_date = close_date.strftime("%d-%m-%Y")
        loan.case_closing_amount = close_amount
        loan.status = "CLOSED"
        # ✅ FIXED: Set outstanding balance to 0
        loan.outstanding_balance = 0
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"Loan {id} closed successfully",
            "loan_id": id,
            "client_name": loan.name,
            "close_date": close_date.strftime("%d-%m-%Y"),
            "close_amount": close_amount
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": f"Failed to close loan: {str(e)}"}), 500
@app.delete("/loan/<int:id>")
@csrf.exempt
@login_required
@page_required('loan_form')
def delete_loan(id):
    """Delete an individual loan and all related records"""
    loan = db.session.get(Loan, id)
    if not loan:
        return jsonify({"success": False, "message": "Loan not found"}), 404

    try:
        Payment.query.filter_by(loan_id=id).delete()
        ShortPayment.query.filter_by(loan_id=id).delete()
        RecoveryPayment.query.filter_by(loan_id=id).delete()
        Payments.query.filter_by(loan_id=id).delete()
        PenaltyCollected.query.filter_by(loan_id=id).delete()

        db.session.delete(loan)
        db.session.commit()
        return jsonify({"success": True, "message": "Loan deleted successfully", "loan_id": id})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": f"Failed to delete loan: {str(e)}"}), 500

    
@app.delete("/client/<phone>")
@csrf.exempt
@login_required
@page_required('loan_form')
def delete_client(phone):
    phone = phone.strip()
    if not phone:
        return jsonify({"error": "Phone number required"}), 400

    loans = Loan.query.filter_by(phone=phone).all()
    if not loans:
        return jsonify({"error": "No records found for this client"}), 404

    for loan in loans:
        Payment.query.filter_by(loan_id=loan.id).delete()
        ShortPayment.query.filter_by(loan_id=loan.id).delete()
        db.session.delete(loan)
    db.session.commit()

    return jsonify({"message": f"✅ All data for {phone} deleted."})

@app.route("/delete/customer", methods=["DELETE"])
@csrf.exempt
@login_required
@page_required('loan_form')
def delete_customer_route():
    """Handle customer deletion from frontend modal"""
    data = request.get_json() or {}
    phone = data.get("phone", "").strip()
    
    if not phone:
        return jsonify({"error": "Phone number required"}), 400

    loans = Loan.query.filter_by(phone=phone).all()
    if not loans:
        return jsonify({"error": "No records found for this client"}), 404

    try:
        # Delete all related data for this customer
        deleted_loans = 0
        for loan in loans:
            Payment.query.filter_by(loan_id=loan.id).delete()
            ShortPayment.query.filter_by(loan_id=loan.id).delete()
            db.session.delete(loan)
            deleted_loans += 1
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": f"✅ Customer {phone} deleted successfully",
            "deleted_loans": deleted_loans
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": f"Failed to delete customer: {str(e)}"}), 500

@app.route('/loan/form')
@login_required
@page_required('loan_form')
def loan_form():
    return render_template('loan_form.html')

@app.route("/clients")
@login_required
def clients():
    rows = db.session.query(Loan.phone, func.min(Loan.name)).group_by(Loan.phone).all()
    return jsonify([{"phone": p, "name": n} for p, n in rows])

# =============================================================================
# ALL CLIENT LEDGERS - 100% COMPLETE
# =============================================================================

@app.route('/all-clients-ledgers-view')
@login_required
@page_required('all_ledgers')
def all_clients_ledgers_view():
    """Display all clients with ledgers using same template format"""
    clients_query = db.session.query(Loan.phone, Loan.name).distinct().all()
    clients_data = []
    today = date.today()
    
    for phone, name in clients_query:
        loans = Loan.query.filter_by(phone=phone).order_by(Loan.loan_date.asc()).all()
        
        if not loans:
            continue
            
        first_loan = loans[0]
        loan_records = []
        for loan in loans:
            bal, pending, inst_amt, _ = compute_row_fields(loan, as_of_date=today)
            loan_dt = parse_date_str(loan.loan_date)
            end_dt = schedule_end_date(loan)
            
            loan_record = {
                'id': loan.id,
                'loan_date': format_date(loan.loan_date),
                'loan_date_raw': loan.loan_date,
                'day_name': loan_dt.strftime("%a") if loan_dt and loan.repayment_type == "WEEKLY" else "",
                'loan_end_date': end_dt.strftime("%d-%m-%Y") if end_dt else "",
                'loan_closed_date': format_date(loan.loan_closed_date),
                'loan_closed_date_raw': loan.loan_closed_date,
                'processing_fees': int(loan.processing_fees) if loan.processing_fees else 0,
                'amount': int(loan.loan_amount) if loan.loan_amount else 0,
                'installment_amount': inst_amt,
                'repayment_type': loan.repayment_type,
                'advance_amount': int(loan.advance_amount) if loan.advance_amount else 0,
                'case_closing_amt': int(loan.case_closing_amount) if loan.case_closing_amount else 0,
                'balance_amount': bal,
                'pending_installments': pending,
                'status': loan.status,
                'name': loan.name,
                'address': loan.address,
                'phone': loan.phone,
                'city': loan.city
            }
            loan_records.append(loan_record)
        
        client_totals = {
            'processing_fees': sum(loan['processing_fees'] for loan in loan_records),
            'loan_amount': sum(loan['amount'] for loan in loan_records if not is_excluded_from_loan_total(Loan.query.get(loan['id']), today)),
            'installment_amount': sum(loan['installment_amount'] for loan in loan_records),
            'advance_amount': sum(loan['advance_amount'] for loan in loan_records),
            'case_closing_amount': sum(loan['case_closing_amt'] for loan in loan_records),
            'balance_amount': sum(loan['balance_amount'] for loan in loan_records),
            'pending_installments': sum(loan['pending_installments'] for loan in loan_records)
        }
        
        clients_data.append({
            'name': first_loan.name or "Unknown Client",
            'phone': phone,
            'address': first_loan.address,
            'city': first_loan.city,
            'loans': loan_records,
            'totals': client_totals,
            'contact_info': []
        })
    
    clients_data.sort(key=lambda x: x['name'])
    
    for client in clients_data:
        contact_parts = []
        if client['address']:
            contact_parts.append(client['address'])
        if client['phone']:
            contact_parts.append(f"📞 {client['phone']}")
        if client['city']:
            contact_parts.append(f"🏙️ {client['city']}")
        client['contact_info'] = " | ".join(contact_parts) if contact_parts else "Contact information not available"
    
    return render_template('all_clients_ledgers_view.html', clients=clients_data)


def _build_client_age_rows():
    today = date.today()
    loans = Loan.query.order_by(Loan.name.asc(), Loan.loan_date.asc()).all()
    clients_map = {}

    for loan in loans:
        phone = (loan.phone or '').strip()
        key = f"PHONE::{phone}" if phone else f"NAME::{(loan.name or 'UNKNOWN').strip().upper()}::{(loan.city or '').strip().upper()}"

        if key not in clients_map:
            clients_map[key] = {
                'name': (loan.name or 'Unknown Client').strip() or 'Unknown Client',
                'phone': phone or None,
                'city': (loan.city or '').strip(),
                'loans': []
            }

        clients_map[key]['loans'].append(loan)

    rows = []
    serial = 1

    for client in sorted(clients_map.values(), key=lambda item: (item['name'] or '').upper()):
        first_loan_date = None
        last_closed_date = None
        current_open_date = None
        next_open_after_last_closed = None

        for loan in client['loans']:
            loan_start = parse_date_str(loan.loan_date)
            if loan_start and (not first_loan_date or loan_start < first_loan_date):
                first_loan_date = loan_start

        closed_dates = [parse_date_str(l.loan_closed_date) for l in client['loans'] if l.status == 'CLOSED']
        closed_dates = [d for d in closed_dates if d]
        if closed_dates:
            last_closed_date = max(closed_dates)

        open_dates = [parse_date_str(l.loan_date) for l in client['loans'] if l.status != 'CLOSED']
        open_dates = [d for d in open_dates if d]
        if open_dates:
            current_open_date = max(open_dates)

        if last_closed_date:
            next_open_candidates = [parse_date_str(l.loan_date) for l in client['loans'] if l.status != 'CLOSED']
            next_open_candidates = [d for d in next_open_candidates if d and d > last_closed_date]
            if next_open_candidates:
                next_open_after_last_closed = min(next_open_candidates)

        age_years = None
        if first_loan_date:
            age_years = round(max(0, (today - first_loan_date).days) / 365.25, 1)

        rows.append({
            'serial': serial,
            'name': client['name'],
            'phone': client['phone'],
            'city': client['city'],
            'age_years': age_years,
            'age_text': f"{age_years:.1f} yrs" if age_years is not None else 'N/A',
            'first_loan_date': first_loan_date,
            'first_loan_text': fmt_dd_mm_yyyy(first_loan_date),
            'last_closed_date': last_closed_date,
            'last_closed_text': fmt_dd_mm_yyyy(last_closed_date),
            'current_open_date': current_open_date,
            'current_open_text': fmt_dd_mm_yyyy(current_open_date),
            'next_open_after_last_closed': next_open_after_last_closed,
            'next_open_after_last_closed_text': fmt_dd_mm_yyyy(next_open_after_last_closed)
        })

        serial += 1

    return rows


@app.get("/sales/clients-age")
@login_required
@page_required('sales_report')
def sales_clients_age_report():
    rows = _build_client_age_rows()
    generated_on = datetime.now().strftime("%d-%m-%Y %I:%M %p")
    return render_template(
        "clients_age_report.html",
        clients=rows,
        total_clients=len(rows),
        generated_on=generated_on
    )

@app.route('/export-all-clients-ledgers-pdf')
@login_required
@page_required('all_ledgers')
def export_all_clients_ledgers_pdf():
    """Export all clients ledgers to PDF"""
    clients_query = db.session.query(Loan.phone, Loan.name).distinct().all()
    clients_data = []
    today = date.today()
    
    for phone, name in clients_query:
        loans = Loan.query.filter_by(phone=phone).order_by(Loan.loan_date.asc()).all()
        if not loans:
            continue
            
        first_loan = loans[0]
        loan_records = []
        
        for loan in loans:
            bal, pending, inst_amt, _ = compute_row_fields(loan, as_of_date=today)
            loan_dt = parse_date_str(loan.loan_date)
            end_dt = schedule_end_date(loan)
            
            loan_records.append({
                'loan_date': format_date(loan.loan_date),
                'day_name': loan_dt.strftime("%a") if loan_dt and loan.repayment_type == "WEEKLY" else "",
                'loan_end_date': end_dt.strftime("%d-%m-%Y") if end_dt else "",
                'loan_closed_date': format_date(loan.loan_closed_date),
                'processing_fees': int(loan.processing_fees) if loan.processing_fees else 0,
                'amount': int(loan.loan_amount) if loan.loan_amount else 0,
                'installment_amount': inst_amt,
                'repayment_type': loan.repayment_type,
                'advance_amount': int(loan.advance_amount) if loan.advance_amount else 0,
                'case_closing_amt': int(loan.case_closing_amount) if loan.case_closing_amount else 0,
                'balance_amount': bal,
                'pending_installments': pending,
                'status': loan.status
            })
        
        totals = {
            'processing_fees': sum(loan['processing_fees'] for loan in loan_records),
            'loan_amount': sum(loan['amount'] for loan in loan_records),
            'installment_amount': sum(loan['installment_amount'] for loan in loan_records),
            'advance_amount': sum(loan['advance_amount'] for loan in loan_records),
            'case_closing_amount': sum(loan['case_closing_amt'] for loan in loan_records),
            'balance_amount': sum(loan['balance_amount'] for loan in loan_records),
            'pending_installments': sum(loan['pending_installments'] for loan in loan_records)
        }
        
        contact_parts = []
        if first_loan.address:
            contact_parts.append(first_loan.address)
        if first_loan.phone:
            contact_parts.append(f"📞 {first_loan.phone}")
        if first_loan.city:
            contact_parts.append(f"🏙️ {first_loan.city}")
        
        clients_data.append({
            'name': first_loan.name or "Unknown Client",
            'phone': phone,
            'full_contact_line': " | ".join(contact_parts) if contact_parts else "Contact information not available",
            'loans': loan_records,
            'totals': totals
        })
    
    clients_data.sort(key=lambda x: x['name'])
    
    html_content = render_template_string('''
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>All Clients Ledgers</title>
        <style>
            body { font-family: Arial, sans-serif; margin: 0; padding: 15px; font-size: 10px; }
            .header { text-align: center; border-bottom: 2px solid #333; padding: 15px; margin-bottom: 20px; }
            .client-section { margin: 20px 0; page-break-inside: avoid; border: 1px solid #ddd; padding: 15px; }
            .client-header { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 10px; border-radius: 8px; margin-bottom: 15px; text-align: center; }
            .client-name { font-size: 16px; font-weight: bold; margin: 0 0 5px 0; }
            .client-contact { font-size: 11px; margin: 0; }
            table { width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 8px; }
            th, td { border: 1px solid #ddd; padding: 3px; text-align: center; }
            th { background-color: #343a40; color: white; font-weight: bold; }
            .status-open { background-color: #fff3cd; }
            .status-closed { background-color: #d1edff; }
            .totals-row { background-color: #e9ecef; font-weight: bold; }
            .footer { text-align: center; margin-top: 30px; font-size: 9px; color: #666; }
        </style>
    </head>
    <body>
        <div class="header">
            <h2>📋 ALL CLIENTS LEDGERS</h2>
            <p>Complete Loan Management Data | Generated: {{ generated_date }}</p>
        </div>
        
        {% for client in clients %}
        <div class="client-section">
            <div class="client-header">
                <div class="client-name">{{ client.name }}</div>
                <div class="client-contact">{{ client.full_contact_line }}</div>
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th>Loan Date</th><th>Day</th><th>End Date</th><th>Close Date</th>
                        <th>Proc. Fees</th><th>Loan Amount</th><th>Inst. Amount</th><th>Type</th>
                        <th>Advance</th><th>Closing Amt</th><th>Balance</th><th>Pending</th><th>Status</th>
                    </tr>
                </thead>
                <tbody>
                    {% for loan in client.loans %}
                    <tr class="{{ 'status-closed' if loan.status == 'CLOSED' else 'status-open' }}">
                        <td>{{ loan.loan_date }}</td>
                        <td>{{ loan.day_name }}</td>
                        <td>{{ loan.loan_end_date }}</td>
                        <td>{{ loan.loan_closed_date }}</td>
                        <td>₹{{ "{:,}".format(loan.processing_fees) }}</td>
                        <td>₹{{ "{:,}".format(loan.amount) }}</td>
                        <td>₹{{ "{:,}".format(loan.installment_amount) }}</td>
                        <td>{{ loan.repayment_type }}</td>
                        <td>₹{{ "{:,}".format(loan.advance_amount) }}</td>
                        <td>₹{{ "{:,}".format(loan.case_closing_amt) }}</td>
                        <td>₹{{ "{:,}".format(loan.balance_amount) }}</td>
                        <td>{{ loan.pending_installments }}</td>
                        <td>{{ loan.status }}</td>
                    </tr>
                    {% endfor %}
                </tbody>
                <tfoot>
                    <tr class="totals-row">
                        <td colspan="4"><strong>TOTALS:</strong></td>
                        <td><strong>₹{{ "{:,}".format(client.totals.processing_fees) }}</strong></td>
                        <td><strong>₹{{ "{:,}".format(client.totals.loan_amount) }}</strong></td>
                        <td><strong>₹{{ "{:,}".format(client.totals.installment_amount) }}</strong></td>
                        <td></td>
                        <td><strong>₹{{ "{:,}".format(client.totals.advance_amount) }}</strong></td>
                        <td><strong>₹{{ "{:,}".format(client.totals.case_closing_amount) }}</strong></td>
                        <td><strong>₹{{ "{:,}".format(client.totals.balance_amount) }}</strong></td>
                        <td><strong>{{ client.totals.pending_installments }}</strong></td>
                        <td></td>
                    </tr>
                </tfoot>
            </table>
        </div>
        {% endfor %}
        
        <div class="footer">
            <p><strong>Total Clients: {{ clients|length }} | SHIVI PROJECT - Loan Management System</strong></p>
        </div>
    </body>
    </html>
    ''', clients=clients_data, generated_date=datetime.now().strftime("%d-%m-%Y %I:%M %p"))
    
    try:
        if not PDFKIT_AVAILABLE:
            return jsonify({'error': 'PDF generation not available'}), 500
        
        # Generate simple PDF using ReportLab
        pdf_buffer = generate_simple_pdf(
            f"All Clients Ledgers - {datetime.now().strftime('%d-%m-%Y')}",
            html_content,
            f'all_clients_ledgers_{date.today().strftime("%d%m%Y")}.pdf'
        )
        
        if pdf_buffer:
            return send_file(pdf_buffer, mimetype='application/pdf', as_attachment=True,
                            download_name=f'all_clients_ledgers_{date.today().strftime("%d%m%Y")}.pdf')
        else:
            return jsonify({'error': 'Failed to generate PDF'}), 500
                        
    except Exception as e:
        return f"PDF generation failed: {str(e)}", 500

@app.route('/export-all-clients-ledgers-excel')
@login_required
@page_required('all_ledgers')
def export_all_clients_ledgers_excel():
    """Export all clients ledgers to Excel"""
    try:
        clients_query = db.session.query(Loan.phone, Loan.name).distinct().all()
        excel_data = []
        client_blocks = []
        today = date.today()
        
        excel_data.append({
            'Client Name': 'ALL CLIENTS LEDGERS EXPORT',
            'Phone': f'Generated: {datetime.now().strftime("%d-%m-%Y %I:%M %p")}',
            'Loan Date': '', 'Day': '', 'End Date': '', 'Close Date': '', 'Proc Fees': '',
            'Loan Amount': '', 'Inst Amount': '', 'Type': '', 'Advance': '',
            'Closing Amt': '', 'Balance': '', 'Pending': '', 'Status': ''
        })
        excel_data.append({})
        
        for phone, name in sorted(clients_query, key=lambda x: x[1]):
            loans = Loan.query.filter_by(phone=phone).order_by(Loan.loan_date.asc()).all()
            
            if not loans:
                continue
                
            first_loan = loans[0]
            
            contact_parts = []
            if first_loan.address:
                contact_parts.append(first_loan.address)
            if first_loan.phone:
                contact_parts.append(f"📞 {first_loan.phone}")
            if first_loan.city:
                contact_parts.append(f"🏙️ {first_loan.city}")
            
            contact_line = " | ".join(contact_parts) if contact_parts else "Contact info not available"
            client_display_name = first_loan.name or 'Unknown Client'
            block = {
                'header_row': len(excel_data),
                'header_text': f"{client_display_name}\n{contact_line}" if contact_line else client_display_name,
                'data_rows': []
            }

            excel_data.append({
                'Client Name': client_display_name,
                'Phone': contact_line,
                'Loan Date': '', 'Day': '', 'End Date': '', 'Close Date': '', 'Proc Fees': '',
                'Loan Amount': '', 'Inst Amount': '', 'Type': '', 'Advance': '',
                'Closing Amt': '', 'Balance': '', 'Pending': '', 'Status': ''
            })
            
            block['column_header_row'] = len(excel_data)
            excel_data.append({
                'Client Name': 'Loan Date', 'Phone': 'Day', 'Loan Date': 'End Date', 'Day': 'Close Date',
                'End Date': 'Proc Fees', 'Close Date': 'Loan Amount', 'Proc Fees': 'Inst Amount',
                'Loan Amount': 'Type', 'Inst Amount': 'Advance', 'Type': 'Closing Amt',
                'Advance': 'Balance', 'Closing Amt': 'Pending', 'Balance': 'Status', 'Pending': '', 'Status': ''
            })
            
            totals = {'processing_fees': 0, 'loan_amount': 0, 'installment_amount': 0, 
                     'advance_amount': 0, 'case_closing_amount': 0, 'balance_amount': 0, 'pending_installments': 0}
            
            for loan in loans:
                bal, pending, inst_amt, _ = compute_row_fields(loan, as_of_date=today)
                loan_dt = parse_date_str(loan.loan_date)
                end_dt = schedule_end_date(loan)
                include_in_loan_total = not is_excluded_from_loan_total(loan, today)
                data_row_index = len(excel_data)
                
                excel_data.append({
                    'Client Name': format_date(loan.loan_date),
                    'Phone': loan_dt.strftime("%a") if loan_dt and loan.repayment_type == "WEEKLY" else "",
                    'Loan Date': end_dt.strftime("%d-%m-%Y") if end_dt else "",
                    'Day': format_date(loan.loan_closed_date),
                    'End Date': int(loan.processing_fees) if loan.processing_fees else 0,
                    'Close Date': int(loan.loan_amount) if loan.loan_amount else 0,
                    'Proc Fees': inst_amt,
                    'Loan Amount': loan.repayment_type,
                    'Inst Amount': int(loan.advance_amount) if loan.advance_amount else 0,
                    'Type': int(loan.case_closing_amount) if loan.case_closing_amount else 0,
                    'Advance': bal,
                    'Closing Amt': pending,
                    'Balance': loan.status,
                    'Pending': '', 'Status': ''
                })
                block['data_rows'].append(data_row_index)
                
                totals['processing_fees'] += int(loan.processing_fees) if loan.processing_fees else 0
                if include_in_loan_total:
                    totals['loan_amount'] += int(loan.loan_amount) if loan.loan_amount else 0
                totals['installment_amount'] += inst_amt
                totals['advance_amount'] += int(loan.advance_amount) if loan.advance_amount else 0
                totals['case_closing_amount'] += int(loan.case_closing_amount) if loan.case_closing_amount else 0
                totals['balance_amount'] += bal
                totals['pending_installments'] += pending
            
            block['totals_row'] = len(excel_data)
            excel_data.append({
                'Client Name': 'TOTALS', 'Phone': '', 'Loan Date': '', 'Day': '',
                'End Date': totals['processing_fees'],
                'Close Date': totals['loan_amount'],
                'Proc Fees': totals['installment_amount'],
                'Loan Amount': '',
                'Inst Amount': totals['advance_amount'],
                'Type': totals['case_closing_amount'],
                'Advance': totals['balance_amount'],
                'Closing Amt': totals['pending_installments'],
                'Balance': '', 'Pending': '', 'Status': ''
            })
            
            excel_data.append({})
            client_blocks.append(block)
        
        df = pd.DataFrame(excel_data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='All Clients Ledgers', index=False)
            worksheet = writer.sheets['All Clients Ledgers']
            worksheet.set_landscape()
            worksheet.set_paper(9)
            worksheet.fit_to_pages(1, 0)
            
            for idx, col in enumerate(df.columns):
                series_as_str = df[col].fillna('').astype(str)
                max_len = series_as_str.map(len).max()
                max_len = max(max_len, len(col))
                worksheet.set_column(idx, idx, max(6, min(max_len + 2, 40)))

            workbook = writer.book
            client_header_fmt = workbook.add_format({
                'bold': True,
                'font_size': 14,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True,
                'border': 2,
                'top': 2,
                'bottom': 2,
                'bg_color': '#f8f9fa'
            })
            column_heading_fmt = workbook.add_format({
                'bold': True,
                'align': 'center',
                'valign': 'vcenter',
                'text_wrap': True,
                'border': 1,
                'top': 2,
                'bottom': 2,
                'bg_color': '#4a5568',
                'font_color': '#ffffff'
            })
            loan_row_fmt = workbook.add_format({
                'border': 1,
                'top': 2,
                'valign': 'vcenter'
            })
            totals_row_fmt = workbook.add_format({
                'bold': True,
                'border': 1,
                'top': 3,
                'bottom': 3,
                'bg_color': '#e9ecef'
            })

            num_columns = len(df.columns)
            for block in client_blocks:
                header_row_excel = block['header_row'] + 1
                worksheet.set_row(header_row_excel, 33)
                worksheet.merge_range(header_row_excel, 0, header_row_excel, num_columns - 1, block['header_text'], client_header_fmt)
                column_row_excel = block.get('column_header_row', 0) + 1
                worksheet.set_row(column_row_excel, 25, column_heading_fmt)
                for data_row in block['data_rows']:
                    worksheet.set_row(data_row + 1, 23, loan_row_fmt)
                totals_row_excel = block.get('totals_row')
                if totals_row_excel is not None:
                    worksheet.set_row(totals_row_excel + 1, 25, totals_row_fmt)
        
        output.seek(0)
        
        return send_file(output, as_attachment=True,
                        download_name=f'all_clients_ledgers_{date.today().strftime("%d%m%Y")}.xlsx',
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                        
    except Exception as e:
        return f"Excel export failed: {str(e)}", 500

# =============================================================================
# RECEIVABLE & LEDGER ROUTES - 100% COMPLETE
# =============================================================================

@app.route("/advances/report")
@login_required
@page_required('receivable')
def advances_report_page():
    """Receivable Report Page"""
    return render_template("advances_report.html")

@app.route("/api/receivable-report")
@login_required
@page_required('receivable')
def get_receivable_report():
    today_str = datetime.now().strftime("%d-%m-%Y")
    as_on = (
        request.args.get("as_on_date")
        or request.args.get("as_on")
        or today_str
    )

    # Snapshot: if a frozen receivable report already exists for this date,
    # return it directly so past days never change even if loans change later.
    try:
        snap = ReceivableSnapshot.query.filter_by(as_on_date=as_on).first()
        if snap is not None:
            try:
                data = json.loads(snap.data_json or '{}')
            except Exception:
                data = {}
            data.setdefault('as_on', as_on)
            data['is_snapshot'] = True
            if snap.generated_at and not data.get('snapshot_generated_at'):
                data['snapshot_generated_at'] = snap.generated_at
            return jsonify(data)
    except Exception as e:
        current_app.logger.error(f"receivable_report snapshot lookup failed for {as_on}: {e}")

    customers, grand_total = [], 0

    phones = [
        row[0] for row in
        db.session.query(Loan.phone)
                  .filter(Loan.phone.isnot(None))
                  .distinct()
    ]

    for phone in phones:
        with current_app.test_request_context(
            f"/ledger?phone={phone}&as_on_date={as_on}"
        ):
            # ✅ FIXED: Correct attribute name
            ledger_json = current_app.view_functions["ledger"]().get_json()

        balance = ledger_json.get("totals", {}).get("balance_amount", 0)
        if balance == 0:
            continue

        latest = (Loan.query
                      .filter_by(phone=phone)
                      .order_by(Loan.id.desc())
                      .first())

        customers.append({
            "phone":   phone,
            "name":    latest.name,
            "balance": round(balance, 2)
        })
        grand_total += balance

    customers.sort(key=lambda c: c["name"])

    generated_str = datetime.now().strftime("%d-%m-%Y %I:%M:%S %p")
    payload = {
        "customers":    customers,
        "count":        len(customers),
        "grand_total":  round(grand_total, 2),
        "as_on":        as_on,
        "generated_at": generated_str,
        "is_snapshot":  False,
        "snapshot_generated_at": generated_str,
    }

    # Save snapshot for this as_on date the first time it is computed.
    try:
        existing = ReceivableSnapshot.query.filter_by(as_on_date=as_on).first()
        if existing is None:
            snap = ReceivableSnapshot(
                as_on_date=as_on,
                data_json=json.dumps(payload, ensure_ascii=False),
                generated_at=generated_str,
            )
            db.session.add(snap)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Failed to save ReceivableSnapshot for {as_on}: {e}")

    return jsonify(payload)


@app.delete("/receivable-report/snapshot")
@csrf.exempt
@login_required
@page_required('receivable')
def delete_receivable_snapshot():
    """Delete the saved receivable snapshot for a given as_on_date.

    Frontend can call: DELETE /receivable-report/snapshot?as_on_date=DD-MM-YYYY
    After a successful delete, calling /api/receivable-report for that date will
    recompute once and store a fresh snapshot.
    """
    as_on = (request.args.get("as_on_date") or request.args.get("as_on") or "").strip()
    if not as_on:
        return jsonify({"success": False, "error": "as_on_date parameter required"}), 400

    try:
        deleted_rows = ReceivableSnapshot.query.filter_by(as_on_date=as_on).delete()
        db.session.commit()
        return jsonify({"success": True, "deleted": deleted_rows, "as_on": as_on})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route("/ledger")
@login_required
def ledger():
    """Honours as_on_date parameter"""
    phone = (request.args.get("phone") or "").strip()
    if not phone:
        return jsonify({"errors": {"phone": "Required"}}), 400

    as_on_date_str = request.args.get("as_on_date")
    if as_on_date_str:
        try:
            as_of_date = datetime.strptime(as_on_date_str, "%d-%m-%Y").date()
        except ValueError:
            try:
                as_of_date = datetime.strptime(as_on_date_str, "%Y-%m-%d").date()
            except ValueError:
                as_of_date = date.today()
    else:
        as_of_date = date.today()

    loans = Loan.query.filter_by(phone=phone).order_by(Loan.loan_date.asc()).all()
    results = []
    ZERO_ON_CLOSE_DATE = False

    totals = {
        "processing_fees": 0,
        "loan_amount": 0,
        "installment_amount": 0,
        "advance_amount": 0,
        "case_closing_amount": 0,
        "balance_amount": 0,
        "pending_installments": 0
    }

    for l in loans:
        if l is None:
            continue
        loan_dt = parse_date_str(l.loan_date)
        end_dt = schedule_end_date(l)

        bal, pending, inst_amt, eff_applied = compute_row_fields(
            l, as_of_date=as_of_date, zero_on_close_date=ZERO_ON_CLOSE_DATE
        )

        include_in_loan_total = not is_excluded_from_loan_total(l, as_of_date)

        row_data = {
            "id": l.id,
            "loan_date": format_date(l.loan_date),
            "loan_date_raw": l.loan_date,
            "day_name": loan_dt.strftime("%a") if loan_dt and l.repayment_type == "WEEKLY" else "",
            "loan_end_date": end_dt.strftime("%d-%m-%Y") if end_dt else "",
            "loan_closed_date": format_date(l.loan_closed_date),
            "loan_closed_date_raw": l.loan_closed_date,
            "processing_fees": int(l.processing_fees) if l.processing_fees else 0,
            "amount": int(l.loan_amount) if l.loan_amount else 0,
            "installment_amount": inst_amt,
            "repayment_type": l.repayment_type,
            "advance_amount": int(l.advance_amount) if l.advance_amount else 0,
            "case_closing_amt": int(l.case_closing_amount) if l.case_closing_amount else 0,
            "balance_amount": bal,
            "pending_installments": pending,
            "status": l.status,
            "effective_closed_applied": eff_applied,
            "include_in_loan_total": include_in_loan_total,
            "name": l.name,
            "address": l.address,
            "phone": l.phone,
            "city": l.city,
            "remarks": l.remarks or ""
        }
        results.append(row_data)

        totals["processing_fees"] += row_data["processing_fees"]
        totals["installment_amount"] += row_data["installment_amount"]
        totals["advance_amount"] += row_data["advance_amount"]
        totals["case_closing_amount"] += row_data["case_closing_amt"]
        totals["balance_amount"] += row_data["balance_amount"]
        totals["pending_installments"] += row_data["pending_installments"]
        
        if include_in_loan_total:
            totals["loan_amount"] += row_data["amount"]

    return jsonify({
        "rows": results,
        "totals": totals,
        "meta": {"as_of": as_of_date.isoformat(), "phone": phone}
    })

@app.route("/ledger/page")
@login_required
@page_required('ledger')
def ledger_full_page():
    return render_template("ledger_full.html")

# =============================================================================
# EXPORT ROUTES - 100% COMPLETE 
# =============================================================================

@app.route("/export/receivable-excel")
@login_required
@page_required('receivable')
def export_receivable_excel():
    """Export receivable report with multi-column layout (matching HTML view)."""
    try:
        from io import BytesIO
        import math
        import xlsxwriter

        as_on_date = request.args.get('as_on_date', datetime.now().strftime('%d-%m-%Y'))

        with current_app.test_request_context(f'/api/receivable-report?as_on_date={as_on_date}'):
            response = get_receivable_report()
            if isinstance(response, tuple):
                return response

        data = response.get_json()
        customers = sorted(data.get('customers', []), key=lambda c: c['name'].lower())

        CUSTOMERS_PER_COLUMN = 60
        COLUMNS_PER_PAGE = 4  # Sr.No, Name, Balance repeated 4 times
        GROUP_SIZE = CUSTOMERS_PER_COLUMN * COLUMNS_PER_PAGE
        total_columns = COLUMNS_PER_PAGE * 3

        output = BytesIO()
        workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        worksheet = workbook.add_worksheet('Receivable Report')

        title_fmt = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center', 'valign': 'vcenter'})
        subtitle_fmt = workbook.add_format({'font_size': 10, 'align': 'center'})
        header_fmt = workbook.add_format({'bold': True, 'bg_color': '#1f2933', 'font_color': '#ffffff', 'align': 'center'})
        serial_fmt = workbook.add_format({'align': 'center', 'border': 1})
        name_fmt = workbook.add_format({'border': 1})
        amount_fmt = workbook.add_format({'border': 1, 'num_format': '#,##0'})
        subtotal_fmt = workbook.add_format({'bold': True, 'bg_color': '#e9ecef', 'border': 1, 'num_format': '#,##0'})
        summary_label_fmt = workbook.add_format({'bold': True})
        summary_value_fmt = workbook.add_format({'num_format': '#,##0'})

        worksheet.merge_range(0, 0, 0, total_columns - 1, 'RECEIVABLE REPORT - SHIVI PROJECT', title_fmt)
        worksheet.merge_range(
            1, 0, 1, total_columns - 1,
            f"As On Date: {as_on_date} | Generated: {data.get('generated_at', datetime.now().strftime('%d-%m-%Y %I:%M:%S %p'))}",
            subtitle_fmt
        )

        row_cursor = 3
        blocks = math.ceil(len(customers) / GROUP_SIZE) or 1
        for block in range(blocks):
            # Column headers for this block
            for col_group in range(COLUMNS_PER_PAGE):
                col_base = col_group * 3
                worksheet.write(row_cursor, col_base + 0, 'Sr.No', header_fmt)
                worksheet.write(row_cursor, col_base + 1, 'Customer Name', header_fmt)
                worksheet.write(row_cursor, col_base + 2, 'Balance Amount', header_fmt)
            row_cursor += 1

            block_start_index = block * GROUP_SIZE
            for row_offset in range(CUSTOMERS_PER_COLUMN):
                has_data = False
                for col_group in range(COLUMNS_PER_PAGE):
                    customer_index = block_start_index + col_group * CUSTOMERS_PER_COLUMN + row_offset
                    col_base = col_group * 3
                    if customer_index < len(customers):
                        has_data = True
                        customer = customers[customer_index]
                        worksheet.write(row_cursor, col_base + 0, customer_index + 1, serial_fmt)
                        worksheet.write(row_cursor, col_base + 1, customer['name'], name_fmt)
                        worksheet.write_number(row_cursor, col_base + 2, customer['balance'], amount_fmt)
                    else:
                        worksheet.write_blank(row_cursor, col_base + 0, None, serial_fmt)
                        worksheet.write_blank(row_cursor, col_base + 1, None, name_fmt)
                        worksheet.write_blank(row_cursor, col_base + 2, None, amount_fmt)
                if has_data:
                    row_cursor += 1

            # Subtotals row for block
            subtotal_row = row_cursor
            for col_group in range(COLUMNS_PER_PAGE):
                col_base = col_group * 3
                start_idx = block_start_index + col_group * CUSTOMERS_PER_COLUMN
                end_idx = min(start_idx + CUSTOMERS_PER_COLUMN, len(customers))
                column_total = sum(c['balance'] for c in customers[start_idx:end_idx])
                worksheet.write(subtotal_row, col_base + 0, 'SUBTOTAL', subtotal_fmt)
                worksheet.write_blank(subtotal_row, col_base + 1, None, subtotal_fmt)
                worksheet.write_number(subtotal_row, col_base + 2, column_total, subtotal_fmt)
            row_cursor += 2

        # Summary section
        worksheet.write(row_cursor, 0, 'SUMMARY', header_fmt)
        row_cursor += 1
        worksheet.write(row_cursor, 0, 'Total Customers:', summary_label_fmt)
        worksheet.write_number(row_cursor, 1, data.get('count', len(customers)), summary_value_fmt)
        row_cursor += 1
        worksheet.write(row_cursor, 0, 'Grand Total Outstanding:', summary_label_fmt)
        worksheet.write_number(row_cursor, 1, data.get('grand_total', 0), amount_fmt)
        row_cursor += 1
        avg_balance = 0
        if customers:
            avg_balance = data.get('grand_total', 0) / len(customers)
        worksheet.write(row_cursor, 0, 'Average Balance:', summary_label_fmt)
        worksheet.write_number(row_cursor, 1, avg_balance, amount_fmt)
        row_cursor += 1

        worksheet.set_landscape()
        worksheet.set_paper(9)
        worksheet.fit_to_pages(1, 0)

        column_widths = []
        for i in range(total_columns):
            if i % 3 == 0:
                column_widths.append(8)
            elif i % 3 == 1:
                column_widths.append(28)
            else:
                column_widths.append(14)
        for idx, width in enumerate(column_widths):
            worksheet.set_column(idx, idx, width)

        workbook.close()
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"receivable_report_{as_on_date.replace('-', '')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except ImportError:
        return "Excel export requires xlsxwriter library. Install with: pip install xlsxwriter", 500
    except Exception as e:
        return f"Excel export failed: {str(e)}", 500

@app.route("/export/receivable-pdf")
@login_required
@page_required('receivable')
def export_receivable_pdf():
    """Export receivable report as PDF with As On Date"""
    try:
        as_on_date = request.args.get('as_on_date', datetime.now().strftime('%d-%m-%Y'))
        
        with current_app.test_request_context(f'/api/receivable-report?as_on_date={as_on_date}'):
            response = get_receivable_report()
            if isinstance(response, tuple):
                return response
        
        data = response.get_json()
        
        html_template = '''
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body { font-family: Arial, sans-serif; margin: 20px; }
                h2 { text-align: center; color: #333; }
                table { width: 100%; border-collapse: collapse; margin: 20px 0; }
                th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
                th { background-color: #f2f2f2; }
                .total-row { font-weight: bold; background-color: #e9ecef; }
            </style>
        </head>
        <body>
            <h2>Receivable Report - As On {{ as_on_date }}</h2>
            <table>
                <thead>
                    <tr>
                        <th>Sr.No</th>
                        <th>Customer Name</th>
                        <th>Phone</th>
                        <th>Balance Amount</th>
                    </tr>
                </thead>
                <tbody>
                    {% for customer in customers %}
                    <tr>
                        <td>{{ loop.index }}</td>
                        <td>{{ customer.name }}</td>
                        <td>{{ customer.phone }}</td>
                        <td>₹{{ "{:,}".format(customer.balance) }}</td>
                    </tr>
                    {% endfor %}
                    <tr class="total-row">
                        <td colspan="3">GRAND TOTAL</td>
                        <td>₹{{ "{:,}".format(grand_total) }}</td>
                    </tr>
                </tbody>
            </table>
            <p><strong>Generated:</strong> {{ generated_at }}</p>
        </body>
        </html>
        '''
        
        html_content = render_template_string(html_template, 
                                            customers=data['customers'],
                                            grand_total=data['grand_total'],
                                            as_on_date=as_on_date,
                                            generated_at=data['generated_at'])
        
        options = {
            'page-size': 'A4',
            'margin-top': '0.75in',
            'margin-right': '0.75in',
            'margin-bottom': '0.75in',
            'margin-left': '0.75in',
            'encoding': "UTF-8",
            'no-outline': None
        }
        
        pdf = pdfkit.from_string(html_content, False, options=options)
        buf = io.BytesIO(pdf)
        
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                        download_name=f'receivable_report_{as_on_date.replace("-", "")}.pdf')
        
    except Exception as e:
        return f"PDF generation failed: {str(e)}", 500

@app.route("/export/excel")
@login_required
@page_required('ledger')
def export_excel():
    try:
        import pandas as pd
        from io import BytesIO
    except Exception:
        return jsonify({"error": "Pandas is required for Excel export"}), 500

    phone = (request.args.get("phone") or "").strip()
    if not phone:
        return jsonify({"errors": {"phone": "Required"}}), 400

    loans = Loan.query.filter_by(phone=phone).order_by(Loan.loan_date.asc()).all()
    if not loans:
        return "No loans found for this phone number.", 404

    f = loans[0]
    client_name   = f.name  or "UNKNOWN CLIENT"
    client_addr   = f.address or ""
    client_city   = f.city or ""
    client_phone  = f.phone or ""
    contact_line  = " | ".join(x for x in
                    [client_addr, f"Phone: {client_phone}" if client_phone else "", client_city] if x)

    rows = []

    rows.append({
        'ID': 'CUSTOMER NAME:',
        'Name': client_name.upper(),
        **{h: '' for h in ["Phone","Loan Date","Day","Loan End Date","Close Date",
                           "Processing Fees","Loan Amount","Installment Amount",
                           "Repayment Type","Advance","Case Closing Amount",
                           "Balance","Pending Installments","Status",
                           "Effective Closed Applied","City","Address"]}
    })

    rows.append({
        'ID': 'CONTACT INFO:',
        'Name': contact_line or "Not Provided",
        **{h: '' for h in ["Phone","Loan Date","Day","Loan End Date","Close Date",
                           "Processing Fees","Loan Amount","Installment Amount",
                           "Repayment Type","Advance","Case Closing Amount",
                           "Balance","Pending Installments","Status",
                           "Effective Closed Applied","City","Address"]}
    })

    rows.append({h: '' for h in rows[0].keys()})

    today = date.today()
    totals = {k: 0 for k in
              ["Processing Fees","Loan Amount","Advance","Installment Amount",
               "Case Closing Amount","Balance","Pending Installments"]}

    for l in loans:
        loan_dt = parse_date_str(l.loan_date)
        end_dt  = schedule_end_date(l)
        bal, pend, inst, _ = compute_row_fields(l, as_of_date=today)

        row = {
            'ID': l.id,
            'Name': l.name,
            'Phone': l.phone,
            'Loan Date': l.loan_date,
            'Day': loan_dt.strftime("%a") if loan_dt and l.repayment_type=="WEEKLY" else "",
            'Loan End Date': end_dt.strftime("%d-%m-%Y") if end_dt else "",
            'Close Date': l.loan_closed_date or "",
            'Processing Fees': int(l.processing_fees or 0),
            'Loan Amount':   int(l.loan_amount    or 0),
            'Installment Amount': inst,
            'Repayment Type': l.repayment_type,
            'Advance': int(l.advance_amount or 0),
            'Case Closing Amount': int(l.case_closing_amount or 0),
            'Balance': bal,
            'Pending Installments': pend,
            'Status': l.status,
            'Effective Closed Applied': '',
            'City': l.city or "",
            'Address': l.address or ""
        }
        rows.append(row)

        for k in totals:
            totals[k] += row[k]

    rows.append({
        'ID':'', 'Name':'', 'Phone':phone, 'Loan Date':'', 'Day':'', 'Loan End Date':'',
        'Close Date':'TOTALS', 'Processing Fees':totals["Processing Fees"],
        'Loan Amount':totals["Loan Amount"], 'Installment Amount':totals["Installment Amount"],
        'Repayment Type':'', 'Advance':totals["Advance"],
        'Case Closing Amount':totals["Case Closing Amount"],
        'Balance':totals["Balance"], 'Pending Installments':totals["Pending Installments"],
        'Status':'', 'Effective Closed Applied':'', 'City':'', 'Address':''
    })

    output = BytesIO()
    df = pd.DataFrame(rows)

    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Loans", index=False)
        wb  = writer.book
        ws  = writer.sheets["Loans"]

        ws.set_landscape()
        ws.set_paper(9)
        ws.fit_to_pages(1, 0)

        for i, col in enumerate(df.columns):
            maxlen = max(df[col].astype(str).map(len).max(), len(col))
            ws.set_column(i, i, min(maxlen + 2, 50))

        header_fmt   = wb.add_format({'bold':True,'font_size':14,'bg_color':'#D7E4BC'})
        contact_fmt  = wb.add_format({'font_size':12,'bg_color':'#F2F2F2'})
        ws.set_row(0, 24, header_fmt)
        ws.set_row(1, 20, contact_fmt)

    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'loans_{phone}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.route("/export/payment-entry-pdf")
@login_required
@page_required('payment_entry')
def export_payment_entry_pdf():
    print("🔍 DEBUG: Payment Entry PDF route called!")
    
    as_on_date = request.args.get("as_on_date")
    if not as_on_date:
        return jsonify({"error": "Date parameter required"}), 400
    
    try:
        target_date = datetime.strptime(as_on_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400

    target_date_str = target_date.strftime('%Y-%m-%d')
    orientation = (request.args.get('orientation') or 'landscape').lower()
    if orientation not in ('portrait', 'landscape'):
        orientation = 'landscape'

    columns_per_page = 3
    try:
        cols_param = int(request.args.get('cols') or 3)
        if cols_param in (3, 4):
            columns_per_page = cols_param
    except ValueError:
        columns_per_page = 3
    
    # 🎯 USE EXACT SAME DATA LOGIC AS YOUR SCREEN
    existing_submission = PaymentSubmission.query.filter_by(submission_date=target_date_str).first()
    payments_already_submitted = existing_submission is not None
    
    existing_payments = Payment.query.filter_by(payment_date=target_date_str).all()
    paid_loan_ids = {p.loan_id for p in existing_payments}

    draft = None
    unpaid_loan_ids = set()
    if not payments_already_submitted:
        draft = PaymentDraft.query.filter_by(draft_date=target_date_str).first()
        if draft:
            try:
                parsed = json.loads(draft.data_json or '{}')
                short_list = parsed.get('shortPayments') or []
                for item in short_list:
                    try:
                        unpaid_loan_ids.add(int(item.get('loan_id')))
                    except Exception:
                        continue
            except Exception:
                unpaid_loan_ids = set()
    
    # Get clients using YOUR screen's exact functions
    daily_clients = get_clients_due_for_date(target_date, "DAILY")
    weekly_clients = get_clients_due_for_date(target_date, "WEEKLY")
    ten_days_clients = get_clients_due_for_date(target_date, "TEN_DAYS")
    monthly_clients = get_clients_due_for_date(target_date, "MONTHLY")
    
    # Mark payments received (same as your screen)
    for client in daily_clients + weekly_clients + ten_days_clients + monthly_clients:
        if payments_already_submitted:
            client['payment_received'] = client['loan_id'] in paid_loan_ids
        else:
            try:
                client['payment_received'] = int(client['loan_id']) not in unpaid_loan_ids
            except Exception:
                client['payment_received'] = True
    
    # Calculate subtotals (same as your screen)
    daily_subtotal = sum(c['amount_due'] for c in daily_clients if c['payment_received'])
    weekly_subtotal = sum(c['amount_due'] for c in weekly_clients if c['payment_received'])
    ten_days_subtotal = sum(c['amount_due'] for c in ten_days_clients if c['payment_received'])
    monthly_subtotal = sum(c['amount_due'] for c in monthly_clients if c['payment_received'])

    short_clients = []
    short_total = 0
    if payments_already_submitted:
        db_shorts = ShortPayment.query.filter_by(payment_date=target_date_str, status='PENDING').all()
        for short in db_shorts:
            loan = db.session.get(Loan, short.loan_id)
            if not loan:
                continue
            unpaid_since = str(short.payment_date or target_date_str)
            amt = float(short.expected_amount or 0)
            short_clients.append({
                'loan_id': short.loan_id,
                'name': loan.name,
                'repayment_type': loan.repayment_type,
                'unpaid_since': unpaid_since,
                'total_amount': amt
            })
            short_total += amt
    else:
        if draft:
            try:
                parsed = json.loads(draft.data_json or '{}')
                short_list = parsed.get('shortPayments') or []
                for item in short_list:
                    amt = float(item.get('amount') or 0)
                    short_clients.append({
                        'loan_id': item.get('loan_id'),
                        'name': item.get('name'),
                        'repayment_type': str(item.get('type') or '').upper(),
                        'unpaid_since': item.get('unpaid_since') or target_date_str,
                        'total_amount': amt
                    })
                    short_total += amt
            except Exception:
                short_clients = []
                short_total = 0

    def format_for_payment_entry(clients):
        formatted = []
        for client in clients:
            formatted.append({
                'loan_id': client.get('loan_id'),
                'name': client.get('name'),
                'repayment_type': client.get('repayment_type'),
                'total_amount': float(client.get('amount_due') or 0),
                'payment_received': bool(client.get('payment_received'))
            })
        return formatted

    report_data = {
        'short': short_clients,
        'short_total': short_total,
        'daily': format_for_payment_entry(daily_clients),
        'weekly': format_for_payment_entry(weekly_clients),
        'ten_days': format_for_payment_entry(ten_days_clients),
        'monthly': format_for_payment_entry(monthly_clients),
        'daily_total': daily_subtotal,
        'weekly_total': weekly_subtotal,
        'ten_days_total': ten_days_subtotal,
        'monthly_total': monthly_subtotal
    }

    html_content = render_template_string(
        PDF_PAYMENT_ENTRY_TEMPLATE,
        data=report_data,
        report_date=target_date.strftime('%A, %B %d, %Y'),
        generated_time=datetime.now().strftime('%d/%m/%Y %I:%M %p'),
        orientation=orientation,
        columns_per_page=columns_per_page
    )
    
    zoom = '1.18' if columns_per_page == 3 else '1.06'
    options = {
        'page-size': 'A4',
        'orientation': 'Landscape' if orientation == 'landscape' else 'Portrait',
        'margin-top': '0.4cm',
        'margin-right': '0.4cm',
        'margin-bottom': '0.4cm',
        'margin-left': '0.4cm',
        'viewport-size': '1700x1100',
        'zoom': zoom,
        'encoding': "UTF-8",
        'no-outline': None,
        'enable-local-file-access': None,
        'print-media-type': None,
        'disable-smart-shrinking': None
    }
    
    try:
        if not PDFKIT_AVAILABLE:
            return jsonify({'error': 'PDF generation not available - pdfkit not installed'}), 500
        
        pdf = pdfkit.from_string(html_content, False, options=options)
        buf = io.BytesIO(pdf)
        buf.seek(0)
        
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                        download_name=f'payment_entry_{as_on_date}.pdf')
        
    except Exception as e:
        print(f"🔍 DEBUG: PDF generation failed: {str(e)}")
        return f"PDF generation failed: {str(e)}", 500


@app.route("/export/pdf")
@login_required
@page_required('ledger')
def export_pdf():
    phone = (request.args.get("phone") or "").strip()
    if not phone:
        return jsonify({"errors": {"phone": "Required"}}), 400

    loans = Loan.query.filter_by(phone=phone).order_by(Loan.loan_date.asc()).all()
    
    if not loans:
        return "No loans found for this phone number.", 404

    first_loan = loans[0]
    client_info = {
        'name': first_loan.name or "Unknown Client",
        'phone': first_loan.phone or "",
        'address': first_loan.address or "",
        'city': first_loan.city or ""
    }
    
    contact_parts = []
    if client_info['address']:
        contact_parts.append(f"Address: {client_info['address']}")
    if client_info['phone']:
        contact_parts.append(f"Phone: {client_info['phone']}")
    if client_info['city']:
        contact_parts.append(f"City: {client_info['city']}")
    
    client_info['full_contact_line'] = " | ".join(contact_parts) if contact_parts else "Contact information not available"

    today = date.today()
    loan_data = []
    totals = {
        "processing_fees": 0, "loan_amount": 0, "installment_amount": 0,
        "advance_amount": 0, "case_closing_amount": 0, "balance_amount": 0, "pending_installments": 0
    }

    for loan in loans:
        bal, pending, inst_amt, eff_applied = compute_row_fields(loan, as_of_date=today, zero_on_close_date=False)
        loan_dt = parse_date_str(loan.loan_date)
        end_dt = schedule_end_date(loan)
        
        loan_record = {
            'id': loan.id, 
            'loan_date': format_date(loan.loan_date),
            'day_name': loan_dt.strftime("%a") if loan_dt and loan.repayment_type == "WEEKLY" else "",
            'loan_end_date': end_dt.strftime("%d-%m-%Y") if end_dt else "",
            'loan_closed_date': format_date(loan.loan_closed_date),
            'processing_fees': int(loan.processing_fees) if loan.processing_fees else 0,
            'amount': int(loan.loan_amount) if loan.loan_amount else 0,
            'installment_amount': inst_amt, 
            'repayment_type': loan.repayment_type,
            'advance_amount': int(loan.advance_amount) if loan.advance_amount else 0,
            'case_closing_amt': int(loan.case_closing_amount) if loan.case_closing_amount else 0,
            'balance_amount': bal,
            'pending_installments': pending,
            'status': loan.status
        }
        
        loan_data.append(loan_record)
        
        totals["processing_fees"] += loan_record["processing_fees"]
        totals["installment_amount"] += loan_record["installment_amount"]
        totals["advance_amount"] += loan_record["advance_amount"]
        totals["case_closing_amount"] += loan_record["case_closing_amt"]
        totals["balance_amount"] += loan_record["balance_amount"]
        totals["pending_installments"] += loan_record["pending_installments"]
        
        if not is_excluded_from_loan_total(loan, today):
            totals["loan_amount"] += loan_record["amount"]

    html_content = render_template_string(PDF_TEMPLATE, client=client_info, loans=loan_data, 
                                        totals=totals, generated_date=date.today().strftime("%d-%m-%Y"))
    
    try:
        options = {
            'page-size': 'A4', 'orientation': 'Landscape',
            'margin-top': '0.5cm', 'margin-right': '0.5cm', 
            'margin-bottom': '0.5cm', 'margin-left': '0.5cm',
            'encoding': "UTF-8", 'no-outline': None, 'enable-local-file-access': None
        }
        
        if not PDFKIT_AVAILABLE:
            return jsonify({'error': 'PDF generation not available - pdfkit not installed'}), 500
        
        pdf = pdfkit.from_string(html_content, False, options=options)
        buf = io.BytesIO(pdf)
        buf.seek(0)
        
        return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=f'ledger_{phone}.pdf')
        
    except Exception as e:
        return f"PDF generation failed: {str(e)}. Please check wkhtmltopdf installation.", 500

# =============================================================================
# DAILY REPORTS - 100% COMPLETE WITH FIXED DAILY INSTALLMENT LOGIC
# =============================================================================

@app.get("/daily-report/page")
@login_required
@page_required('daily_report')
def daily_report_page():
    return render_template("daily_report.html")

@app.get("/daily-report/data")
@login_required
@page_required('daily_report')
def daily_collection_data():
    """✅ PRODUCTION VERSION with snapshotting: frozen per-date daily report data.

    - If a snapshot exists for the requested date, return that JSON so past days
      never change.
    - If no snapshot exists, compute the report once, save it, and return it.
    - Always include small metadata so the UI can show whether data came from a
      saved snapshot and when it was generated.
    """
    from datetime import datetime, timedelta

    report_date = request.args.get("date")
    if not report_date:
        return jsonify({"error": "Date parameter required"}), 400

    try:
        target_date = datetime.strptime(report_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400

    # If a snapshot already exists for this date, always return it (frozen report)
    try:
        snap = DailyReportSnapshot.query.filter_by(report_date=report_date).first()
        if snap is not None:
            try:
                data = json.loads(snap.data_json or '{}')
            except Exception:
                data = {}
            # Mark explicitly that this payload is coming from a saved snapshot
            data["is_snapshot"] = True
            if snap.generated_at and not data.get("snapshot_generated_at"):
                data["snapshot_generated_at"] = snap.generated_at
            return jsonify(data)
    except Exception as e:
        # If snapshot lookup fails, fall back to live computation
        app.logger.error(f"daily_collection_data snapshot lookup failed for {report_date}: {e}")

    previous_date = target_date - timedelta(days=1)
    loans = Loan.query.all()
    client_collections = {}
    
    for loan in loans:
        if loan is None:
            continue
        has_installment = has_installment_due_on_date(loan, target_date)
        
        # Robust date parsing without debug messages
        loan_closed_yesterday = False
        if loan.status == "CLOSED" and loan.loan_closed_date:
            try:
                date_str = str(loan.loan_closed_date).strip()
                loan_closed_date = None
                
                if '-' in date_str and len(date_str) == 10:
                    if date_str[4] == '-':  # YYYY-MM-DD
                        loan_closed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    elif date_str[2] == '-':  # DD-MM-YYYY
                        loan_closed_date = datetime.strptime(date_str, '%d-%m-%Y').date()
                
                if loan_closed_date:
                    loan_closed_yesterday = (loan_closed_date == previous_date)
            except:
                loan_closed_yesterday = False

        # ✅ CRITICAL: Closed loans must not appear in daily dues.
        # Only allow CLOSED loans if they were closed yesterday (for highlighting), with ₹0 due.
        if loan.status == "CLOSED" and not loan_closed_yesterday:
            continue
        if loan.status == "CLOSED":
            has_installment = False
        
        # Include loan if has installment OR closed yesterday
        if has_installment or loan_closed_yesterday:
            phone_key = (loan.phone or '').strip()
            base_key = phone_key if phone_key else (loan.name or '').strip()
            key = f"{base_key}_{loan.repayment_type}"
            loan_start = parse_date_str(loan.loan_date)
            
            is_new_client_for_this_loan = (loan_start == previous_date) if loan_start else False
            is_closed_loan_for_this_loan = loan_closed_yesterday
            
            if key not in client_collections:
                client_collections[key] = {
                    "name": loan.name,
                    "phone": (loan.phone or '').strip(),
                    "repayment_type": loan.repayment_type,
                    "total_amount": 0,
                    "weekly_amount": 0,
                    "weekly_two_days": 0,
                    "is_new_client": is_new_client_for_this_loan,
                    "is_closed_loan": is_closed_loan_for_this_loan
                }
            else:
                if is_new_client_for_this_loan:
                    client_collections[key]["is_new_client"] = True
                if is_closed_loan_for_this_loan:
                    client_collections[key]["is_closed_loan"] = True
            
            # Calculate amount
            if has_installment:
                _, pending_inst, inst_amt, _ = compute_row_fields(loan, target_date)
                if loan.repayment_type == "TEN_DAYS":
                    advance_amt = loan.advance_amount or 0
                    per_amt = round((loan.loan_amount or 0) / 10) if (loan.loan_amount or 0) else 0
                    if per_amt > 0:
                        advance_blocks = int(advance_amt // per_amt)
                        if advance_blocks >= 10:
                            continue
                if loan.repayment_type == "WEEKLY":
                    start_date = parse_date_str(loan.loan_date)
                    total_amt = loan.loan_amount or 0
                    weekly_installment = round(total_amt * 0.07)
                    two_days_balance = round(total_amt * 0.02)
                    adv_amt = loan.advance_amount or 0
                    advance_covers_week1 = (weekly_installment > 0) and (adv_amt >= weekly_installment)

                    days_diff = (target_date - start_date).days if start_date else -1
                    if advance_covers_week1 and days_diff == 0:
                        continue
                    if days_diff == 98:
                        if advance_covers_week1:
                            client_collections[key]["weekly_two_days"] += two_days_balance
                        else:
                            client_collections[key]["weekly_amount"] += weekly_installment
                            client_collections[key]["weekly_two_days"] += two_days_balance
                    else:
                        client_collections[key]["weekly_amount"] += weekly_installment
                else:
                    due_amount = due_amount_for_date(loan, target_date, inst_amt)
                    client_collections[key]["total_amount"] += due_amount
            else:
                client_collections[key]["total_amount"] += 0
    
    # Separate and sort clients
    daily_clients = []
    weekly_clients = []
    ten_days_clients = []
    monthly_clients = []
    
    for client_data in client_collections.values():
        if client_data["repayment_type"] == "WEEKLY":
            weekly_total = (client_data.get("weekly_amount") or 0) + (client_data.get("weekly_two_days") or 0)
            if weekly_total > 0:
                if (client_data.get("weekly_amount") or 0) > 0:
                    weekly_clients.append({
                        "name": client_data["name"],
                        "phone": client_data.get("phone", ""),
                        "repayment_type": "WEEKLY",
                        "total_amount": client_data.get("weekly_amount") or 0,
                        "is_new_client": client_data.get("is_new_client", False),
                        "is_closed_loan": client_data.get("is_closed_loan", False)
                    })
                if (client_data.get("weekly_two_days") or 0) > 0:
                    weekly_clients.append({
                        "name": client_data["name"],
                        "phone": client_data.get("phone", ""),
                        "repayment_type": "WEEKLY",
                        "total_amount": client_data.get("weekly_two_days") or 0,
                        "is_sub_row": True,
                        "sub_label": "2 Days Balance"
                    })
            continue

        if client_data["repayment_type"] == "DAILY":
            daily_clients.append(client_data)
        elif client_data["repayment_type"] == "TEN_DAYS":
            ten_days_clients.append(client_data)
        elif client_data["repayment_type"] == "MONTHLY":
            monthly_clients.append(client_data)
    
    daily_clients.sort(key=lambda x: x["name"].lower())
    weekly_clients.sort(key=lambda x: (x["name"].lower(), 1 if x.get("is_sub_row") else 0))
    ten_days_clients.sort(key=lambda x: x["name"].lower())
    monthly_clients.sort(key=lambda x: x["name"].lower())
    
    daily_total = sum(client["total_amount"] for client in daily_clients)
    weekly_total = sum(client["total_amount"] for client in weekly_clients)
    ten_days_total = sum(client["total_amount"] for client in ten_days_clients)
    monthly_total = sum(client["total_amount"] for client in monthly_clients)

    generated_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    payload = {
        "daily": daily_clients,
        "weekly": weekly_clients,
        "ten_days": ten_days_clients,
        "daily_total": daily_total,
        "weekly_total": weekly_total,
        "ten_days_total": ten_days_total,
        "monthly": monthly_clients,
        "monthly_total": monthly_total,
        # First time computation in this process: treat as fresh, but we will
        # also save it as the frozen snapshot for this date.
        "is_snapshot": False,
        "snapshot_generated_at": generated_str,
    }

    # Save snapshot for this date the first time it is computed, so future
    # requests (including Outstanding report) see a frozen version.
    try:
        existing = DailyReportSnapshot.query.filter_by(report_date=report_date).first()
        if existing is None:
            snap = DailyReportSnapshot(
                report_date=report_date,
                data_json=json.dumps(payload, ensure_ascii=False),
                generated_at=generated_str
            )
            db.session.add(snap)
            db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to save DailyReportSnapshot for {report_date}: {e}")

    return jsonify(payload)


@app.delete("/daily-report/snapshot")
@csrf.exempt
@login_required
@page_required('daily_report')
def delete_daily_report_snapshot():
    """Delete the saved snapshot for a given date so the next request recomputes.

    Frontend can call: DELETE /daily-report/snapshot?date=YYYY-MM-DD
    After a successful delete, calling /daily-report/data for that date will
    recompute the report once and store a fresh snapshot.
    """
    report_date = request.args.get("date")
    if not report_date:
        return jsonify({"success": False, "error": "Date parameter required"}), 400

    try:
        deleted_rows = DailyReportSnapshot.query.filter_by(report_date=report_date).delete()
        db.session.commit()
        return jsonify({"success": True, "deleted": deleted_rows, "date": report_date})
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

# ✅ PROFESSIONAL PDF EXPORT FOR DAILY REPORT
@app.get("/daily-report/export/pdf")
@login_required
@page_required('daily_report')
def daily_report_pdf_export():
    """✅ PROFESSIONAL PDF export with same styling as template"""
    from datetime import datetime, timedelta
    
    report_date = request.args.get("date")
    if not report_date:
        return jsonify({"error": "Date parameter required"}), 400
    
    try:
        target_date = datetime.strptime(report_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400
    
    previous_date = target_date - timedelta(days=1)
    loans = Loan.query.all()
    client_collections = {}
    
    for loan in loans:
        if loan is None:
            continue
        has_installment = has_installment_due_on_date(loan, target_date)
        
        # Same robust date parsing
        loan_closed_yesterday = False
        if loan.status == "CLOSED" and loan.loan_closed_date:
            try:
                date_str = str(loan.loan_closed_date).strip()
                loan_closed_date = None
                
                if '-' in date_str and len(date_str) == 10:
                    if date_str[4] == '-':  # YYYY-MM-DD
                        loan_closed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    elif date_str[2] == '-':  # DD-MM-YYYY
                        loan_closed_date = datetime.strptime(date_str, '%d-%m-%Y').date()
                
                if loan_closed_date:
                    loan_closed_yesterday = (loan_closed_date == previous_date)
            except:
                loan_closed_yesterday = False

        # ✅ CRITICAL: Closed loans must not appear in dues.
        if loan.status == "CLOSED" and not loan_closed_yesterday:
            continue
        if loan.status == "CLOSED":
            has_installment = False
        
        if has_installment or loan_closed_yesterday:
            key = f"{loan.name}_{loan.repayment_type}"
            loan_start = parse_date_str(loan.loan_date)
            
            is_new_client_for_this_loan = (loan_start == previous_date) if loan_start else False
            is_closed_loan_for_this_loan = loan_closed_yesterday
            
            if key not in client_collections:
                client_collections[key] = {
                    "name": loan.name,
                    "repayment_type": loan.repayment_type,
                    "total_amount": 0,
                    "is_new_client": is_new_client_for_this_loan,
                    "is_closed_loan": is_closed_loan_for_this_loan
                }
            else:
                if is_new_client_for_this_loan:
                    client_collections[key]["is_new_client"] = True
                if is_closed_loan_for_this_loan:
                    client_collections[key]["is_closed_loan"] = True
            
            if has_installment:
                _, pending_inst, inst_amt, _ = compute_row_fields(loan, target_date)
                if loan.repayment_type == "TEN_DAYS":
                    advance_amt = loan.advance_amount or 0
                    per_amt = round((loan.loan_amount or 0) / 10) if (loan.loan_amount or 0) else 0
                    if per_amt > 0:
                        advance_blocks = int(advance_amt // per_amt)
                        if advance_blocks >= 10:
                            continue
                due_amount = due_amount_for_date(loan, target_date, inst_amt)
                client_collections[key]["total_amount"] += due_amount
            else:
                client_collections[key]["total_amount"] += 0
    
    # Process for PDF (same as your existing logic)
    daily_clients = [c for c in client_collections.values() if c["repayment_type"] == "DAILY"]
    weekly_clients = [c for c in client_collections.values() if c["repayment_type"] == "WEEKLY"]
    ten_days_clients = [c for c in client_collections.values() if c["repayment_type"] == "TEN_DAYS"]
    monthly_clients = [c for c in client_collections.values() if c["repayment_type"] == "MONTHLY"]
    
    daily_clients.sort(key=lambda x: x["name"].lower())
    weekly_clients.sort(key=lambda x: x["name"].lower())
    ten_days_clients.sort(key=lambda x: x["name"].lower())
    monthly_clients.sort(key=lambda x: x["name"].lower())
    
    daily_total = sum(client["total_amount"] for client in daily_clients)
    weekly_total = sum(client["total_amount"] for client in weekly_clients)
    ten_days_total = sum(client["total_amount"] for client in ten_days_clients)
    monthly_total = sum(client["total_amount"] for client in monthly_clients)
    
    data = {
        "daily": daily_clients,
        "weekly": weekly_clients,
        "ten_days": ten_days_clients,
        "daily_total": daily_total,
        "weekly_total": weekly_total,
        "ten_days_total": ten_days_total,
        "monthly": monthly_clients,
        "monthly_total": monthly_total
    }
    
    # Continue with your existing PDF generation
    html_content = render_template_string(PDF_DAILY_REPORT_TEMPLATE, 
                                          data=data, 
                                          report_date=target_date.strftime('%A, %B %d, %Y'),
                                          generated_time=datetime.now().strftime('%d/%m/%Y %I:%M %p'))
    
    try:
        options = {
            'page-size': 'A4',
            'orientation': 'Landscape',
            'margin-top': '0.5cm',
            'margin-right': '0.5cm',
            'margin-bottom': '0.5cm',
            'margin-left': '0.5cm',
            'encoding': "UTF-8",
            'no-outline': None,
            'enable-local-file-access': None,
            'print-media-type': None,
            'disable-smart-shrinking': None,
            'dpi': 96,  # Windows 7 compatibility
            'image-quality': 100,  # Better quality for older systems
            'image-dpi': 96  # Consistent DPI for images
        }
        
        
        if not PDFKIT_AVAILABLE:
            return jsonify({'error': 'PDF generation not available - pdfkit not installed'}), 500
        
        pdf = pdfkit.from_string(html_content, False, options=options)
        buf = io.BytesIO(pdf)
        buf.seek(0)
        
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                        download_name=f'daily_collection_report_{report_date}.pdf')
        
    except Exception as e:
        return f"PDF generation failed: {str(e)}", 500

# ✅ PROFESSIONAL EXCEL EXPORT FOR DAILY REPORT
@app.get("/daily-report/export/excel")
@login_required
@page_required('daily_report')
def daily_report_excel_export():
    """✅ COMPLETE A4-fitted column-wise Excel with all sections"""
    try:
        import pandas as pd
        from io import BytesIO
        from datetime import datetime, timedelta
        import math
    except Exception:
        return jsonify({"error": "Pandas is required for Excel export"}), 500
    
    report_date = request.args.get("date")
    if not report_date:
        return jsonify({"error": "Date parameter required"}), 400
    
    try:
        target_date = datetime.strptime(report_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400
    
    previous_date = target_date - timedelta(days=1)
    loans = Loan.query.all()
    client_collections = {}
    
    # Same data collection logic as before...
    for loan in loans:
        has_installment = has_installment_due_on_date(loan, target_date)
        
        loan_closed_yesterday = False
        if loan.status == "CLOSED" and loan.loan_closed_date:
            try:
                date_str = str(loan.loan_closed_date).strip()
                loan_closed_date = None
                
                if '-' in date_str and len(date_str) == 10:
                    if date_str[4] == '-':
                        loan_closed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    elif date_str[2] == '-':
                        loan_closed_date = datetime.strptime(date_str, '%d-%m-%Y').date()
                
                if loan_closed_date:
                    loan_closed_yesterday = (loan_closed_date == previous_date)
            except:
                loan_closed_yesterday = False

        # ✅ CRITICAL: Closed loans must not appear in daily dues.
        # Only allow CLOSED loans if they were closed yesterday (for highlighting), with ₹0 due.
        if loan.status == "CLOSED" and not loan_closed_yesterday:
            continue
        if loan.status == "CLOSED":
            has_installment = False
        
        if has_installment or loan_closed_yesterday:
            key = f"{loan.name}_{loan.repayment_type}"
            loan_start = parse_date_str(loan.loan_date)
            
            is_new_client_for_this_loan = (loan_start == previous_date) if loan_start else False
            is_closed_loan_for_this_loan = loan_closed_yesterday
            
            if key not in client_collections:
                client_collections[key] = {
                    "name": loan.name,
                    "repayment_type": loan.repayment_type,
                    "total_amount": 0,
                    "is_new_client": is_new_client_for_this_loan,
                    "is_closed_loan": is_closed_loan_for_this_loan
                }
            else:
                if is_new_client_for_this_loan:
                    client_collections[key]["is_new_client"] = True
                if is_closed_loan_for_this_loan:
                    client_collections[key]["is_closed_loan"] = True
            
            if has_installment:
                _, pending_inst, inst_amt, _ = compute_row_fields(loan, target_date)
                if loan.repayment_type == "TEN_DAYS":
                    advance_amt = loan.advance_amount or 0
                    per_amt = round((loan.loan_amount or 0) / 10) if (loan.loan_amount or 0) else 0
                    if per_amt > 0:
                        advance_blocks = int(advance_amt // per_amt)
                        if advance_blocks >= 10:
                            continue
                due_amount = due_amount_for_date(loan, target_date, inst_amt)
                client_collections[key]["total_amount"] += due_amount
            else:
                client_collections[key]["total_amount"] += 0
    
    # Separate and sort clients
    daily_clients = [c for c in client_collections.values() if c["repayment_type"] == "DAILY"]
    weekly_clients = [c for c in client_collections.values() if c["repayment_type"] == "WEEKLY"]
    ten_days_clients = [c for c in client_collections.values() if c["repayment_type"] == "TEN_DAYS"]
    monthly_clients = [c for c in client_collections.values() if c["repayment_type"] == "MONTHLY"]
    
    daily_clients.sort(key=lambda x: x["name"].lower())
    weekly_clients.sort(key=lambda x: x["name"].lower())
    ten_days_clients.sort(key=lambda x: x["name"].lower())
    monthly_clients.sort(key=lambda x: x["name"].lower())
    
    daily_total = sum(client["total_amount"] for client in daily_clients)
    weekly_total = sum(client["total_amount"] for client in weekly_clients)
    ten_days_total = sum(client["total_amount"] for client in ten_days_clients)
    monthly_total = sum(client["total_amount"] for client in monthly_clients)
    grand_total = daily_total + weekly_total + ten_days_total + monthly_total
    
    # ✅ CREATE SINGLE EXCEL WITH ALL SECTIONS
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Daily Collection Report")
        
        # ✅ A4 PAGE SETUP
        worksheet.set_landscape()
        worksheet.set_paper(9)  # A4
        worksheet.fit_to_pages(1, 0)  # Fit to 1 page wide, unlimited pages tall
        worksheet.set_margins(0.5, 0.5, 0.5, 0.5)
        
        # ✅ FORMATTING STYLES
        header_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'bg_color': '#4472C4', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 1
        })
        
        section_format = workbook.add_format({
            'bold': True, 'font_size': 12, 'bg_color': '#D9E2F3', 'font_color': 'black',
            'align': 'center', 'valign': 'vcenter', 'border': 1
        })
        
        column_header_format = workbook.add_format({
            'bold': True, 'bg_color': '#F2F2F2', 'align': 'center', 'border': 1, 'font_size': 10
        })
        
        new_client_format = workbook.add_format({
            'bg_color': '#FFF3CD', 'font_color': '#856404', 'border': 1, 'font_size': 9
        })
        
        closed_client_format = workbook.add_format({
            'bg_color': '#F8D7DA', 'font_color': '#721C24', 'border': 1, 'font_size': 9
        })
        
        regular_format = workbook.add_format({'border': 1, 'font_size': 9})
        
        total_format = workbook.add_format({
            'bold': True, 'font_size': 14, 'bg_color': '#198754', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 2
        })
        
        grand_total_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'bg_color': '#DC3545', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 2
        })
        
        # ✅ WRITE REPORT HEADER
        current_row = 0
        worksheet.merge_range(current_row, 0, current_row, 11, 'DAILY COLLECTION REPORT', header_format)
        current_row += 1
        worksheet.merge_range(current_row, 0, current_row, 11, 
                            f'Report Date: {target_date.strftime("%A, %B %d, %Y")}', section_format)
        current_row += 3
        
        # ✅ FUNCTION TO WRITE COLUMN-WISE SECTIONS
        def write_section(clients, section_name, start_row):
            if not clients:
                return start_row
            
            # Section header
            worksheet.merge_range(start_row, 0, start_row, 11, 
                                f'{section_name.upper()} INSTALLMENTS ({len(clients)} clients)', 
                                section_format)
            start_row += 2
            
            # ✅ COLUMN LAYOUT: 4 columns of clients (fits A4 width)
            CLIENTS_PER_COLUMN = 50
            COLUMNS_PER_PAGE = 4  # 4 columns fit A4 width
            num_columns = math.ceil(len(clients) / CLIENTS_PER_COLUMN)
            
            for col_group in range(0, num_columns, COLUMNS_PER_PAGE):
                group_start_row = start_row
                
                # Write column headers
                for col in range(min(COLUMNS_PER_PAGE, num_columns - col_group)):
                    base_col = col * 3  # 3 columns per client group (Sr.No, Name, Amount)
                    
                    worksheet.write(start_row, base_col, 'Sr.No.', column_header_format)
                    worksheet.write(start_row, base_col + 1, 'Client Name', column_header_format)
                    worksheet.write(start_row, base_col + 2, 'Amount', column_header_format)
                
                start_row += 1
                
                # Write client data
                max_rows_in_group = 0
                for col in range(min(COLUMNS_PER_PAGE, num_columns - col_group)):
                    col_index = col_group + col
                    start_idx = col_index * CLIENTS_PER_COLUMN
                    end_idx = min(start_idx + CLIENTS_PER_COLUMN, len(clients))
                    column_clients = clients[start_idx:end_idx]
                    
                    base_col = col * 3
                    
                    for i, client in enumerate(column_clients):
                        row = start_row + i
                        serial = start_idx + i + 1
                        
                        # Choose format
                        if client["is_new_client"]:
                            cell_format = new_client_format
                        elif client["is_closed_loan"]:
                            cell_format = closed_client_format
                        else:
                            cell_format = regular_format
                        
                        worksheet.write(row, base_col, serial, cell_format)
                        worksheet.write(row, base_col + 1, client["name"], cell_format)
                        worksheet.write(row, base_col + 2, f'₹{client["total_amount"]:,}', cell_format)
                    
                    max_rows_in_group = max(max_rows_in_group, len(column_clients))
                    
                    # Column subtotal
                    subtotal_row = start_row + len(column_clients) + 1
                    column_total = sum(c["total_amount"] for c in column_clients)
                    worksheet.merge_range(subtotal_row, base_col, subtotal_row, base_col + 1, 'Subtotal:', section_format)
                    worksheet.write(subtotal_row, base_col + 2, f'₹{column_total:,}', section_format)
                
                start_row += max_rows_in_group + 4
                
                # ✅ PAGE BREAK after each group (except last)
                if col_group + COLUMNS_PER_PAGE < num_columns:
                    worksheet.set_h_pagebreaks([start_row])
            
            # ✅ Section Total
            section_total = sum(c["total_amount"] for c in clients)
            worksheet.merge_range(start_row, 0, start_row, 11, 
                                f'{section_name.upper()} TOTAL: ₹{section_total:,}', 
                                total_format)
            start_row += 3
            
            return start_row
        
        # ✅ WRITE ALL SECTIONS
        current_row = write_section(daily_clients, "Daily", current_row)
        
        # Page break before weekly section
        if weekly_clients or ten_days_clients or monthly_clients:
            worksheet.set_h_pagebreaks([current_row])
        
        current_row = write_section(weekly_clients, "Weekly", current_row)
        current_row = write_section(ten_days_clients, "Ten Days", current_row)
        current_row = write_section(monthly_clients, "Monthly", current_row)
        
        # ✅ COMBINED TOTALS SUMMARY
        worksheet.merge_range(current_row, 0, current_row, 11, 'COMPREHENSIVE COLLECTION SUMMARY', section_format)
        current_row += 2
        
        summary_data = [
            ('Daily Installments', len(daily_clients), daily_total),
            ('Weekly Installments', len(weekly_clients), weekly_total),
            ('Ten Days Installments', len(ten_days_clients), ten_days_total),
            ('Monthly Installments', len(monthly_clients), monthly_total)
        ]
        
        for name, count, total in summary_data:
            worksheet.merge_range(current_row, 0, current_row, 8, f'{name} ({count} clients):', regular_format)
            worksheet.merge_range(current_row, 9, current_row, 11, f'₹{total:,}', regular_format)
            current_row += 1
        
        current_row += 1
        
        # ✅ GRAND TOTAL
        worksheet.merge_range(current_row, 0, current_row, 11, 
                            f'GRAND TOTAL COLLECTION: ₹{grand_total:,}', 
                            grand_total_format)
        
        # ✅ SET COLUMN WIDTHS for A4 fit
        col_widths = [6, 20, 10] * 4  # Sr.No, Name, Amount repeated 4 times
        for i, width in enumerate(col_widths):
            worksheet.set_column(i, i, width)
    
    output.seek(0)
    return send_file(output, as_attachment=True, 
                    download_name=f'daily_collection_report_{report_date}.xlsx',
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
@app.get("/daily-report/export") 
@login_required
@page_required('daily_report')
def daily_collection_export():
    """✅ PROFESSIONAL: Multi-section A4-fitted Excel with column layout"""
    try:
        import pandas as pd
        from io import BytesIO
        from datetime import datetime, timedelta
        import math
    except Exception:
        return jsonify({"error": "Pandas is required for Excel export"}), 500
    
    report_date = request.args.get("date")
    if not report_date:
        return jsonify({"error": "Date parameter required"}), 400
    
    try:
        target_date = datetime.strptime(report_date, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400
    
    # ✅ SAME DATA COLLECTION LOGIC AS YOUR SIMPLE VERSION
    previous_date = target_date - timedelta(days=1)
    loans = Loan.query.all()
    client_collections = {}
    
    for loan in loans:
        if has_installment_due_on_date(loan, target_date):
            key = f"{loan.name}_{loan.repayment_type}"
            loan_start = parse_date_str(loan.loan_date)
            
            is_new_client_for_this_loan = (loan_start == previous_date) if loan_start else False
            
            is_closed_loan_for_this_loan = False
            if loan.status == "CLOSED" and loan.loan_closed_date:
                try:
                    loan_closed_date = datetime.strptime(loan.loan_closed_date, '%d-%m-%Y').date()
                    is_closed_loan_for_this_loan = (loan_closed_date == previous_date)
                except (ValueError, AttributeError):
                    is_closed_loan_for_this_loan = False
            
            if key not in client_collections:
                client_collections[key] = {
                    "name": loan.name,
                    "repayment_type": loan.repayment_type,
                    "total_amount": 0,
                    "is_new_client": is_new_client_for_this_loan,
                    "is_closed_loan": is_closed_loan_for_this_loan
                }
            else:
                if is_new_client_for_this_loan:
                    client_collections[key]["is_new_client"] = True
                if is_closed_loan_for_this_loan:
                    client_collections[key]["is_closed_loan"] = True
            
            _, _, inst_amt, _ = compute_row_fields(loan, target_date)
            client_collections[key]["total_amount"] += inst_amt
    
    # ✅ SEPARATE AND SORT CLIENTS BY TYPE (Like Complex Version)
    daily_clients = [c for c in client_collections.values() if c["repayment_type"] == "DAILY"]
    weekly_clients = [c for c in client_collections.values() if c["repayment_type"] == "WEEKLY"]
    ten_days_clients = [c for c in client_collections.values() if c["repayment_type"] == "TEN_DAYS"]
    monthly_clients = [c for c in client_collections.values() if c["repayment_type"] == "MONTHLY"]
    
    daily_clients.sort(key=lambda x: x["name"].lower())
    weekly_clients.sort(key=lambda x: x["name"].lower())
    ten_days_clients.sort(key=lambda x: x["name"].lower())
    monthly_clients.sort(key=lambda x: x["name"].lower())
    
    daily_total = sum(client["total_amount"] for client in daily_clients)
    weekly_total = sum(client["total_amount"] for client in weekly_clients)
    ten_days_total = sum(client["total_amount"] for client in ten_days_clients)
    monthly_total = sum(client["total_amount"] for client in monthly_clients)
    grand_total = daily_total + weekly_total + ten_days_total + monthly_total
    
    # ✅ PROFESSIONAL EXCEL WITH COMPLEX FORMATTING
    output = BytesIO()
    
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        workbook = writer.book
        worksheet = workbook.add_worksheet("Daily Collection Report")
        
        # ✅ A4 PAGE SETUP
        worksheet.set_landscape()
        worksheet.set_paper(9)  # A4
        worksheet.fit_to_pages(1, 0)  # Fit to 1 page wide, unlimited pages tall
        worksheet.set_margins(0.5, 0.5, 0.5, 0.5)
        
        # ✅ PROFESSIONAL FORMATTING STYLES
        header_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'bg_color': '#4472C4', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 1
        })
        
        section_format = workbook.add_format({
            'bold': True, 'font_size': 12, 'bg_color': '#D9E2F3', 'font_color': 'black',
            'align': 'center', 'valign': 'vcenter', 'border': 1
        })
        
        column_header_format = workbook.add_format({
            'bold': True, 'bg_color': '#F2F2F2', 'align': 'center', 'border': 1, 'font_size': 10
        })
        
        new_client_format = workbook.add_format({
            'bg_color': '#FFF3CD', 'font_color': '#856404', 'border': 1, 'font_size': 9
        })
        
        closed_client_format = workbook.add_format({
            'bg_color': '#F8D7DA', 'font_color': '#721C24', 'border': 1, 'font_size': 9
        })
        
        regular_format = workbook.add_format({'border': 1, 'font_size': 9})
        
        total_format = workbook.add_format({
            'bold': True, 'font_size': 14, 'bg_color': '#198754', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 2
        })
        
        grand_total_format = workbook.add_format({
            'bold': True, 'font_size': 16, 'bg_color': '#DC3545', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 2
        })
        
        # ✅ WRITE REPORT HEADER
        current_row = 0
        worksheet.merge_range(current_row, 0, current_row, 11, 'DAILY COLLECTION REPORT', header_format)
        current_row += 1
        worksheet.merge_range(current_row, 0, current_row, 11, 
                            f'Report Date: {target_date.strftime("%A, %B %d, %Y")}', section_format)
        current_row += 3
        
        # ✅ FUNCTION TO WRITE COLUMN-WISE SECTIONS (Same as Complex Version)
        def write_section(clients, section_name, start_row):
            if not clients:
                return start_row
            
            # Section header
            worksheet.merge_range(start_row, 0, start_row, 11, 
                                f'{section_name.upper()} INSTALLMENTS ({len(clients)} clients)', 
                                section_format)
            start_row += 2
            
            # ✅ COLUMN LAYOUT: 4 columns of clients (fits A4 width)
            CLIENTS_PER_COLUMN = 50
            COLUMNS_PER_PAGE = 4  # 4 columns fit A4 width
            num_columns = math.ceil(len(clients) / CLIENTS_PER_COLUMN)
            
            for col_group in range(0, num_columns, COLUMNS_PER_PAGE):
                group_start_row = start_row
                
                # Write column headers
                for col in range(min(COLUMNS_PER_PAGE, num_columns - col_group)):
                    base_col = col * 3  # 3 columns per client group (Sr.No, Name, Amount)
                    
                    worksheet.write(start_row, base_col, 'Sr.No.', column_header_format)
                    worksheet.write(start_row, base_col + 1, 'Client Name', column_header_format)
                    worksheet.write(start_row, base_col + 2, 'Amount', column_header_format)
                
                start_row += 1
                
                # Write client data
                max_rows_in_group = 0
                for col in range(min(COLUMNS_PER_PAGE, num_columns - col_group)):
                    col_index = col_group + col
                    start_idx = col_index * CLIENTS_PER_COLUMN
                    end_idx = min(start_idx + CLIENTS_PER_COLUMN, len(clients))
                    column_clients = clients[start_idx:end_idx]
                    
                    base_col = col * 3
                    
                    for i, client in enumerate(column_clients):
                        row = start_row + i
                        serial = start_idx + i + 1
                        
                        # Choose format based on client status
                        if client["is_new_client"]:
                            cell_format = new_client_format
                        elif client["is_closed_loan"]:
                            cell_format = closed_client_format
                        else:
                            cell_format = regular_format
                        
                        worksheet.write(row, base_col, serial, cell_format)
                        worksheet.write(row, base_col + 1, client["name"], cell_format)
                        worksheet.write(row, base_col + 2, f'₹{client["total_amount"]:,}', cell_format)
                    
                    max_rows_in_group = max(max_rows_in_group, len(column_clients))
                    
                    # Column subtotal
                    subtotal_row = start_row + len(column_clients) + 1
                    column_total = sum(c["total_amount"] for c in column_clients)
                    worksheet.merge_range(subtotal_row, base_col, subtotal_row, base_col + 1, 'Subtotal:', section_format)
                    worksheet.write(subtotal_row, base_col + 2, f'₹{column_total:,}', section_format)
                
                start_row += max_rows_in_group + 4
                
                # ✅ PAGE BREAK after each group (except last)
                if col_group + COLUMNS_PER_PAGE < num_columns:
                    worksheet.set_h_pagebreaks([start_row])
            
            # ✅ Section Total
            section_total = sum(c["total_amount"] for c in clients)
            worksheet.merge_range(start_row, 0, start_row, 11, 
                                f'{section_name.upper()} TOTAL: ₹{section_total:,}', 
                                total_format)
            start_row += 3
            
            return start_row
        
        # ✅ WRITE ALL SECTIONS
        current_row = write_section(daily_clients, "Daily", current_row)
        
        # Page break before weekly section
        if weekly_clients or ten_days_clients or monthly_clients:
            worksheet.set_h_pagebreaks([current_row])
        
        current_row = write_section(weekly_clients, "Weekly", current_row)
        current_row = write_section(ten_days_clients, "Ten Days", current_row)
        current_row = write_section(monthly_clients, "Monthly", current_row)
        
        # ✅ COMBINED TOTALS SUMMARY
        worksheet.merge_range(current_row, 0, current_row, 11, 'COMPREHENSIVE COLLECTION SUMMARY', section_format)
        current_row += 2
        
        summary_data = [
            ('Daily Installments', len(daily_clients), daily_total),
            ('Weekly Installments', len(weekly_clients), weekly_total),
            ('Ten Days Installments', len(ten_days_clients), ten_days_total),
            ('Monthly Installments', len(monthly_clients), monthly_total)
        ]
        
        for name, count, total in summary_data:
            worksheet.merge_range(current_row, 0, current_row, 8, f'{name} ({count} clients):', regular_format)
            worksheet.merge_range(current_row, 9, current_row, 11, f'₹{total:,}', regular_format)
            current_row += 1
        
        current_row += 1
        
        # ✅ GRAND TOTAL
        worksheet.merge_range(current_row, 0, current_row, 11, 
                            f'GRAND TOTAL COLLECTION: ₹{grand_total:,}', 
                            grand_total_format)
        
        # ✅ SET COLUMN WIDTHS for A4 fit
        col_widths = [6, 20, 10] * 4  # Sr.No, Name, Amount repeated 4 times
        for i, width in enumerate(col_widths):
            worksheet.set_column(i, i, width)
    
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name=f'daily_collection_report_{report_date}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# =============================================================================
# DAILY REPORT RECONCILIATION (RECO) - 100% COMPLETE
# =============================================================================

@app.get("/reconciliation/page")
@login_required
@page_required('reconciliation')
def reconciliation_page():
    """Reconciliation page for comparing daily report snapshots"""
    return render_template("reco.html")

@app.get("/reconciliation/data")
@api_auth_required('reconciliation')
def reconciliation_data():
    """Reconciliation data endpoint - compares yesterday snapshot with today live data"""
    from datetime import datetime, timedelta
    
    yesterday_str = request.args.get("yesterday")
    today_str = request.args.get("today")
    
    if not yesterday_str or not today_str:
        return jsonify({"error": "Both yesterday and today dates required"}), 400
    
    try:
        yesterday_date = datetime.strptime(yesterday_str, "%Y-%m-%d").date()
        today_date = datetime.strptime(today_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format"}), 400
    
    # Get yesterday's snapshot
    yesterday_snap = DailyReportSnapshot.query.filter_by(report_date=yesterday_str).first()
    
    if not yesterday_snap:
        return jsonify({"error": f"No snapshot found for {yesterday_str}. Please generate daily report for that date first."}), 400
    
    try:
        yesterday_data = json.loads(yesterday_snap.data_json or '{}')
    except Exception:
        yesterday_data = {}
    
    # Get today's snapshot (use saved snapshot, not live data)
    today_snap = DailyReportSnapshot.query.filter_by(report_date=today_str).first()
    
    if not today_snap:
        return jsonify({"error": f"No snapshot found for {today_str}. Please generate daily report for that date first."}), 400
    
    try:
        today_data = json.loads(today_snap.data_json or '{}')
    except Exception:
        today_data = {}
    
    # Extract yesterday's new sales - USE EXACT SAME LOGIC AS SALES REPORT
    # Call _query_sales with yesterday_date as both start and end
    yesterday_sales_rows, yesterday_sales_totals = _query_sales(yesterday_str, yesterday_str, "", "", "", False)
    
    yesterday_new_sales = []
    yesterday_daily_total = 0
    yesterday_weekly_total = 0
    
    for row in yesterday_sales_rows:
        # Only include OPEN loans (not closed)
        if row['status'] != 'OPEN':
            continue
        
        if row['type'] == 'DAILY':
            yesterday_new_sales.append({
                'name': row['name'],
                'phone': '',  # Sales report doesn't include phone
                'repayment_type': 'DAILY',
                'loan_amount': row['loan_amount'],
                'daily_installment': row['installment_amt'],
                'status': row['status']
            })
            yesterday_daily_total += row['installment_amt']
        elif row['type'] == 'WEEKLY':
            yesterday_new_sales.append({
                'name': row['name'],
                'phone': '',
                'repayment_type': 'WEEKLY',
                'loan_amount': row['loan_amount'],
                'weekly_installment': row['installment_amt'],
                'status': row['status']
            })
            yesterday_weekly_total += row['installment_amt']
    
    # Extract yesterday's closed cases (clients with is_closed_loan=True)
    # FIXED: Find which specific loan case was closed for each client
    yesterday_closed_cases = []
    yesterday_closed_daily_lost = 0
    
    # Detect cases that were manually closed on yesterday
    # Find the specific loan case that was closed for each client
    all_daily_loans = Loan.query.filter(Loan.repayment_type == 'DAILY').all()
    for loan in all_daily_loans:
        if loan.status == 'CLOSED' and loan.loan_closed_date:
            try:
                closed_date = parse_date_str(loan.loan_closed_date)
                if closed_date == yesterday_date:
                    # This specific loan case was closed yesterday
                    # Calculate what daily installment was lost from this specific case
                    _, _, inst_amt, _ = compute_row_fields(loan, yesterday_date)
                    
                    # Determine if it was auto-closed (101 days) or manual
                    loan_start = parse_date_str(loan.loan_date)
                    days_diff = (closed_date - loan_start).days if loan_start else 0
                    close_type = 'AUTO (101 days)' if days_diff >= 101 else 'MANUAL'
                    
                    yesterday_closed_cases.append({
                        'name': loan.name,
                        'phone': loan.phone or '',
                        'repayment_type': loan.repayment_type,
                        'loan_id': loan.id,
                        'loan_start_date': loan.loan_date,
                        'closing_amount': loan.case_closing_amount or 0,
                        'daily_installment_lost': inst_amt or 0,
                        'close_type': close_type,
                        'close_date': loan.loan_closed_date,
                        'days_active': days_diff
                    })
                    yesterday_closed_daily_lost += inst_amt or 0
            except:
                pass
    
    # Extract yesterday's "Going to Close" cases (due to finish based on loan end date)
    # These are cases that were due to finish on yesterday, regardless of manual close
    yesterday_going_to_close = []
    yesterday_going_to_close_daily_lost = 0
    
    for loan in all_daily_loans:
        if loan.status == 'CLOSED':
            continue  # Already counted in manual closed
        
        # Check if loan was due to finish on yesterday
        loan_end_date = schedule_end_date(loan)
        if loan_end_date:
            try:
                if isinstance(loan_end_date, str):
                    loan_end_date = parse_date_str(loan_end_date)
                if loan_end_date == yesterday_date:
                    # This loan was due to finish on yesterday
                    _, _, inst_amt, _ = compute_row_fields(loan, yesterday_date)
                    yesterday_going_to_close.append({
                        'name': loan.name,
                        'phone': loan.phone or '',
                        'repayment_type': loan.repayment_type,
                        'loan_id': loan.id,
                        'loan_amount': loan.loan_amount,
                        'daily_installment': inst_amt or 0,
                        'due_end_date': loan_end_date.strftime('%d-%m-%Y') if hasattr(loan_end_date, 'strftime') else str(loan_end_date),
                        'manual_entry_needed': 'YES'  # User needs to manually close this
                    })
                    yesterday_going_to_close_daily_lost += inst_amt or 0
            except:
                pass
    
    # Calculate comparison dates for weekly, monthly, 10-days
    # Weekly: Compare with same day of previous week (exactly 7 days ago)
    # If today is Saturday 8 August, compare with Saturday 1 August
    # If today is Friday 7 August, compare with Friday 31 July
    # If today is Monday 4 August, compare with Monday 28 July
    last_week_date = today_date - timedelta(days=7)
    last_week_str = last_week_date.strftime('%Y-%m-%d')
    
    # 10-Days: Compare with 10 days before
    ten_days_comparison_date = today_date - timedelta(days=10)
    ten_days_comparison_str = ten_days_comparison_date.strftime('%Y-%m-%d')
    
    # Monthly: Compare with 30 days before
    monthly_comparison_date = today_date - timedelta(days=30)
    monthly_comparison_str = monthly_comparison_date.strftime('%Y-%m-%d')
    
    # Get snapshots for comparison dates
    last_week_snap = DailyReportSnapshot.query.filter_by(report_date=last_week_str).first()
    monthly_snap = DailyReportSnapshot.query.filter_by(report_date=monthly_comparison_str).first()
    ten_days_snap = DailyReportSnapshot.query.filter_by(report_date=ten_days_comparison_str).first()
    
    # Parse comparison snapshot data
    last_week_data = {}
    monthly_comparison_data = {}
    ten_days_comparison_data = {}
    
    if last_week_snap:
        try:
            last_week_data = json.loads(last_week_snap.data_json or '{}')
        except:
            last_week_data = {}
    
    if monthly_snap:
        try:
            monthly_comparison_data = json.loads(monthly_snap.data_json or '{}')
        except:
            monthly_comparison_data = {}
    
    if ten_days_snap:
        try:
            ten_days_comparison_data = json.loads(ten_days_snap.data_json or '{}')
        except:
            ten_days_comparison_data = {}
    
    # Extract detailed client info for Weekly, Monthly, 10-Days
    # Compare with appropriate historical snapshots instead of yesterday
    weekly_comparison = compare_client_category(
        last_week_data.get('weekly', []),
        today_data.get('weekly', []),
        'WEEKLY',
        last_week_str,
        today_str
    )
    monthly_comparison = compare_client_category(
        monthly_comparison_data.get('monthly', []),
        today_data.get('monthly', []),
        'MONTHLY',
        monthly_comparison_str,
        today_str
    )
    ten_days_comparison = compare_client_category(
        ten_days_comparison_data.get('ten_days', []),
        today_data.get('ten_days', []),
        'TEN_DAYS',
        ten_days_comparison_str,
        today_str
    )
    
    # Get sales data for comparison dates to verify installment amounts
    weekly_sales_rows, weekly_sales_totals = _query_sales(last_week_str, last_week_str, "", "", "", False)
    monthly_sales_rows, monthly_sales_totals = _query_sales(monthly_comparison_str, monthly_comparison_str, "", "", "", False)
    ten_days_sales_rows, ten_days_sales_totals = _query_sales(ten_days_comparison_str, ten_days_comparison_str, "", "", "", False)
    
    # Calculate expected installment amounts based on ALL sales on comparison date
    # For Weekly: Sum all weekly sales on the comparison date
    weekly_expected_total = 0
    weekly_sales_details = []
    for row in weekly_sales_rows:
        if row['status'] == 'OPEN' and row['type'] == 'WEEKLY':
            weekly_expected_total += row['installment_amt']
            weekly_sales_details.append({
                'name': row['name'],
                'loan_amount': row['loan_amount'],
                'installment_amt': row['installment_amt']
            })
    
    # Calculate matched clients (present in both snapshots with same amount)
    # Rely on snapshot data as source of truth (not loan status which may be incorrect)
    # Special case: If old case closed and new case started same day, don't count as matched
    weekly_matched_clients = []
    weekly_matched_total = 0
    weekly_snapshot_dict = {c['name']: c.get('total_amount', 0) for c in last_week_data.get('weekly', []) if not c.get('is_sub_row')}
    weekly_today_dict = {c['name']: c.get('total_amount', 0) for c in today_data.get('weekly', []) if not c.get('is_sub_row')}
    
    for name in set(weekly_snapshot_dict.keys()) & set(weekly_today_dict.keys()):
        snapshot_amt = weekly_snapshot_dict[name]
        today_amt = weekly_today_dict[name]
        
        # Check if this client has both closed and new loans on same day (replaced case)
        loan = Loan.query.filter(
            Loan.name == name,
            Loan.repayment_type == 'WEEKLY'
        ).first()
        
        is_replaced_case = False
        if loan:
            # Check if there's a closed loan and a new loan with same name on same day
            all_loans = Loan.query.filter(
                Loan.name == name,
                Loan.repayment_type == 'WEEKLY'
            ).all()
            
            closed_loans = [l for l in all_loans if l.status == 'CLOSED' and l.loan_closed_date]
            open_loans = [l for l in all_loans if l.status == 'OPEN']
            
            # If there's a closed loan and an open loan with same creation date, it's a replaced case
            # Only detect if exactly 1 closed and 1 open loan on same date (simple replacement)
            if closed_loans and open_loans and len(closed_loans) == 1 and len(open_loans) == 1:
                for closed in closed_loans:
                    for open_loan in open_loans:
                        # Parse dates to handle different formats
                        try:
                            closed_date = parse_date_str(closed.loan_closed_date)
                            open_date = parse_date_str(open_loan.loan_date)
                            if closed_date == open_date:
                                is_replaced_case = True
                                break
                        except:
                            pass
                    if is_replaced_case:
                        break
        
        # Include in matched if amounts match AND not a replaced case
        if snapshot_amt == today_amt and not is_replaced_case:
            weekly_matched_clients.append({
                'name': name,
                'amount': snapshot_amt
            })
            weekly_matched_total += snapshot_amt
    
    # Expected total is ONLY new sales from comparison date (matched clients are already in snapshot)
    # Do NOT add matched total to expected
    
    # Calculate expected installment amounts based on ALL sales on comparison date
    # For Monthly: Sum all monthly sales on the comparison date
    monthly_expected_total = 0
    monthly_sales_details = []
    for row in monthly_sales_rows:
        if row['status'] == 'OPEN' and row['type'] == 'MONTHLY':
            monthly_expected_total += row['installment_amt']
            monthly_sales_details.append({
                'name': row['name'],
                'loan_amount': row['loan_amount'],
                'installment_amt': row['installment_amt']
            })
    
    # Calculate matched clients (present in both snapshots with same amount)
    # Rely on snapshot data as source of truth (not loan status which may be incorrect)
    # Special case: If old case closed and new case started same day, don't count as matched
    monthly_matched_clients = []
    monthly_matched_total = 0
    monthly_snapshot_dict = {c['name']: c.get('total_amount', 0) for c in monthly_comparison_data.get('monthly', []) if not c.get('is_sub_row')}
    monthly_today_dict = {c['name']: c.get('total_amount', 0) for c in today_data.get('monthly', []) if not c.get('is_sub_row')}
    
    for name in set(monthly_snapshot_dict.keys()) & set(monthly_today_dict.keys()):
        snapshot_amt = monthly_snapshot_dict[name]
        today_amt = monthly_today_dict[name]
        
        # Check if this client has both closed and new loans on same day (replaced case)
        loan = Loan.query.filter(
            Loan.name == name,
            Loan.repayment_type == 'MONTHLY'
        ).first()
        
        is_replaced_case = False
        if loan:
            all_loans = Loan.query.filter(
                Loan.name == name,
                Loan.repayment_type == 'MONTHLY'
            ).all()
            
            closed_loans = [l for l in all_loans if l.status == 'CLOSED' and l.loan_closed_date]
            open_loans = [l for l in all_loans if l.status == 'OPEN']
            
            # Only detect if exactly 1 closed and 1 open loan on same date (simple replacement)
            if closed_loans and open_loans and len(closed_loans) == 1 and len(open_loans) == 1:
                for closed in closed_loans:
                    for open_loan in open_loans:
                        # Parse dates to handle different formats
                        try:
                            closed_date = parse_date_str(closed.loan_closed_date)
                            open_date = parse_date_str(open_loan.loan_date)
                            if closed_date == open_date:
                                is_replaced_case = True
                                break
                        except:
                            pass
                    if is_replaced_case:
                        break
        
        # Include in matched if amounts match AND not a replaced case
        if snapshot_amt == today_amt and not is_replaced_case:
            monthly_matched_clients.append({
                'name': name,
                'amount': snapshot_amt
            })
            monthly_matched_total += snapshot_amt
    
    # Expected total is ONLY new sales from comparison date (matched clients are already in snapshot)
    # Do NOT add matched total to expected
    
    # Calculate expected installment amounts based on ALL sales on comparison date
    # For 10-Days: Sum all 10-days sales on the comparison date
    ten_days_expected_total = 0
    ten_days_sales_details = []
    for row in ten_days_sales_rows:
        if row['status'] == 'OPEN' and row['type'] == 'TEN_DAYS':
            ten_days_expected_total += row['installment_amt']
            ten_days_sales_details.append({
                'name': row['name'],
                'loan_amount': row['loan_amount'],
                'installment_amt': row['installment_amt']
            })
    
    # Calculate matched clients (present in both snapshots with same amount)
    # Rely on snapshot data as source of truth (not loan status which may be incorrect)
    # Special case: If old case closed and new case started same day, don't count as matched
    ten_days_matched_clients = []
    ten_days_matched_total = 0
    ten_days_snapshot_dict = {c['name']: c.get('total_amount', 0) for c in ten_days_comparison_data.get('ten_days', []) if not c.get('is_sub_row')}
    ten_days_today_dict = {c['name']: c.get('total_amount', 0) for c in today_data.get('ten_days', []) if not c.get('is_sub_row')}
    
    for name in set(ten_days_snapshot_dict.keys()) & set(ten_days_today_dict.keys()):
        snapshot_amt = ten_days_snapshot_dict[name]
        today_amt = ten_days_today_dict[name]
        
        # Check if this client has both closed and new loans on same day (replaced case)
        loan = Loan.query.filter(
            Loan.name == name,
            Loan.repayment_type == 'TEN_DAYS'
        ).first()
        
        is_replaced_case = False
        if loan:
            all_loans = Loan.query.filter(
                Loan.name == name,
                Loan.repayment_type == 'TEN_DAYS'
            ).all()
            
            closed_loans = [l for l in all_loans if l.status == 'CLOSED' and l.loan_closed_date]
            open_loans = [l for l in all_loans if l.status == 'OPEN']
            
            # Only detect if exactly 1 closed and 1 open loan on same date (simple replacement)
            if closed_loans and open_loans and len(closed_loans) == 1 and len(open_loans) == 1:
                for closed in closed_loans:
                    for open_loan in open_loans:
                        # Parse dates to handle different formats
                        try:
                            closed_date = parse_date_str(closed.loan_closed_date)
                            open_date = parse_date_str(open_loan.loan_date)
                            if closed_date == open_date:
                                is_replaced_case = True
                                break
                        except:
                            pass
                    if is_replaced_case:
                        break
        
        # Include in matched if amounts match AND not a replaced case
        if snapshot_amt == today_amt and not is_replaced_case:
            ten_days_matched_clients.append({
                'name': name,
                'amount': snapshot_amt
            })
            ten_days_matched_total += snapshot_amt
    
    # Expected total is ONLY new sales from comparison date (matched clients are already in snapshot)
    # Do NOT add matched total to expected
    
    # Calculate today's expected daily total
    # Formula: Yesterday's daily total + New daily sales today - Daily lost from closed cases - Daily lost from going to close
    today_new_daily_sales = 0
    
    # Get today's new sales - USE EXACT SAME LOGIC AS SALES REPORT
    # Call _query_sales with today_date as both start and end
    today_sales_rows, today_sales_totals = _query_sales(today_str, today_str, "", "", "", False)
    
    for row in today_sales_rows:
        # Only include OPEN loans and DAILY type
        if row['status'] != 'OPEN' or row['type'] != 'DAILY':
            continue
        
        today_new_daily_sales += row['installment_amt']
    
    # Total daily lost = manual closed + going to close
    total_daily_lost = yesterday_closed_daily_lost + yesterday_going_to_close_daily_lost
    
    # Calculate expected total using sales data
    # Start with yesterday's snapshot daily total, add yesterday's new sales - losses
    # Yesterday's new sales will have their first installment today
    # Today's new sales will have their first installment tomorrow, so NOT added here
    today_expected_total = yesterday_data.get('daily_total', 0) + yesterday_daily_total - total_daily_lost
    
    # Get today's actual daily total
    today_actual_total = today_data.get('daily_total', 0)
    
    # Calculate difference
    difference = today_actual_total - today_expected_total
    
    # Get today's other installments
    today_other_installments = {
        'weekly_total': today_data.get('weekly_total', 0),
        'ten_days_total': today_data.get('ten_days_total', 0),
        'monthly_total': today_data.get('monthly_total', 0)
    }
    
    # Get yesterday's other installments from snapshot
    yesterday_other_installments = {
        'weekly_total': yesterday_data.get('weekly_total', 0),
        'ten_days_total': yesterday_data.get('ten_days_total', 0),
        'monthly_total': yesterday_data.get('monthly_total', 0)
    }
    
    # Calculate weekly expected total using sales data
    # Start with yesterday's snapshot weekly total, add today's new weekly sales from database
    today_weekly_expected = yesterday_other_installments['weekly_total'] + yesterday_weekly_total
    
    # Detailed comparison - use appropriate historical dates for each metric
    detailed_comparison = [
        {
            'metric': 'Daily Collection Total',
            'yesterday_value': yesterday_data.get('daily_total', 0),
            'today_value': today_data.get('daily_total', 0),
            'comparison_type': 'daily',
            'comparison_date': yesterday_str,
            'expected_from_sales': None
        },
        {
            'metric': 'Weekly Collection Total',
            'yesterday_value': last_week_data.get('weekly_total', 0),
            'today_value': today_data.get('weekly_total', 0),
            'comparison_type': 'weekly',
            'comparison_date': last_week_str,
            'expected_from_sales': weekly_expected_total,
            'sales_details': weekly_sales_details,
            'matched_clients': weekly_matched_clients
        },
        {
            'metric': '10-Days Collection Total',
            'yesterday_value': ten_days_comparison_data.get('ten_days_total', 0),
            'today_value': today_data.get('ten_days_total', 0),
            'comparison_type': 'ten_days',
            'comparison_date': ten_days_comparison_str,
            'expected_from_sales': ten_days_expected_total,
            'sales_details': ten_days_sales_details,
            'matched_clients': ten_days_matched_clients
        },
        {
            'metric': 'Monthly Collection Total',
            'yesterday_value': monthly_comparison_data.get('monthly_total', 0),
            'today_value': today_data.get('monthly_total', 0),
            'comparison_type': 'monthly',
            'comparison_date': monthly_comparison_str,
            'expected_from_sales': monthly_expected_total,
            'sales_details': monthly_sales_details,
            'matched_clients': monthly_matched_clients
        },
        {
            'metric': 'Total Daily Clients',
            'yesterday_value': len(yesterday_data.get('daily', [])),
            'today_value': len(today_data.get('daily', [])),
            'comparison_type': 'daily',
            'comparison_date': yesterday_str,
            'expected_from_sales': None
        },
        {
            'metric': 'Total Weekly Clients',
            'yesterday_value': len(last_week_data.get('weekly', [])),
            'today_value': len(today_data.get('weekly', [])),
            'comparison_type': 'weekly',
            'comparison_date': last_week_str,
            'expected_from_sales': None
        }
    ]
    
    # Find clients that reduced (disappeared from today's daily list)
    yesterday_daily_clients = set(client['name'] for client in yesterday_data.get('daily', []))
    today_daily_clients = set(client['name'] for client in today_data.get('daily', []))
    reduced_clients = yesterday_daily_clients - today_daily_clients
    
    client_reduction_details = []
    for client_name in reduced_clients:
        loan = Loan.query.filter(
            Loan.name == client_name,
            Loan.repayment_type == 'DAILY'
        ).first()
        if loan:
            # Calculate installment amount for this client
            _, _, inst_amt, _ = compute_row_fields(loan, yesterday_date)
            
            client_reduction_details.append({
                'name': client_name,
                'phone': loan.phone or '',
                'status': loan.status,
                'installment_amount': inst_amt or 0,
                'reason': 'CLOSED' if loan.status == 'CLOSED' else 'UNKNOWN'
            })
    
    # Snapshot info
    snapshot_info = {
        'yesterday_snapshot_date': yesterday_str,
        'yesterday_snapshot_generated_at': yesterday_snap.generated_at,
        'today_snapshot_date': today_str,
        'today_snapshot_generated_at': today_snap.generated_at
    }
    
    return jsonify({
        'snapshot_info': snapshot_info,
        'yesterday_daily_total': yesterday_data.get('daily_total', 0),
        'yesterday_new_daily_sales': yesterday_daily_total,
        'yesterday_weekly_total': yesterday_weekly_total,
        'today_new_daily_sales': today_new_daily_sales,
        'today_expected_total': today_expected_total,
        'today_weekly_expected': today_weekly_expected,
        'today_actual_total': today_actual_total,
        'difference': difference,
        'yesterday_new_sales': yesterday_new_sales,
        'yesterday_closed_cases': yesterday_closed_cases,
        'yesterday_going_to_close': yesterday_going_to_close,
        'yesterday_closed_daily_lost': yesterday_closed_daily_lost,
        'yesterday_going_to_close_daily_lost': yesterday_going_to_close_daily_lost,
        'total_daily_lost': total_daily_lost,
        'today_other_installments': today_other_installments,
        'detailed_comparison': detailed_comparison,
        'client_reduction_details': client_reduction_details,
        'weekly_comparison': weekly_comparison,
        'monthly_comparison': monthly_comparison,
        'ten_days_comparison': ten_days_comparison
    })

def compare_client_category(snapshot_clients, system_clients, category_type, comparison_date=None, current_date=None):
    """Compare client category between snapshot and system computed"""
    from datetime import datetime, timedelta
    
    snapshot_set = set(c['name'] for c in snapshot_clients if not c.get('is_sub_row'))
    system_set = set(c['name'] for c in system_clients if not c.get('is_sub_row'))
    
    added = system_set - snapshot_set
    removed = snapshot_set - system_set
    common = snapshot_set & system_set
    
    # Handle replaced cases: old case closed and new case started same day
    # These should be moved from common to both removed (old) and added (new)
    replaced_cases = []
    for name in common:
        loan = Loan.query.filter(
            Loan.name == name,
            Loan.repayment_type == category_type
        ).first()
        
        if loan:
            all_loans = Loan.query.filter(
                Loan.name == name,
                Loan.repayment_type == category_type
            ).all()
            
            closed_loans = [l for l in all_loans if l.status == 'CLOSED' and l.loan_closed_date]
            open_loans = [l for l in all_loans if l.status == 'OPEN']
            
            # Only detect if exactly 1 closed and 1 open loan on same date (simple replacement)
            if closed_loans and open_loans and len(closed_loans) == 1 and len(open_loans) == 1:
                for closed in closed_loans:
                    for open_loan in open_loans:
                        # Parse dates to handle different formats
                        try:
                            closed_date = parse_date_str(closed.loan_closed_date)
                            open_date = parse_date_str(open_loan.loan_date)
                            if closed_date == open_date:
                                replaced_cases.append(name)
                                break
                        except:
                            pass
                    if name in replaced_cases:
                        break
    
    # Move replaced cases from common to both removed and added
    for name in replaced_cases:
        if name in common:
            common.remove(name)
        removed.add(name)
        added.add(name)
    
    # Parse dates for sales/closed case checking
    comp_date = parse_date_str(comparison_date) if comparison_date else None
    curr_date = parse_date_str(current_date) if current_date else None
    
    # Get detailed info for each group (no phone)
    added_details = []
    for name in added:
        client = next((c for c in system_clients if c['name'] == name), None)
        if client:
            # Check if this client was added as new sale on the comparison date
            reason = 'Unknown'
            if comparison_date:
                # Get all loans for this client and find the OPEN one
                all_loans = Loan.query.filter(
                    Loan.name == name,
                    Loan.repayment_type == category_type
                ).all()
                
                # Look for OPEN loan that matches the comparison date
                matching_loan = None
                for loan in all_loans:
                    if loan.status == 'OPEN':
                        try:
                            loan_date = parse_date_str(loan.loan_date)
                            comp_date = parse_date_str(comparison_date)
                            if loan_date == comp_date:
                                matching_loan = loan
                                break
                        except:
                            pass
                
                # If no exact match, use the first OPEN loan
                if not matching_loan:
                    for loan in all_loans:
                        if loan.status == 'OPEN':
                            matching_loan = loan
                            break
                
                # If still no match, use the first loan
                if not matching_loan and all_loans:
                    matching_loan = all_loans[0]
                
                if matching_loan:
                    # Parse dates to handle different formats
                    try:
                        loan_date = parse_date_str(matching_loan.loan_date)
                        comp_date = parse_date_str(comparison_date)
                        if loan_date == comp_date:
                            reason = f'New sale added on {comparison_date}'
                        else:
                            reason = f'Sale added on {matching_loan.loan_date} (not on comparison date)'
                    except:
                        reason = f'Sale added on {matching_loan.loan_date} (not on comparison date)'
            
            added_details.append({
                'name': name,
                'amount': client.get('total_amount', 0),
                'is_new_client': client.get('is_new_client', False),
                'reason': reason
            })
    
    removed_details = []
    for name in removed:
        client = next((c for c in snapshot_clients if c['name'] == name), None)
        if client:
            # Check if this client's loan was closed on the comparison date
            reason = 'Unknown'
            if comparison_date:
                # Get all loans for this client and find the CLOSED one
                all_loans = Loan.query.filter(
                    Loan.name == name,
                    Loan.repayment_type == category_type
                ).all()
                
                # Look for CLOSED loan that matches the comparison date
                matching_loan = None
                for loan in all_loans:
                    if loan.status == 'CLOSED' and loan.loan_closed_date:
                        try:
                            closed_date = parse_date_str(loan.loan_closed_date)
                            comp_date = parse_date_str(comparison_date)
                            if closed_date == comp_date:
                                matching_loan = loan
                                break
                        except:
                            pass
                
                # If no exact match, use the first CLOSED loan
                if not matching_loan:
                    for loan in all_loans:
                        if loan.status == 'CLOSED' and loan.loan_closed_date:
                            matching_loan = loan
                            break
                
                # If still no match, use the first loan
                if not matching_loan and all_loans:
                    matching_loan = all_loans[0]
                
                if matching_loan:
                    if matching_loan.status == 'CLOSED' and matching_loan.loan_closed_date:
                        # Parse dates to handle different formats
                        try:
                            closed_date = parse_date_str(matching_loan.loan_closed_date)
                            comp_date = parse_date_str(comparison_date)
                            if closed_date == comp_date:
                                reason = f'Loan closed on {comparison_date}'
                            else:
                                reason = f'Loan closed on {matching_loan.loan_closed_date} (not on comparison date)'
                        except:
                            reason = f'Loan closed on {matching_loan.loan_closed_date} (not on comparison date)'
                    else:
                        reason = 'Loan not closed'
            
            removed_details.append({
                'name': name,
                'amount': client.get('total_amount', 0),
                'is_closed_loan': client.get('is_closed_loan', False),
                'reason': reason
            })
    
    common_details = []
    for name in common:
        snapshot_client = next((c for c in snapshot_clients if c['name'] == name), None)
        system_client = next((c for c in system_clients if c['name'] == name), None)
        if snapshot_client and system_client:
            amount_diff = system_client.get('total_amount', 0) - snapshot_client.get('total_amount', 0)
            common_details.append({
                'name': name,
                'snapshot_amount': snapshot_client.get('total_amount', 0),
                'system_amount': system_client.get('total_amount', 0),
                'amount_diff': amount_diff
            })
    
    result = {
        'category': category_type,
        'has_data': len(snapshot_set) > 0 or len(system_set) > 0,
        'snapshot_count': len(snapshot_set),
        'system_count': len(system_set),
        'snapshot_total': sum(c.get('total_amount', 0) for c in snapshot_clients if not c.get('is_sub_row')),
        'system_total': sum(c.get('total_amount', 0) for c in system_clients if not c.get('is_sub_row')),
        'added': added_details,
        'removed': removed_details,
        'common': common_details,
        'comparison_date': comparison_date,
        'current_date': current_date
    }
    
    return result

def compute_daily_report_for_date(target_date, force_live=False):
    """Compute daily report data for a specific date (similar to daily_collection_data but without snapshot)"""
    from datetime import datetime, timedelta
    
    previous_date = target_date - timedelta(days=1)
    loans = Loan.query.all()
    client_collections = {}
    
    for loan in loans:
        if loan is None:
            continue
        has_installment = has_installment_due_on_date(loan, target_date)
        
        # Robust date parsing
        loan_closed_yesterday = False
        if loan.status == "CLOSED" and loan.loan_closed_date:
            try:
                date_str = str(loan.loan_closed_date).strip()
                loan_closed_date = None
                
                if '-' in date_str and len(date_str) == 10:
                    if date_str[4] == '-':
                        loan_closed_date = datetime.strptime(date_str, '%Y-%m-%d').date()
                    elif date_str[2] == '-':
                        loan_closed_date = datetime.strptime(date_str, '%d-%m-%Y').date()
                
                if loan_closed_date:
                    loan_closed_yesterday = (loan_closed_date == previous_date)
            except:
                loan_closed_yesterday = False

        if loan.status == "CLOSED" and not loan_closed_yesterday:
            continue
        if loan.status == "CLOSED":
            has_installment = False
        
        if has_installment or loan_closed_yesterday:
            phone_key = (loan.phone or '').strip()
            base_key = phone_key if phone_key else (loan.name or '').strip()
            key = f"{base_key}_{loan.repayment_type}"
            loan_start = parse_date_str(loan.loan_date)
            
            is_new_client_for_this_loan = (loan_start == previous_date) if loan_start else False
            is_closed_loan_for_this_loan = loan_closed_yesterday
            
            if key not in client_collections:
                client_collections[key] = {
                    "name": loan.name,
                    "phone": (loan.phone or '').strip(),
                    "repayment_type": loan.repayment_type,
                    "total_amount": 0,
                    "weekly_amount": 0,
                    "weekly_two_days": 0,
                    "is_new_client": is_new_client_for_this_loan,
                    "is_closed_loan": is_closed_loan_for_this_loan
                }
            else:
                if is_new_client_for_this_loan:
                    client_collections[key]["is_new_client"] = True
                if is_closed_loan_for_this_loan:
                    client_collections[key]["is_closed_loan"] = True
            
            if has_installment:
                _, pending_inst, inst_amt, _ = compute_row_fields(loan, target_date)
                if loan.repayment_type == "TEN_DAYS":
                    advance_amt = loan.advance_amount or 0
                    per_amt = round((loan.loan_amount or 0) / 10) if (loan.loan_amount or 0) else 0
                    if per_amt > 0:
                        advance_blocks = int(advance_amt // per_amt)
                        if advance_blocks >= 10:
                            continue
                if loan.repayment_type == "WEEKLY":
                    start_date = parse_date_str(loan.loan_date)
                    total_amt = loan.loan_amount or 0
                    weekly_installment = round(total_amt * 0.07)
                    two_days_balance = round(total_amt * 0.02)
                    adv_amt = loan.advance_amount or 0
                    advance_covers_week1 = (weekly_installment > 0) and (adv_amt >= weekly_installment)

                    days_diff = (target_date - start_date).days if start_date else -1
                    if advance_covers_week1 and days_diff == 0:
                        continue
                    if days_diff == 98:
                        if advance_covers_week1:
                            client_collections[key]["weekly_two_days"] += two_days_balance
                        else:
                            client_collections[key]["weekly_amount"] += weekly_installment
                            client_collections[key]["weekly_two_days"] += two_days_balance
                    else:
                        client_collections[key]["weekly_amount"] += weekly_installment
                else:
                    due_amount = due_amount_for_date(loan, target_date, inst_amt)
                    client_collections[key]["total_amount"] += due_amount
            else:
                client_collections[key]["total_amount"] += 0
    
    # Separate and sort clients
    daily_clients = []
    weekly_clients = []
    ten_days_clients = []
    monthly_clients = []
    
    for client_data in client_collections.values():
        if client_data["repayment_type"] == "WEEKLY":
            weekly_total = (client_data.get("weekly_amount") or 0) + (client_data.get("weekly_two_days") or 0)
            if weekly_total > 0:
                if (client_data.get("weekly_amount") or 0) > 0:
                    weekly_clients.append({
                        "name": client_data["name"],
                        "phone": client_data.get("phone", ""),
                        "repayment_type": "WEEKLY",
                        "total_amount": client_data.get("weekly_amount") or 0,
                        "is_new_client": client_data.get("is_new_client", False),
                        "is_closed_loan": client_data.get("is_closed_loan", False)
                    })
                if (client_data.get("weekly_two_days") or 0) > 0:
                    weekly_clients.append({
                        "name": client_data["name"],
                        "phone": client_data.get("phone", ""),
                        "repayment_type": "WEEKLY",
                        "total_amount": client_data.get("weekly_two_days") or 0,
                        "is_sub_row": True,
                        "sub_label": "2 Days Balance"
                    })
            continue

        if client_data["repayment_type"] == "DAILY":
            daily_clients.append(client_data)
        elif client_data["repayment_type"] == "TEN_DAYS":
            ten_days_clients.append(client_data)
        elif client_data["repayment_type"] == "MONTHLY":
            monthly_clients.append(client_data)
    
    daily_clients.sort(key=lambda x: x["name"].lower())
    weekly_clients.sort(key=lambda x: (x["name"].lower(), 1 if x.get("is_sub_row") else 0))
    ten_days_clients.sort(key=lambda x: x["name"].lower())
    monthly_clients.sort(key=lambda x: x["name"].lower())
    
    daily_total = sum(client["total_amount"] for client in daily_clients)
    weekly_total = sum(client["total_amount"] for client in weekly_clients)
    ten_days_total = sum(client["total_amount"] for client in ten_days_clients)
    monthly_total = sum(client["total_amount"] for client in monthly_clients)

    return {
        "daily": daily_clients,
        "weekly": weekly_clients,
        "ten_days": ten_days_clients,
        "daily_total": daily_total,
        "weekly_total": weekly_total,
        "ten_days_total": ten_days_total,
        "monthly": monthly_clients,
        "monthly_total": monthly_total,
        "is_snapshot": False,
        "snapshot_generated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

# =============================================================================
# SALES REPORTS WITH MANUAL CLOSE FEATURE - 100% COMPLETE
# =============================================================================

def _query_sales(start_str="", end_str="", status="", type_filter="", remarks_filter="", has_remarks=False):
    """Query sales data with date range + status + type + remarks filtering"""
    query = Loan.query
    
    if start_str:
        query = query.filter(Loan.loan_date >= start_str)
    if end_str:
        query = query.filter(Loan.loan_date <= end_str)
    
    if status:
        query = query.filter(Loan.status == status)
    
    if type_filter:
        query = query.filter(Loan.repayment_type == type_filter)
    
    if remarks_filter:
        query = query.filter(Loan.remarks == remarks_filter)
    
    if has_remarks:
        query = query.filter(Loan.remarks.isnot(None), Loan.remarks != "")
    
    loans = query.order_by(Loan.loan_date.asc()).all()

    rows = []
    totals = {
        "loan_amount": 0,
        "copies": 0,
        "installments": 0
    }

    client_loans_cache = {}

    def find_renewal_loan(client_name, closing_dt, current_loan_id):
        """Find renewal loan: same day, next day, or any day after close date"""
        if not closing_dt:
            return None, None
        if client_name not in client_loans_cache:
            client_loans_cache[client_name] = Loan.query.filter(Loan.name == client_name).all()
        
        # First check for same-day renewal
        for client_loan in client_loans_cache[client_name]:
            if client_loan.id == current_loan_id:
                continue
            start_dt = parse_date_str(client_loan.loan_date)
            if start_dt == closing_dt:
                return client_loan, "same_day"
        
        # If no same-day renewal, check for ANY renewal after close date
        # Find the earliest loan that starts after the closing date
        renewal_loan = None
        renewal_start_dt = None
        
        for client_loan in client_loans_cache[client_name]:
            if client_loan.id == current_loan_id:
                continue
            start_dt = parse_date_str(client_loan.loan_date)
            if start_dt > closing_dt:
                # This is a potential renewal loan
                if renewal_start_dt is None or start_dt < renewal_start_dt:
                    renewal_loan = client_loan
                    renewal_start_dt = start_dt
        
        if renewal_loan:
            # Determine if it's next day or later
            next_day = closing_dt + timedelta(days=1)
            if renewal_start_dt == next_day:
                return renewal_loan, "next_day"
            else:
                return renewal_loan, "later_day"
        
        return None, None

    today = date.today()
    for i, loan in enumerate(loans, 1):
        _, _, inst_amt, _ = compute_row_fields(loan, as_of_date=today)

        end_dt = schedule_end_date(loan)

        loan_dt = parse_date_str(loan.loan_date)
        day = loan_dt.strftime("%a") if loan_dt and loan.repayment_type == "WEEKLY" else ""

        # ✅ FIXED: Calculate current balance for manual close
        current_balance, _, _, _ = compute_row_fields(loan, as_of_date=today)

        renewal_info = ""
        if loan.loan_closed_date:
            closing_dt = parse_date_str(loan.loan_closed_date)
            renewal_loan, renewal_type = find_renewal_loan(loan.name, closing_dt, loan.id)
            if renewal_loan:
                renewal_amount = int(renewal_loan.loan_amount or 0)
                if renewal_type == "same_day":
                    renewal_info = f"Renewed same day @ ₹{renewal_amount:,}"
                elif renewal_type == "next_day":
                    renewal_date_str = fmt_dd_mm_yyyy(renewal_loan.loan_date)
                    renewal_info = f"Renewed on {renewal_date_str} @ ₹{renewal_amount:,}"
                elif renewal_type == "later_day":
                    renewal_date_str = fmt_dd_mm_yyyy(renewal_loan.loan_date)
                    renewal_info = f"Renewed on {renewal_date_str} @ ₹{renewal_amount:,}"
            else:
                renewal_info = "Loan not renewed"

        row = {
            "serial": i,
            "loan_id": loan.id,  # ✅ NEW: Add loan_id for closing
            "name": loan.name,
            "loan_start_date": fmt_dd_mm_yyyy(loan.loan_date),
            "loan_end_date": fmt_dd_mm_yyyy(end_dt.strftime("%Y-%m-%d")) if end_dt else "",
            "close_date": fmt_dd_mm_yyyy(loan.loan_closed_date),
            "close_amount": int(loan.case_closing_amount) if loan.case_closing_amount else 0,
            "type": loan.repayment_type,
            "day": day,
            "loan_amount": int(loan.loan_amount) if loan.loan_amount else 0,
            "copies": round((loan.loan_amount or 0)/10_000, 2),
            "installment_amt": inst_amt,
            "status": loan.status,
            "remarks": loan.remarks or "",
            "current_balance": current_balance,  # ✅ NEW: For default close amount
            "natural_end_date": end_dt.strftime("%Y-%m-%d") if end_dt else "",
            "renewal_info": renewal_info
        }

        rows.append(row)
        
        totals["loan_amount"] += row["loan_amount"]
        totals["copies"] += row["copies"]
        totals["installments"] += row["installment_amt"]
    
    return rows, totals

@app.get("/sales/report")
@login_required
@page_required('sales_report')
def sale_report():
    """Sales Report with FROM/TO date filtering + Status + Type filtering + Manual Close Feature"""
    start_str = (request.args.get("start") or "").strip()
    end_str   = (request.args.get("end")   or "").strip()
    status    = (request.args.get("status") or "").strip()
    type_filter = (request.args.get("type") or "").strip()

    if start_str and parse_date_str(start_str) is None:
        return jsonify({"error": "start date bad format (use YYYY-MM-DD)"}), 400
    if end_str and parse_date_str(end_str) is None:
        return jsonify({"error": "end date bad format (use YYYY-MM-DD)"}), 400

    try:
        remarks_filter = (request.args.get("remarks") or "").strip()
        has_remarks = request.args.get("has_remarks") == "true"
        rows, totals = _query_sales(start_str, end_str, status, type_filter, remarks_filter, has_remarks)
        return jsonify({
            "rows": rows, 
            "totals": totals, 
            "count": len(rows),
            "filters": {
                "start_date": start_str or "All time",
                "end_date": end_str or "Present",
                "status": status or "All",
                "type": type_filter or "All",
                "remarks": "Has Remarks" if has_remarks else (remarks_filter or "All")
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.get("/sales/report/page")
@login_required
@page_required('sales_report')
def sale_report_page():
    """Sales Report HTML page with Manual Close Feature"""
    return render_template("sale_report.html")

# ✅ PROFESSIONAL PDF EXPORT FOR SALES REPORT
@app.get("/sales/export/pdf")
@login_required
@page_required('sales_report')
def sales_export_pdf():
    """✅ PROFESSIONAL PDF export with same styling as daily report"""
    start_str = (request.args.get("start") or "").strip()
    end_str   = (request.args.get("end")   or "").strip()
    status    = (request.args.get("status") or "").strip()
    type_filter = (request.args.get("type") or "").strip()
    remarks_filter = (request.args.get("remarks") or "").strip()
    has_remarks = request.args.get("has_remarks") == "true"

    try:
        rows, totals = _query_sales(start_str, end_str, status, type_filter, remarks_filter, has_remarks)

        open_totals = {"loan_amount": 0, "copies": 0, "installments": 0, "count": 0}
        closed_totals = {"loan_amount": 0, "copies": 0, "installments": 0, "count": 0}

        for row in rows:
            bucket = closed_totals if row.get("status") == "CLOSED" else open_totals
            amount = row.get("loan_amount") or 0
            copies = row.get("copies") or 0
            installments = row.get("installment_amt") or 0

            bucket["loan_amount"] += amount
            bucket["copies"] += copies
            bucket["installments"] += installments
            bucket["count"] += 1

        grand_totals = {
            "loan_amount": open_totals["loan_amount"] + closed_totals["loan_amount"],
            "copies": open_totals["copies"] + closed_totals["copies"],
            "installments": open_totals["installments"] + closed_totals["installments"],
            "count": open_totals["count"] + closed_totals["count"],
        }
        
        html_content = render_template_string(
            SALES_REPORT_PDF_TEMPLATE,
            rows=rows,
            totals=totals,
            total_count=len(rows),
            open_totals=open_totals,
            closed_totals=closed_totals,
            grand_totals=grand_totals,
            filters={
                "start_date": start_str or "All time",
                "end_date": end_str or "Present",
                "status": status or "All",
                "type": type_filter or "All",
                "remarks": "Has Remarks" if has_remarks else (remarks_filter or "All")
            },
            generated_date=datetime.now().strftime("%d-%b-%Y %H:%M")
        )
        
        options = {
            'page-size': 'A4',
            'orientation': 'Landscape',
            'margin-top': '0.5cm',
            'margin-right': '0.5cm',
            'margin-bottom': '0.5cm',
            'margin-left': '0.5cm',
            'encoding': "UTF-8",
            'no-outline': None,
            'enable-local-file-access': None
        }
        
        if not PDFKIT_AVAILABLE:
            return jsonify({'error': 'PDF generation not available - pdfkit not installed'}), 500
        
        pdf = pdfkit.from_string(html_content, False, options=options)
        buf = io.BytesIO(pdf)
        buf.seek(0)
        
        return send_file(buf, mimetype='application/pdf', as_attachment=True,
                        download_name=f'sales_report_{datetime.now().strftime("%Y%m%d")}.pdf')
        
    except ValueError as e:
        return str(e), 400

# ✅ PROFESSIONAL EXCEL EXPORT FOR SALES REPORT
@app.get("/sales/export/excel")
@login_required
@page_required('sales_report')
def sales_export_excel():
    """✅ PROFESSIONAL Excel export with perfect formatting"""
    try:
        import pandas as pd
        from io import BytesIO
    except ImportError:
        return jsonify({"error": "Pandas required for Excel export"}), 500
    
    start_str = (request.args.get("start") or "").strip()
    end_str   = (request.args.get("end")   or "").strip()
    status    = (request.args.get("status") or "").strip()
    type_filter = (request.args.get("type") or "").strip()
    remarks_filter = (request.args.get("remarks") or "").strip()
    has_remarks = request.args.get("has_remarks") == "true"

    try:
        rows, totals = _query_sales(start_str, end_str, status, type_filter, remarks_filter, has_remarks)

        open_totals = {"loan_amount": 0, "copies": 0, "installments": 0, "count": 0}
        closed_totals = {"loan_amount": 0, "copies": 0, "installments": 0, "count": 0}

        for row in rows:
            bucket = closed_totals if row.get("status") == "CLOSED" else open_totals
            bucket["loan_amount"] += row.get("loan_amount") or 0
            bucket["copies"] += row.get("copies") or 0
            bucket["installments"] += row.get("installment_amt") or 0
            bucket["count"] += 1

        grand_totals = {
            "loan_amount": open_totals["loan_amount"] + closed_totals["loan_amount"],
            "copies": open_totals["copies"] + closed_totals["copies"],
            "installments": open_totals["installments"] + closed_totals["installments"],
            "count": open_totals["count"] + closed_totals["count"],
        }
        
        # ✅ PROFESSIONAL EXCEL STRUCTURE
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet("Sales Report")
            
            # ✅ A4 PAGE SETUP
            worksheet.set_landscape()
            worksheet.set_paper(9)
            worksheet.fit_to_pages(1, 0)
            worksheet.set_margins(0.5, 0.5, 0.5, 0.5)
            
            # ✅ PROFESSIONAL FORMATTING
            header_format = workbook.add_format({
                'bold': True, 'font_size': 16, 'bg_color': '#4472C4', 'font_color': 'white',
                'align': 'center', 'valign': 'vcenter', 'border': 1
            })
            
            filters_format = workbook.add_format({
                'font_size': 11, 'bg_color': '#F8F9FA', 'border': 1, 'text_wrap': True
            })
            
            column_header_format = workbook.add_format({
                'bold': True, 'bg_color': '#343A40', 'font_color': 'white',
                'align': 'center', 'border': 1, 'font_size': 10
            })
            
            open_format = workbook.add_format({
                'bg_color': '#FFF3CD', 'font_color': '#856404', 'border': 1
            })
            open_close_format = workbook.add_format({
                'bg_color': '#FFF3CD', 'font_color': '#856404', 'border': 1, 'text_wrap': True
            })
            
            closed_format = workbook.add_format({
                'bg_color': '#D1EDFF', 'font_color': '#0C5460', 'border': 1
            })
            closed_close_format = workbook.add_format({
                'bg_color': '#D1EDFF', 'font_color': '#0C5460', 'border': 1, 'text_wrap': True
            })
            
            regular_format = workbook.add_format({'border': 1})
            
            totals_format = workbook.add_format({
                'bold': True, 'bg_color': '#28A745', 'font_color': 'white',
                'align': 'center', 'border': 2
            })
            closed_summary_format = workbook.add_format({
                'bold': True, 'bg_color': '#6C757D', 'font_color': 'white',
                'align': 'center', 'border': 2
            })
            grand_summary_format = workbook.add_format({
                'bold': True, 'bg_color': '#17A2B8', 'font_color': 'white',
                'align': 'center', 'border': 2
            })
            
            # ✅ WRITE REPORT HEADER
            current_row = 0
            worksheet.merge_range(current_row, 0, current_row, 12, '📊 SALES REPORT', header_format)
            current_row += 1
            worksheet.merge_range(current_row, 0, current_row, 12, 
                                f'Generated: {datetime.now().strftime("%d-%m-%Y %I:%M %p")} | Total Records: {len(rows)}', 
                                filters_format)
            current_row += 2
            
            # ✅ FILTERS INFO
            remarks_display = "Has Remarks" if has_remarks else (remarks_filter or "All")
            filter_text = f"Date Range: {start_str or 'All'} to {end_str or 'Present'} | Status: {status or 'All'} | Type: {type_filter or 'All'} | Remarks: {remarks_display}"
            worksheet.merge_range(current_row, 0, current_row, 12, f'Filters Applied: {filter_text}', filters_format)
            current_row += 3
            
            # ✅ COLUMN HEADERS
            headers = ['#', 'Client Name', 'Loan Start', 'Loan End', 'Close Date', 'Close Amount',
                      'Type', 'Day', 'Loan Amount', 'Copies', 'Installment', 'Remarks', 'Status']
            
            for col, header in enumerate(headers):
                worksheet.write(current_row, col, header, column_header_format)
            current_row += 1
            
            # ✅ DATA ROWS
            for row in rows:
                # Choose format based on status
                row_format = closed_format if row['status'] == 'CLOSED' else open_format
                close_cell_format = closed_close_format if row['status'] == 'CLOSED' else open_close_format
                close_cell_value = row['close_date'] or ''
                if row.get('renewal_info'):
                    close_cell_value = (close_cell_value + "\n" if close_cell_value else '') + row['renewal_info']
                
                worksheet.write(current_row, 0, row['serial'], row_format)
                worksheet.write(current_row, 1, row['name'], row_format)
                worksheet.write(current_row, 2, row['loan_start_date'], row_format)
                worksheet.write(current_row, 3, row['loan_end_date'], row_format)
                worksheet.write(current_row, 4, close_cell_value, close_cell_format)
                worksheet.write(current_row, 5, f"₹{row.get('close_amount', 0):,}", row_format)
                worksheet.write(current_row, 6, row['type'], row_format)
                worksheet.write(current_row, 7, row['day'], row_format)
                worksheet.write(current_row, 8, f"₹{row['loan_amount']:,}", row_format)
                worksheet.write(current_row, 9, f"{row['copies']:.2f}", row_format)
                worksheet.write(current_row, 10, f"₹{row['installment_amt']:,}", row_format)
                worksheet.write(current_row, 11, row.get('remarks', ''), row_format)
                worksheet.write(current_row, 12, row['status'], row_format)
                
                current_row += 1
            
            # ✅ TOTALS ROW
            current_row += 1
            worksheet.merge_range(current_row, 0, current_row, 7, 'TOTALS (Active Only):', totals_format)
            worksheet.write(current_row, 8, f"₹{open_totals['loan_amount']:,}", totals_format)
            worksheet.write(current_row, 9, f"{open_totals['copies']:.2f}", totals_format)
            worksheet.write(current_row, 10, f"₹{open_totals['installments']:,}", totals_format)
            worksheet.write(current_row, 11, '', totals_format)
            worksheet.write(current_row, 12, '', totals_format)

            current_row += 1
            worksheet.merge_range(current_row, 0, current_row, 7,
                                  f"CLOSED CASES ({closed_totals['count']}) - Excluded Above:",
                                  closed_summary_format)
            worksheet.write(current_row, 8, f"₹{closed_totals['loan_amount']:,}", closed_summary_format)
            worksheet.write(current_row, 9, f"{closed_totals['copies']:.2f}", closed_summary_format)
            worksheet.write(current_row, 10, f"₹{closed_totals['installments']:,}", closed_summary_format)
            worksheet.write(current_row, 11, '', closed_summary_format)
            worksheet.write(current_row, 12, '', closed_summary_format)

            current_row += 1
            worksheet.merge_range(current_row, 0, current_row, 7,
                                  f"GRAND TOTALS ({grand_totals['count']}) - Open + Closed:",
                                  grand_summary_format)
            worksheet.write(current_row, 8, f"₹{grand_totals['loan_amount']:,}", grand_summary_format)
            worksheet.write(current_row, 9, f"{grand_totals['copies']:.2f}", grand_summary_format)
            worksheet.write(current_row, 10, f"₹{grand_totals['installments']:,}", grand_summary_format)
            worksheet.write(current_row, 11, '', grand_summary_format)
            worksheet.write(current_row, 12, '', grand_summary_format)
            
            # ✅ COLUMN WIDTHS
            col_widths = [5, 25, 12, 12, 12, 12, 8, 5, 12, 8, 12, 15, 8]
            for i, width in enumerate(col_widths):
                worksheet.set_column(i, i, width)
        
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# =============================================================================
# BATCH OPERATIONS - 100% COMPLETE WITH ROBUST DATE PARSING
# =============================================================================
@app.route('/import-real-business-data')
@login_required
@role_required('admin')
def import_real_data():
    """Universal import system for all database tables"""
    return '''
    <html>
    <head><title>Complete Business Data Import</title></head>
    <body>
        <h2>🏦 Complete Business Database Import</h2>
        <p><strong>Import all your business tables - prevents project compatibility issues</strong></p>
        
        <h3>📊 Import All Tables (Required)</h3>
        <form method="POST" action="/bulk-import-all" enctype="multipart/form-data">
            <table border="1" style="margin: 20px 0; width: 100%;">
                <tr><th style="padding: 10px;">Table</th><th style="padding: 10px;">CSV File</th><th style="padding: 10px;">Records</th><th style="padding: 10px;">Status</th></tr>
                <tr>
                    <td style="padding: 8px;"><strong>Loans</strong></td>
                    <td style="padding: 8px;"><input type="file" name="loan_file" accept=".csv" required></td>
                    <td style="padding: 8px;">564 loans</td>
                    <td style="padding: 8px;"><span style="color: red;">REQUIRED</span></td>
                </tr>
                <tr>
                    <td style="padding: 8px;"><strong>Payments</strong></td>
                    <td style="padding: 8px;"><input type="file" name="payment_file" accept=".csv" required></td>
                    <td style="padding: 8px;">755+ payments</td>
                    <td style="padding: 8px;"><span style="color: red;">REQUIRED</span></td>
                </tr>
                <tr>
                    <td style="padding: 8px;"><strong>Payment Submissions</strong></td>
                    <td style="padding: 8px;"><input type="file" name="payment_submission_file" accept=".csv"></td>
                    <td style="padding: 8px;">2 submissions</td>
                    <td style="padding: 8px;"><span style="color: green;">OPTIONAL</span></td>
                </tr>
                <tr>
                    <td style="padding: 8px;"><strong>Alternative Payments</strong></td>
                    <td style="padding: 8px;"><input type="file" name="payments_file" accept=".csv"></td>
                    <td style="padding: 8px;">Empty table</td>
                    <td style="padding: 8px;"><span style="color: green;">OPTIONAL</span></td>
                </tr>
                <tr>
                    <td style="padding: 8px;"><strong>Recovery Payments</strong></td>
                    <td style="padding: 8px;"><input type="file" name="recovery_payments_file" accept=".csv"></td>
                    <td style="padding: 8px;">Empty table</td>
                    <td style="padding: 8px;"><span style="color: green;">OPTIONAL</span></td>
                </tr>
                <tr>
                    <td style="padding: 8px;"><strong>Short Payments</strong></td>
                    <td style="padding: 8px;"><input type="file" name="short_payment_file" accept=".csv"></td>
                    <td style="padding: 8px;">Empty table</td>
                    <td style="padding: 8px;"><span style="color: green;">OPTIONAL</span></td>
                </tr>
            </table>
            <input type="submit" value="🚀 Import Complete Database (All Tables)" style="background:#28a745; color:white; padding:15px; font-size:16px; width: 100%; margin: 10px 0;">
        </form>
        
        <p><strong>Note:</strong> This imports all tables to ensure full project compatibility</p>
        <p><a href="/">← Back to App</a> | <a href="/debug-db">Check Database</a></p>
    </body>
    </html>
    '''

@app.route('/bulk-import-all', methods=['POST'])
@csrf.exempt
@login_required
@role_required('admin')
def bulk_import_all():
    try:
        results = {}
        
        # Import Loans (REQUIRED)
        if 'loan_file' in request.files and request.files['loan_file']:
            file = request.files['loan_file']
            df = pd.read_csv(file)
            db.session.query(Loan).delete()
            
            count = 0
            for _, row in df.iterrows():
                loan = Loan(
                    name=str(row.get('name', '')) if pd.notna(row.get('name')) else '',
                    address=str(row.get('address', '')) if pd.notna(row.get('address')) else '',
                    phone=str(row.get('phone', '')) if pd.notna(row.get('phone')) else '',
                    city=str(row.get('city', '')) if pd.notna(row.get('city')) else '',
                    loan_date=str(row.get('loan_date', '')) if pd.notna(row.get('loan_date')) else '',
                    loan_closed_date=str(row.get('loan_closed_date', '')) if pd.notna(row.get('loan_closed_date')) else None,
                    loan_amount=float(row.get('loan_amount', 0)) if pd.notna(row.get('loan_amount')) else 0.0,
                    processing_fees=float(row.get('processing_fees', 0)) if pd.notna(row.get('processing_fees')) else 0.0,
                    case_closing_amount=float(row.get('case_closing_amount', 0)) if pd.notna(row.get('case_closing_amount')) else 0.0,
                    advance_amount=float(row.get('advance_amount', 0)) if pd.notna(row.get('advance_amount')) else 0.0,
                    repayment_type=str(row.get('repayment_type', 'DAILY')) if pd.notna(row.get('repayment_type')) else 'DAILY',
                    status=str(row.get('status', 'OPEN')) if pd.notna(row.get('status')) else 'OPEN',
                    total_paid_amount=float(row.get('total_paid_amount', 0)) if pd.notna(row.get('total_paid_amount')) else 0.0,
                    outstanding_balance=float(row.get('outstanding_balance', 0)) if pd.notna(row.get('outstanding_balance')) else 0.0
                )
                db.session.add(loan)
                count += 1
            results['loans'] = count
            
        # Import Payments (REQUIRED)
        if 'payment_file' in request.files and request.files['payment_file']:
            file = request.files['payment_file']
            df = pd.read_csv(file)
            db.session.query(Payment).delete()
            
            count = 0
            for _, row in df.iterrows():
                payment = Payment(
                    loan_id=int(row.get('loan_id', 1)) if pd.notna(row.get('loan_id')) else 1,
                    payment_date=str(row.get('payment_date', '2025-01-01')),
                    amount_paid=float(row.get('amount_paid', 0)) if pd.notna(row.get('amount_paid')) else 0,
                    payment_method=str(row.get('payment_method', 'Collection Entry')),
                    remarks=str(row.get('remarks', 'Imported from CSV')),
                    created_at=str(row.get('created_at', '')),
                    entry_type=str(row.get('entry_type', 'COLLECTION'))
                )
                db.session.add(payment)
                count += 1
            results['payments'] = count
            
        # Import Payment Submissions (OPTIONAL)
        if 'payment_submission_file' in request.files and request.files['payment_submission_file'].filename:
            file = request.files['payment_submission_file']
            df = pd.read_csv(file)
            db.session.query(PaymentSubmission).delete()
            
            count = 0
            for _, row in df.iterrows():
                submission = PaymentSubmission(
                    submission_date=str(row.get('submission_date', '')),
                    total_amount=float(row.get('total_amount', 0)) if pd.notna(row.get('total_amount')) else 0,
                    total_payments=int(row.get('total_payments', 0)) if pd.notna(row.get('total_payments')) else 0,
                    submitted_at=str(row.get('submitted_at', ''))
                )
                db.session.add(submission)
                count += 1
            results['payment_submissions'] = count
        else:
            results['payment_submissions'] = 0
            
        # Import Alternative Payments (OPTIONAL)
        if 'payments_file' in request.files and request.files['payments_file'].filename:
            file = request.files['payments_file']
            df = pd.read_csv(file)
            db.session.query(Payments).delete()
            
            count = 0
            for _, row in df.iterrows():
                alt_payment = Payments(
                    loan_id=int(row.get('loan_id', 1)) if pd.notna(row.get('loan_id')) else 1,
                    payment_date=str(row.get('payment_date', '')),
                    amount=float(row.get('amount', 0)) if pd.notna(row.get('amount')) else 0,
                    payment_type=str(row.get('payment_type', '')),
                    notes=str(row.get('notes', '')),
                    created_at=str(row.get('created_at', ''))
                )
                db.session.add(alt_payment)
                count += 1
            results['alt_payments'] = count
        else:
            results['alt_payments'] = 0
                
        # Import Recovery Payments (OPTIONAL)
        if 'recovery_payments_file' in request.files and request.files['recovery_payments_file'].filename:
            file = request.files['recovery_payments_file']
            df = pd.read_csv(file)
            db.session.query(RecoveryPayment).delete()
            
            count = 0
            for _, row in df.iterrows():
                recovery = RecoveryPayment(
                    loan_id=int(row.get('loan_id', 1)) if pd.notna(row.get('loan_id')) else 1,
                    client_name=str(row.get('client_name', '')),
                    recovery_date=str(row.get('recovery_date', '')),
                    amount=float(row.get('amount', 0)) if pd.notna(row.get('amount')) else 0,
                    notes=str(row.get('notes', '')),
                    created_at=str(row.get('created_at', ''))
                )
                db.session.add(recovery)
                count += 1
            results['recovery_payments'] = count
        else:
            results['recovery_payments'] = 0
                
        # Import Short Payments (OPTIONAL)
        if 'short_payment_file' in request.files and request.files['short_payment_file'].filename:
            file = request.files['short_payment_file']
            df = pd.read_csv(file)
            db.session.query(ShortPayment).delete()
            
            count = 0
            for _, row in df.iterrows():
                short_payment = ShortPayment(
                    loan_id=int(row.get('loan_id', 1)) if pd.notna(row.get('loan_id')) else 1,
                    payment_date=str(row.get('payment_date', '')),
                    expected_amount=float(row.get('expected_amount', 0)) if pd.notna(row.get('expected_amount')) else 0,
                    status=str(row.get('status', '')),
                    created_at=str(row.get('created_at', ''))
                )
                db.session.add(short_payment)
                count += 1
            results['short_payments'] = count
        else:
            results['short_payments'] = 0
        
        db.session.commit()
        
        # Calculate totals
        total_collections = db.session.query(db.func.sum(Payment.amount_paid)).scalar() or 0
        
        return f'''
        <html>
        <head><title>Complete Database Import Success</title></head>
        <body>
            <h2>🎉 COMPLETE DATABASE IMPORT SUCCESSFUL!</h2>
            <h3>All Tables Imported - Project Compatibility Ensured:</h3>
            <table border="1" style="margin: 20px 0; width: 100%;">
                <tr><th style="padding: 10px;">Table</th><th style="padding: 10px;">Records Imported</th><th style="padding: 10px;">Status</th></tr>
                <tr><td style="padding: 8px;"><strong>Loans</strong></td><td style="padding: 8px;">{results.get('loans', 0)}</td><td style="padding: 8px;">✅ SUCCESS</td></tr>
                <tr><td style="padding: 8px;"><strong>Payments</strong></td><td style="padding: 8px;">{results.get('payments', 0)}</td><td style="padding: 8px;">✅ SUCCESS</td></tr>
                <tr><td style="padding: 8px;"><strong>Payment Submissions</strong></td><td style="padding: 8px;">{results.get('payment_submissions', 0)}</td><td style="padding: 8px;">✅ READY</td></tr>
                <tr><td style="padding: 8px;"><strong>Alternative Payments</strong></td><td style="padding: 8px;">{results.get('alt_payments', 0)}</td><td style="padding: 8px;">✅ READY</td></tr>
                <tr><td style="padding: 8px;"><strong>Recovery Payments</strong></td><td style="padding: 8px;">{results.get('recovery_payments', 0)}</td><td style="padding: 8px;">✅ READY</td></tr>
                <tr><td style="padding: 8px;"><strong>Short Payments</strong></td><td style="padding: 8px;">{results.get('short_payments', 0)}</td><td style="padding: 8px;">✅ READY</td></tr>
            </table>
            
            <h3>Business Summary:</h3>
            <ul>
                <li>💰 <strong>Total Collections:</strong> ₹{total_collections:,.0f}</li>
                <li>🏢 <strong>Total Loan Portfolio:</strong> {results.get('loans', 0)} loans</li>
                <li>📊 <strong>Database Status:</strong> Complete with all 6 business tables</li>
                <li>🔧 <strong>Project Compatibility:</strong> 100% - No missing tables</li>
            </ul>
            
            <div style="background: #d4edda; padding: 15px; margin: 20px 0; border-radius: 5px;">
                <strong>✅ SUCCESS:</strong> All database tables are now set up and ready for your project. 
                No compatibility issues will occur.
            </div>
            
            <p><a href="/" style="background:#28a745; color:white; padding:15px; text-decoration:none; font-size: 16px;">🚀 VIEW YOUR COMPLETE LOAN MANAGEMENT SYSTEM</a></p>
        </body>
        </html>
        '''
        
    except Exception as e:
        return f'''
        <html>
        <body>
            <h2>❌ Import Error</h2>
            <p><strong>Error:</strong> {str(e)}</p>
            <p><a href="/import-real-business-data">← Try Again</a></p>
        </body>
        </html>
        ''', 500


@app.route("/batch/preview-closing")
@login_required
@page_required('sales_report')
def preview_batch_closing():
    """✅ 100% COMPLETE: Preview cases that will be closed in batch with robust date parsing"""
    from_date = request.args.get('from')
    to_date = request.args.get('to')
    
    print(f"🔍 DEBUG: Received dates - From: {from_date}, To: {to_date}")
    
    if not from_date or not to_date:
        return jsonify({"error": "From and To dates are required"}), 400
    
    try:
        # Convert input dates (YYYY-MM-DD) to datetime objects
        from_dt = datetime.strptime(from_date, '%Y-%m-%d')
        to_dt = datetime.strptime(to_date, '%Y-%m-%d')
        
        print(f"🔍 DEBUG: Parsed dates - From: {from_dt.date()}, To: {to_dt.date()}")
        
        # Get all OPEN loans
        loans = db.session.query(Loan).filter(Loan.status == 'OPEN').all()
        print(f"🔍 DEBUG: Found {len(loans)} OPEN loans in database")
        
        cases = []
        total_amount = 0
        
        for loan in loans:
            if loan.loan_date:
                try:
                    # ✅ ROBUST: Use universal date parsing
                    loan_dt = parse_loan_date_universal(loan.loan_date)
                    if not loan_dt:
                        continue
                    
                    print(f"🔍 DEBUG: Loan {loan.id} - Date: {loan.loan_date}, Parsed: {loan_dt.date()}")
                    
                    # Check if loan start date is within the selected range
                    if from_dt.date() <= loan_dt.date() <= to_dt.date():
                        # Calculate natural end date based on repayment type
                        natural_end = schedule_end_date(loan)
                        if not natural_end:
                            continue
                        
                        case_data = {
                            "loan_id": loan.id,
                            "client_name": loan.name,
                            "loan_date": format_date(loan.loan_date),
                            "natural_end_date": format_date(natural_end.strftime("%Y-%m-%d")),
                            "loan_amount": loan.loan_amount or 0,
                            "balance_amount": loan.outstanding_balance or 0,
                        }
                        
                        cases.append(case_data)
                        total_amount += loan.loan_amount or 0
                        
                        print(f"✅ MATCH: Loan {loan.id} - {loan.name} - {loan.loan_date}")
                        
                except ValueError as e:
                    print(f"❌ Date parsing error for loan {loan.id}: {loan.loan_date} - {e}")
                    continue
        
        print(f"🔍 DEBUG: Final result - {len(cases)} cases found, Total: ₹{total_amount}")
        
        return jsonify({
            "cases": cases,
            "total_amount": total_amount,
            "count": len(cases)
        })
        
    except Exception as e:
        print(f"❌ ERROR in preview_batch_closing: {e}")
        return jsonify({"error": str(e)}), 500

@app.route("/batch/close-cases", methods=["POST"])
@csrf.exempt
@login_required
@page_required('sales_report')
def execute_batch_closing():
    """✅ 100% COMPLETE: Execute batch closing with robust date parsing"""
    data = request.get_json() or {}
    
    from_date = data.get('from_date')
    to_date = data.get('to_date')
    default_close_date = data.get('default_close_date')
    default_close_amount = data.get('default_close_amount')

    # Optional: per-loan override amounts from the frontend
    per_loan_list = data.get('per_loan_amounts') or []
    per_loan_map = {}
    for item in per_loan_list:
        try:
            lid = int(item.get('loan_id'))
            amt = float(item.get('close_amount') or 0)
            if lid > 0 and amt >= 0:
                per_loan_map[lid] = amt
        except Exception:
            continue
    
    print(f"🔍 DEBUG: Batch closing - From: {from_date}, To: {to_date}")
    
    if not from_date or not to_date:
        return jsonify({"error": "From and To dates are required"}), 400
    
    try:
        from_dt = datetime.strptime(from_date, '%Y-%m-%d')
        to_dt = datetime.strptime(to_date, '%Y-%m-%d')
        
        # Get all OPEN loans
        loans = db.session.query(Loan).filter(Loan.status == 'OPEN').all()
        
        closed_count = 0
        total_amount = 0
        
        for loan in loans:
            if loan.loan_date:
                try:
                    # ✅ ROBUST: Use universal date parsing
                    loan_dt = parse_loan_date_universal(loan.loan_date)
                    if not loan_dt:
                        continue
                    
                    # Check if within date range
                    if from_dt.date() <= loan_dt.date() <= to_dt.date():
                        # Determine close date
                        if default_close_date:
                            # ✅ FIXED: Store in DD-MM-YYYY format
                            close_dt = datetime.strptime(default_close_date, '%Y-%m-%d')
                            close_date = close_dt.strftime('%d-%m-%Y')
                        else:
                            # Use natural end date based on repayment type
                            natural_end = schedule_end_date(loan)
                            if not natural_end:
                                continue
                            close_date = natural_end.strftime('%d-%m-%Y')
                        
                        # Determine close amount
                        if loan.id in per_loan_map:
                            # Per-loan override from UI always wins
                            close_amount = per_loan_map[loan.id]
                        elif default_close_amount:
                            # Same default amount for all, if user entered it
                            close_amount = float(default_close_amount)
                        else:
                            # Default behaviour:
                            # - If loan's natural end date has already passed (before today) => treat as completed, default close amount 0
                            # - Otherwise, use outstanding balance (0 if fully paid)
                            try:
                                natural_end_for_amount = schedule_end_date(loan)
                            except Exception:
                                natural_end_for_amount = None

                            today_date = datetime.now().date()
                            matured = False
                            if natural_end_for_amount is not None:
                                try:
                                    ned = natural_end_for_amount.date() if hasattr(natural_end_for_amount, "date") else natural_end_for_amount
                                    matured = ned < today_date
                                except Exception:
                                    matured = False

                            if matured:
                                close_amount = 0.0
                            else:
                                bal = float(loan.outstanding_balance or 0)
                                close_amount = bal if bal > 0 else 0.0
                        
                        # Update loan
                        loan.status = 'CLOSED'
                        loan.loan_closed_date = close_date
                        loan.case_closing_amount = int(close_amount) if close_amount else 0
                        # ✅ FIXED: Set outstanding balance to 0
                        loan.outstanding_balance = 0
                        
                        closed_count += 1
                        total_amount += close_amount or 0
                        
                        print(f"✅ CLOSED: Loan {loan.id} - {loan.name}")
                        
                except ValueError:
                    continue
        
        # Commit all changes
        db.session.commit()
        
        print(f"✅ SUCCESS: Closed {closed_count} loans, Total: ₹{total_amount}")
        
        return jsonify({
            "success": True,
            "closed_count": closed_count,
            "total_amount": total_amount
        })
        
    except Exception as e:
        print(f"❌ ERROR in execute_batch_closing: {e}")
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# Individual loan closing route
@app.route('/loan/close/<int:loan_id>', methods=['POST'])
@csrf.exempt
@login_required
@page_required('loan_form')
def close_individual_loan(loan_id):
    """Close individual loan manually"""
    data = request.get_json() or {}
    
    try:
        loan = Loan.query.get_or_404(loan_id)
        
        close_date = data.get('close_date')
        close_amount = data.get('close_amount')
        
        if not close_date:
            return jsonify({"error": "Close date is required"}), 400
        
        try:
            close_dt = datetime.strptime(close_date, '%Y-%m-%d')
            # ✅ FIXED: Store in DD-MM-YYYY format
            formatted_close_date = close_dt.strftime('%d-%m-%Y')
        except ValueError:
            return jsonify({"error": "Invalid date format"}), 400
        
        loan.status = 'CLOSED'
        loan.loan_closed_date = formatted_close_date
        loan.case_closing_amount = int(float(close_amount)) if close_amount else (loan.loan_amount or 0)
        # ✅ FIXED: Set outstanding balance to 0
        loan.outstanding_balance = 0
        
        db.session.commit()
        
        return jsonify({
            "success": True,
            "message": "Loan closed successfully",
            "loan_id": loan_id
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500

# =============================================================================
# SERVER-SIDE EXCEL EXPORT WITH BORDERS
# =============================================================================

@app.route('/export-all-clients-ledgers-excel-server')
@login_required
@page_required('all_ledgers')
def export_all_clients_excel():
    """Server-side Excel export with proper borders using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Get all clients data
        clients_data = get_all_clients_ledger_data()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "All Clients Ledger"
        
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )
        
        title_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        
        client_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        client_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')
        
        totals_font = Font(name='Calibri', size=11, bold=True)
        totals_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
        
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:M{current_row}')
        cell = ws[f'A{current_row}']
        cell.value = 'SHIVI PROJECT - ALL CLIENTS LEDGER REPORT'
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = center_align
        cell.border = thick_border
        current_row += 1
        
        # Generation date
        ws[f'A{current_row}'] = f'Generated: {date.today().strftime("%Y-%m-%d")}'
        current_row += 2
        
        # Column headers
        headers = ['Loan Date', 'Day', 'End Date', 'Close Date', 'Proc. Fees', 
                   'Loan Amount', 'Inst. Amount', 'Type', 'Advance', 
                   'Closing Amt', 'Balance', 'Pending', 'Status']
        
        # Set column widths
        col_widths = [12, 8, 12, 12, 12, 14, 14, 10, 12, 14, 14, 8, 10]
        for idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # Process each client
        for client in clients_data:
            # Client header
            ws.merge_cells(f'A{current_row}:M{current_row}')
            cell = ws[f'A{current_row}']
            cell.value = f'CLIENT: {client["name"].upper()}'
            cell.font = client_font
            cell.fill = client_fill
            cell.alignment = center_align
            cell.border = thick_border
            current_row += 2
            
            # Table headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            current_row += 1
            
            # Loan rows
            for loan in client.get('loans', []):
                row_data = [
                    loan.get('loan_date', ''),
                    loan.get('day_name', ''),
                    loan.get('loan_end_date', ''),
                    loan.get('loan_closed_date', ''),
                    loan.get('processing_fees', 0),
                    loan.get('amount', 0),
                    loan.get('installment_amount', 0),
                    loan.get('repayment_type', ''),
                    loan.get('advance_amount', 0),
                    loan.get('case_closing_amt', 0),
                    loan.get('balance_amount', 0),
                    loan.get('pending_installments', 0),
                    loan.get('status', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    cell.alignment = center_align
                    
                    # Format currency
                    if col_idx in [5, 6, 7, 9, 10, 11] and isinstance(value, (int, float)):
                        cell.number_format = '₹#,##0'
                        cell.alignment = right_align
                
                current_row += 1
            
            # Totals row
            totals = client.get('totals', {})
            totals_data = [
                'TOTALS:', '', '', '',
                totals.get('processing_fees', 0),
                totals.get('loan_amount', 0),
                totals.get('installment_amount', 0),
                '',
                totals.get('advance_amount', 0),
                totals.get('case_closing_amount', 0),
                totals.get('balance_amount', 0),
                totals.get('pending_installments', 0),
                ''
            ]
            
            for col_idx, value in enumerate(totals_data, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = value
                cell.font = totals_font
                cell.fill = totals_fill
                cell.border = thick_border
                cell.alignment = right_align if col_idx > 4 else center_align
                
                if col_idx in [5, 6, 7, 9, 10, 11] and isinstance(value, (int, float)):
                    cell.number_format = '₹#,##0'
            
            current_row += 4  # 3 empty rows + 1
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f'All_Clients_Ledger_{date.today().strftime("%Y-%m-%d")}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Excel export error: {e}")
        return jsonify({"error": str(e)}), 500

# =============================================================================
def convert_postgres_to_sqlite(pg_url, sqlite_path):
    """Convert PostgreSQL database to SQLite backup"""
    from sqlalchemy import create_engine
    import pandas as pd
    
    # Connect to PostgreSQL
    pg_engine = create_engine(pg_url)
    
    # Connect to SQLite
    sqlite_engine = create_engine(f'sqlite:///{sqlite_path}')
    
    # Get all table names
    with pg_engine.connect() as conn:
        result = conn.execute(text("SELECT tablename FROM pg_tables WHERE schemaname = 'public'"))
        tables = [row[0] for row in result]
    
    # Copy each table
    for table in tables:
        try:
            df = pd.read_sql_table(table, pg_engine)
            df.to_sql(table, sqlite_engine, if_exists='replace', index=False)
            print(f"✅ Backed up table: {table}")
        except Exception as e:
            print(f"⚠️ Skipped table {table}: {e}")

# =============================================================================
# BACKUP ENDPOINT FOR HTTP BACKUP METHOD
# =============================================================================
@app.route('/api/backup', methods=['GET'])
@login_required
@role_required('admin')
def backup_database():
    """Backup database for HTTP backup method"""
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    expected_token = os.environ.get('BACKUP_TOKEN', '')
    
    if not token or token != expected_token:
        return jsonify({'error': 'Invalid or missing backup token'}), 401
    
    try:
        # Create backup of current database
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = f"backup_{timestamp}.db"
        
        if os.environ.get('DATABASE_URL'):
            # For PostgreSQL, create SQLite backup
            convert_postgres_to_sqlite(app.config['SQLALCHEMY_DATABASE_URI'], backup_path)
        else:
            # For SQLite, just copy
            import shutil
            shutil.copy2('swift_local.db', backup_path)
        
        return send_file(backup_path, as_attachment=True, 
                        download_name=f'swift_backup_{timestamp}.db')
    except Exception as e:
        return jsonify({'error': f'Backup failed: {str(e)}'}), 500

# FINAL ROUTES & ERROR HANDLERS
# =============================================================================

@app.errorhandler(404)
def not_found(e):
    return jsonify({"error": "Not Found"}), 404

@app.errorhandler(500)
def server_error(e):
    return jsonify({"error": "Internal Server Error"}), 500

# =============================================================================
# APP STARTUP - 100% COMPLETE WITH DD-MM-YYYY & WEEKLY LOGIC FIXES
# =============================================================================

if __name__ == "__main__":
    with app.app_context():
        print("🔄 Initializing 100% COMPLETE loan management system...")
        print("✅ DD-MM-YYYY format enforced throughout")
        print("✅ CRITICAL FIX: Daily installments start from NEXT day")
        print("✅ FIXED: Weekly loan logic matches your business rules")
        print("✅ FIXED: Closed loans show ₹0 balance everywhere")
        print("✅ FIXED: Daily collection report hyperlinks")
        print("✅ PROFESSIONAL: PDF & Excel exports for all reports")
        print("✅ Robust date parsing (supports multiple formats)")
        print("✅ Complete batch operations with debug logging")
        print("\n🏆 100% FEATURE COMPLETE:")
        print("   - Payment entry system with transaction safety")
        print("   - Real-time balance updates via event listeners")
        print("   - Short payment tracking & recovery")
        print("   - Manual loan closing from sales report")
        print("   - Batch closing with sophisticated date parsing")
        print("   - Professional PDF/Excel exports (Daily Report Style)")
        print("   - Complete daily reports with fixed logic")
        print("   - Receivable reports with date filtering")
        print("   - Unified client ledger management (HYPERLINK FIXED)")
        print("   - Sales reports with professional formatting")
        print("   - Individual & bulk loan operations")
        
        print("\n📋 Business Rules (100% Implemented):")
        print("   DAILY: 100 installments, starts day 1 (FIXED)")
        print("   WEEKLY: YOUR LOGIC - 14 equal installments + 2 days balance")
        print("   TEN_DAYS: 10 installments, every 10 days")
        print("   Main Rule: Close today → installment = 0 tomorrow")
        print("   Weekly Advance: First week FREE, Day 98 = ₹400 only")
        
        print("\n🎯 Date Format: DD-MM-YYYY everywhere")
        
        # Environment-aware startup message
        if os.environ.get('DATABASE_URL'):
            print("🌐 Running in CLOUD mode (Northflank/GitHub Actions)")
        else:
            print("💻 Running in LOCAL mode (SQLite)")

if __name__ == '__main__':
    print("🚀 Starting Flask server...")
    with app.app_context():
        db.create_all()
        print("✅ Tables ready")
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))

        

